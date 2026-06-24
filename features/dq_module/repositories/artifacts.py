"""Excel workbook export — equivalent of STATpy's results-export utilities."""

from __future__ import annotations


def render_excel(metrics: dict) -> bytes:
    """Generate Excel workbook with all dashboard data tables."""
    import io
    import openpyxl
    from openpyxl.styles import PatternFill, Font

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    q = metrics["latest_quarter"]
    d = metrics["by_quarter"].get(q, {})

    def add_sheet(name, headers, rows):
        ws = wb.create_sheet(name[:31])
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="0F1D35")
            cell.font = Font(bold=True, color="FFFFFF")
        for row in rows:
            ws.append(list(row.values()) if isinstance(row, dict) else row)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18
        return ws

    # Summary
    ws = wb.create_sheet("Summary")
    ws.append(["Metric", "Value"])
    ws.append(["Quarter", q])
    ws.append(["Snapshot Date", d.get("snapshot_date", "")])
    ws.append(["Total Records", d.get("total_records", "")])
    ws.append(["Overall Completeness %", d.get("completeness", {}).get("overall_pct", "")])
    ws.append(["Rules Failed", d.get("business_rules", {}).get("rules_failed", "")])
    ws.append(["Critical Failures", d.get("business_rules", {}).get("critical_failures", "")])
    ws.append(["Avg PSI", d.get("drift", {}).get("avg_psi", "")])
    ws.append(["DQ Rating", d.get("governance", {}).get("dq_rating", "")])
    ws.append(["Run ID", metrics["run_id"]])
    ws.append(["Last Refresh", metrics["last_refresh"]])

    comp = d.get("completeness", {})
    add_sheet("Completeness", ["Column", "Missing %", "Missing N", "Severity", "Critical Col"],
              comp.get("by_column", []))

    rules = d.get("business_rules", {})
    add_sheet("Business Rules", ["Code", "Description", "Severity", "Domain", "Failed Records", "Failure Rate %", "Passed"],
              rules.get("by_rule", []))

    drift = d.get("drift", {})
    add_sheet("Distribution Drift", ["Column", "PSI", "Level", "Cur Mean", "Cur Median", "Cur Std Dev", "Ref Mean", "Ref Median", "Ref Std Dev"],
              drift.get("by_variable", []))

    pop = d.get("population", {})
    add_sheet("Population by Segment", ["Segment", "Prior Q", "New", "New %", "Dropped", "Dropped %", "Current", "Net Change", "Net Change %"],
              pop.get("by_segment", []))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
