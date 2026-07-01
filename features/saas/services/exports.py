"""SAAS workspace export and report builders.

This module owns the non-Dash mechanics for report HTML and Excel workbooks.
The Dash callbacks prepare filter state and call these functions, keeping the
callback module focused on interaction wiring.
"""

from __future__ import annotations

from datetime import datetime
from html import escape as html_escape
import math
import statistics

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText, Text
from openpyxl.chart.title import Title
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import CharacterProperties, Paragraph, ParagraphProperties, RegularTextRun, RichTextProperties
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..data_access import SAAS_PAGE_DATA
from ..domain.metrics import BASELINE_SCENARIO_VALUE


def build_saas_report_html(sections: list[tuple[str, object]], meta_lines: list[str]) -> str:
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    meta_items = "".join(f"<li>{html_escape(line)}</li>" for line in meta_lines)

    if sections:
        chart_blocks = []
        for index, (title, fig) in enumerate(sections):
            # Fixed pixel size keeps the chart from being laid out at the on-screen
            # viewport width and then clipped when the browser paginates for print.
            fig = fig.update_layout(autosize=False, width=900, height=420)
            chart_html = fig.to_html(
                full_html=False,
                include_plotlyjs="cdn" if index == 0 else False,
                config={"responsive": False},
            )
            chart_blocks.append(
                f'<section class="saas-report-chart"><h2>{html_escape(title)}</h2>{chart_html}</section>'
            )
        body = "\n".join(chart_blocks)
    else:
        body = "<p>No charts match the current filters.</p>"

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>SAAS MEV Report</title>
<style>
  body {{ font-family: -apple-system, Helvetica, Arial, sans-serif; margin: 24px; color: #1f2933; }}
  h1 {{ font-size: 20px; margin-bottom: 4px; }}
  .saas-report-meta {{ font-size: 13px; color: #52606d; margin-bottom: 24px; }}
  .saas-report-meta ul {{ margin: 4px 0 0; padding-left: 18px; }}
  .saas-report-chart {{ margin-bottom: 32px; page-break-inside: avoid; }}
  .saas-report-chart h2 {{ font-size: 15px; margin-bottom: 8px; }}
  .saas-report-chart .plotly-graph-div {{ margin: 0; }}
  @media print {{
    @page {{ size: landscape; margin: 12mm; }}
    .saas-report-chart {{ page-break-after: always; }}
  }}
</style>
</head>
<body>
  <h1>Scenario Analysis as a Service (SAAS) &mdash; MEV Report</h1>
  <div class="saas-report-meta">
    Generated {generated_at}
    <ul>{meta_items}</ul>
  </div>
  {body}
  <p class="saas-report-print-hint">To save this report as a PDF, open this file in a browser and use Print &rarr; Save as PDF.</p>
</body>
</html>"""


def _model_descriptive_label(model_name: str) -> str:
    """Model's descriptive name, falling back to the raw model name."""
    descriptive_map = SAAS_PAGE_DATA.get("model_descriptive_name_map", {})
    return descriptive_map.get(model_name) or model_name

SAAS_METRIC_COLUMNS: list[tuple[str, str, str]] = [
    # --- Identifiers ---
    ("model", "Model", "Model name (as used in the GMIS system)."),
    ("model_descriptive", "Model Descriptive Name", "Model descriptive name (from the model_characteristic sheet)."),
    ("mev_type", "MEV Type", "Whether the MEV is a Raw or a Transformed variable."),
    ("mev", "MEV", "MEV name, shown using the MEV Label mode selected on the dashboard."),
    ("model_contribution", "Model Contribution", "The MEV's contribution to the model, shown as a percentage. Blank for raw MEVs."),
    # --- Historical statistics ---
    ("history_min", "History Min", "Minimum of the historical (Quarter <= 0) MEV values - the data behind the Min-Max lines."),
    ("history_max", "History Max", "Maximum of the historical (Quarter <= 0) MEV values - the data behind the Min-Max lines."),
    ("history_mean", "History Mean", "Mean of the historical (Quarter <= 0) MEV values."),
    ("history_std", "History STD", "Population standard deviation (ddof=0) of the historical MEV values."),
    # --- Historical timing ---
    ("time_of_min", "Date of Min", "Date on which the historical minimum value occurs."),
    ("time_of_max", "Date of Max", "Date on which the historical maximum value occurs."),
    ("minmax_daterange", "Historical Min-Max Date Range", "Date range spanned by the historical (Quarter <= 0) data: earliest historical date to latest historical date."),
    # --- Chosen-scenario projection vs history ({scenario} is replaced with the
    #     selected scenario's label, e.g. Severe / Baseline / Other) ---
    ("sevadv_min", "{scenario} Min", "Minimum of the projection (Quarter > 0) MEV values for the chosen scenario."),
    ("is_min_beyond", "Is {scenario} Min Beyond?", "Yes if the projection min is lower than the historical min."),
    ("sevadv_max", "{scenario} Max", "Maximum of the projection (Quarter > 0) MEV values for the chosen scenario."),
    ("is_max_beyond", "Is {scenario} Max Beyond?", "Yes if the projection max is greater than the historical max."),
    # --- Lower bound (History Min - 2 STD) and its breach tests ---
    ("history_min_2std", "History Min - 2 STD", "History Min - 2 x History STD."),
    ("is_min_lt_2std", "Is {scenario} Min < History Min - 2 STD?", "Yes if the projection min is less than History Min - 2 x History STD."),
    ("is_baseline_min_lt_2std", "Is Baseline Min < History Min - 2 STD?", "Yes if the projection min is less than History Min - 2 x History STD (using Baseline scenario data)."),
    # --- Upper bound (History Max + 2 STD) and its breach tests ---
    ("history_max_2std", "History Max + 2 STD", "History Max + 2 x History STD."),
    ("is_max_gt_2std", "Is {scenario} Max > History Max + 2 STD?", "Yes if the projection max is greater than History Max + 2 x History STD."),
    ("is_baseline_max_gt_2std", "Is Baseline Max > History Max + 2 STD?", "Yes if the projection max is greater than History Max + 2 x History STD (using Baseline scenario data)."),
    # --- Overall ---
    ("conclusion", "Conclusion", "Yes if any of the +/-2-STD breach tests in this sheet are Yes; otherwise No."),
]

_BASELINE_METRIC_KEYS = {"is_baseline_min_lt_2std", "is_baseline_max_gt_2std"}

_BOOLEAN_METRIC_KEYS = {
    "is_min_beyond", "is_max_beyond", "is_min_lt_2std", "is_max_gt_2std",
    "is_baseline_min_lt_2std", "is_baseline_max_gt_2std", "conclusion",
}

_DATE_METRIC_KEYS = {"time_of_min", "time_of_max"}

_FLOAT_METRIC_KEYS = {
    "history_min", "history_max", "history_mean", "history_std",
    "sevadv_min", "sevadv_max",
    "history_min_2std", "history_max_2std",
}

_PERCENT_METRIC_KEYS = {"model_contribution"}

def active_metric_columns(baseline_available: bool, scenario_label: str, scenario_value: str) -> list[tuple[str, str, str]]:
    """Metric columns to output, with chosen-scenario headers resolved.

    The dedicated Baseline columns are dropped when the Baseline scenario is not
    in the dataset, or when Baseline itself is the chosen scenario (in which case
    the chosen-scenario columns already represent the Baseline projection).
    """
    include_baseline = baseline_available and scenario_value != BASELINE_SCENARIO_VALUE
    label = scenario_label or "Scenario"
    resolved: list[tuple[str, str, str]] = []
    for key, header, description in SAAS_METRIC_COLUMNS:
        if key in _BASELINE_METRIC_KEYS and not include_baseline:
            continue
        resolved.append((key, header.replace("{scenario}", label), description.replace("{scenario}", label)))
    return resolved

def _excel_format_bool(value) -> str:
    if value is None:
        return ""
    return "Yes" if value else "No"

def _write_saas_excel_readme(ws, meta_lines: list[str], columns: list[tuple[str, str, str]], model_tabs: list[tuple[str, str]] | None = None, scenario_label: str | None = None) -> None:
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    wrap = Alignment(wrap_text=True, vertical="top")

    ws["A1"] = "Scenario Analysis as a Service (SAAS) - Historical Range Analysis"
    ws["A1"].font = title_font

    section_font = Font(bold=True, size=12)

    row = 3
    for line in meta_lines:
        ws.cell(row=row, column=1, value=line)
        row += 1

    # Metrics tab column definitions.
    row += 1
    ws.cell(row=row, column=1, value="Summary tab - column definitions").font = section_font
    row += 1
    ws.cell(row=row, column=1, value="Column").font = header_font
    ws.cell(row=row, column=2, value="Description").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=2).fill = header_fill
    row += 1
    for _key, header, description in columns:
        ws.cell(row=row, column=1, value=header)
        cell = ws.cell(row=row, column=2, value=description)
        cell.alignment = wrap
        row += 1

    # Worksheet guide: tab names mapped to the model they represent (chart tab
    # names are truncated to Excel's 31-character limit, so long model names may
    # not be fully visible on the tab itself).
    row += 1
    ws.cell(row=row, column=1, value="Worksheets in this file").font = section_font
    row += 1
    ws.cell(row=row, column=1, value="Worksheet").font = header_font
    ws.cell(row=row, column=2, value="Contents").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=2).fill = header_fill
    row += 1
    ws.cell(row=row, column=1, value="README")
    ws.cell(row=row, column=2, value="Column definitions (this sheet).").alignment = wrap
    row += 1
    ws.cell(row=row, column=1, value="Summary")
    ws.cell(row=row, column=2, value="Scenario metrics for every model and MEV.").alignment = wrap
    row += 1
    for tab_title, model_name in (model_tabs or []):
        descriptive = _model_descriptive_label(model_name)
        ws.cell(row=row, column=1, value=tab_title)
        contents = (
            f"Charts for {descriptive} ({model_name}) - History and {scenario_label} scenario projection."
            if scenario_label
            else f"Charts for {descriptive} ({model_name})"
        )
        ws.cell(row=row, column=2, value=contents).alignment = wrap
        row += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 90

def _write_saas_excel_metrics(ws, metric_rows: list[dict], columns: list[tuple[str, str, str]]) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D4ED8")

    if not metric_rows:
        ws["A1"] = "No metrics match the current filters and scenario."
        return

    for col_index, (_key, header, _description) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_index, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for row_index, record in enumerate(metric_rows, start=2):
        for col_index, (key, _header, _description) in enumerate(columns, start=1):
            value = record.get(key)
            cell = ws.cell(row=row_index, column=col_index)
            if key in _BOOLEAN_METRIC_KEYS:
                cell.value = _excel_format_bool(value)
            elif key in _DATE_METRIC_KEYS:
                cell.value = value
                if value is not None:
                    cell.number_format = "yyyy-mm-dd"
            elif key in _PERCENT_METRIC_KEYS:
                cell.value = None if value is None else float(value)
                cell.number_format = "0.0%"
            elif key in _FLOAT_METRIC_KEYS:
                cell.value = None if value is None else float(value)
                cell.number_format = "0.0000"
            else:
                cell.value = value

    ws.freeze_panes = "F2"
    for col_index, (key, header, _description) in enumerate(columns, start=1):
        width = 16
        if key == "model":
            width = 18
        elif key == "model_descriptive":
            width = 24
        elif key == "mev_type":
            width = 16
        elif key == "mev":
            width = 30
        elif key == "model_contribution":
            width = 18
        elif key == "minmax_daterange":
            width = 26
        elif len(header) > 22:
            width = min(46, len(header) + 2)
        ws.column_dimensions[get_column_letter(col_index)].width = width

_SAAS_EXCEL_SERIES_STYLES = [
    ("1E3A8A", 28575, None),      # History - solid navy, ~2.25pt
    ("93C5FD", 22225, "dash"),    # Projection - dashed faded blue, ~1.75pt
    ("0F766E", 19050, "dash"),    # Historical Min - dashed teal, ~1.5pt
    ("0F766E", 19050, "dash"),    # Historical Max - dashed teal, ~1.5pt
]

_EXCEL_INVALID_SHEET_CHARS = set("[]:*?/\\")

def _excel_safe_sheet_title(name: str, used: set[str]) -> str:
    cleaned = "".join("_" if ch in _EXCEL_INVALID_SHEET_CHARS else ch for ch in str(name or "")).strip()
    cleaned = cleaned[:31] or "Model"
    candidate = cleaned
    counter = 1
    while candidate.lower() in used:
        suffix = f"_{counter}"
        candidate = cleaned[: 31 - len(suffix)] + suffix
        counter += 1
    used.add(candidate.lower())
    return candidate

def _style_saas_excel_chart(chart: LineChart) -> None:
    for series, (color, width, dash) in zip(chart.series, _SAAS_EXCEL_SERIES_STYLES):
        line = LineProperties(w=width, solidFill=color)
        if dash:
            line.prstDash = dash
        graphical_properties = GraphicalProperties()
        graphical_properties.line = line
        series.graphicalProperties = graphical_properties
        series.smooth = False

def _saas_excel_axis_text(rotation: int = 0, size: int = 900) -> RichText:
    """Axis text properties: small font, optional rotation (degrees)."""
    return RichText(
        bodyPr=RichTextProperties(rot=int(rotation * -60000), vert="horz"),
        p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=size)), endParaRPr=CharacterProperties(sz=size))],
    )

def _saas_excel_nice_step(value_range: float, target_ticks: int = 5) -> tuple[float, str]:
    """Pick a clean y-axis major unit (~target_ticks intervals) and a matching
    number format, using the 1/2/5 x 10^k "nice numbers" rule."""
    if value_range is None or value_range <= 0:
        return 1.0, "0.00"
    raw = value_range / target_ticks
    exponent = math.floor(math.log10(raw))
    normalized = raw / (10 ** exponent)
    if normalized < 1.5:
        nice, step_exponent = 1, exponent
    elif normalized < 3:
        nice, step_exponent = 2, exponent
    elif normalized < 7:
        nice, step_exponent = 5, exponent
    else:
        nice, step_exponent = 1, exponent + 1  # 10 x 10^exponent
    step = nice * (10 ** step_exponent)
    decimals = max(0, -step_exponent)
    number_format = "0" if decimals == 0 else "0." + "0" * decimals
    return step, number_format

def _saas_family_chart_rows(specs: list[dict]) -> list[list[dict]]:
    """Arrange a model's chart specs into rows: each row is one transformed MEV
    followed by its raw MEVs (from the model's family map). Transformed families
    are ordered by label; anything left over is packed three per row."""
    if not specs:
        return []
    model_name = specs[0].get("model")
    family_map = SAAS_PAGE_DATA.get("model_mev_family_map", {}).get(model_name, {})
    spec_by_name: dict[str, dict] = {}
    for spec in specs:
        spec_by_name.setdefault(spec.get("mev_name"), spec)

    used: set = set()
    rows: list[list[dict]] = []
    transformed_present = [
        (name, raw_names) for name, raw_names in family_map.items() if name in spec_by_name
    ]
    transformed_present.sort(key=lambda item: (spec_by_name[item[0]].get("y_title") or "").lower())
    for transformed_name, raw_names in transformed_present:
        if transformed_name in used:
            continue
        row = [spec_by_name[transformed_name]]
        used.add(transformed_name)
        for raw_name in raw_names:
            raw_spec = spec_by_name.get(raw_name)
            if raw_spec is not None and raw_name not in used:
                row.append(raw_spec)
                used.add(raw_name)
        rows.append(row)

    leftovers = [spec for spec in specs if spec.get("mev_name") not in used]
    for index in range(0, len(leftovers), 3):
        rows.append(leftovers[index:index + 3])

    return rows

def _add_saas_chart(ws_charts, ws_data, spec: dict, scenario_label: str, data_row: int, anchor_col_index: int, anchor_row_num: int) -> int:
    """Write a single chart's data block and place the styled chart at the anchor.
    Returns the next free row in the (hidden) Chart Data sheet."""
    history = spec["history"]
    projection = spec["projection"]
    mev_label = spec.get("y_title") or spec.get("title") or ""

    headers = ["MEV", "Period", "History", f"{scenario_label} (Projection)", "History Min", "History Max"]
    for col_index, header in enumerate(headers, start=1):
        ws_data.cell(row=data_row, column=col_index, value=header)
    first_data_row = data_row + 1
    for offset, label in enumerate(spec["labels"]):
        current = first_data_row + offset
        ws_data.cell(row=current, column=1, value=mev_label)
        ws_data.cell(row=current, column=2, value=label)
        ws_data.cell(row=current, column=3, value=history[offset])
        ws_data.cell(row=current, column=4, value=projection[offset])
        ws_data.cell(row=current, column=5, value=spec["hist_min"])
        ws_data.cell(row=current, column=6, value=spec["hist_max"])
    last_data_row = max(first_data_row, first_data_row + len(spec["labels"]) - 1)

    chart = LineChart()
    chart.title = spec["title"]
    chart.style = 2
    chart.height = 9
    chart.width = 15
    chart.y_axis.title = None  # the chart title already names the MEV
    chart.x_axis.title = None  # quarter labels are self-explanatory; frees space for the bottom legend
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.y_axis.majorGridlines = None  # remove horizontal gridlines

    # Scale the y-axis so the historical Min/Max band fills the central half of
    # the plot: the Historical Min line sits at ~25% of the height and the
    # Historical Max at ~75% (min - 0.5*band .. max + 0.5*band). The axis is
    # widened to include any projection excursions so nothing is clipped; the
    # baseline (x-axis) sits at the axis minimum.
    plotted = [value for value in history if value is not None]
    plotted += [value for value in projection if value is not None]
    hist_low = spec.get("hist_min")
    hist_high = spec.get("hist_max")
    if hist_low is not None:
        plotted.append(hist_low)
    if hist_high is not None:
        plotted.append(hist_high)
    if plotted:
        data_min = min(plotted)
        data_max = max(plotted)
        if hist_low is not None and hist_high is not None and hist_high > hist_low:
            band = hist_high - hist_low
            axis_min = min(hist_low - 0.5 * band, data_min)
            axis_max = max(hist_high + 0.5 * band, data_max)
        else:
            pad = (data_max - data_min) * 0.1 or 1.0
            axis_min = data_min - pad
            axis_max = data_max + pad
        chart.y_axis.scaling.min = axis_min
        chart.y_axis.scaling.max = axis_max
        chart.y_axis.crosses = "min"

        # Clean tick frequency (~8 intervals) and matching rounding for labels.
        major_unit, axis_number_format = _saas_excel_nice_step(axis_max - axis_min, target_ticks=8)
        chart.y_axis.majorUnit = major_unit
        chart.y_axis.numFmt = axis_number_format
        chart.y_axis.txPr = _saas_excel_axis_text(rotation=0)

    # Thin out and rotate the period labels so they don't crowd each other.
    point_count = len(spec["labels"])
    label_skip = max(1, round(point_count / 12))
    chart.x_axis.tickLblSkip = label_skip
    chart.x_axis.tickMarkSkip = label_skip
    chart.x_axis.txPr = _saas_excel_axis_text(rotation=45)
    # Keep the period labels at the bottom of the plot rather than at the y=0
    # crossing (where they would sit in the middle of the data for MEVs that go
    # negative).
    chart.x_axis.tickLblPos = "low"

    # Pin the plot to the upper area and the legend just below it, with only a
    # small gap for the rotated x-axis labels. (chart.layout drives the inner
    # plot area; openpyxl copies it onto plot_area on write.)
    chart.layout = Layout(
        manualLayout=ManualLayout(
            layoutTarget="inner", xMode="edge", yMode="edge",
            x=0.10, y=0.08, w=0.86, h=0.64,
        )
    )
    if chart.legend is not None:
        chart.legend.position = "b"
        chart.legend.overlay = False
        chart.legend.layout = Layout(
            manualLayout=ManualLayout(
                xMode="edge", yMode="edge",
                x=0.10, y=0.84, w=0.86, h=0.10,
            )
        )

    data_ref = Reference(ws_data, min_col=3, max_col=6, min_row=data_row, max_row=last_data_row)
    cats_ref = Reference(ws_data, min_col=2, min_row=first_data_row, max_row=last_data_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    _style_saas_excel_chart(chart)

    ws_charts.add_chart(chart, f"{get_column_letter(anchor_col_index)}{anchor_row_num}")
    return last_data_row + 2

def _write_saas_model_charts(ws_charts, ws_data, specs: list[dict], scenario_label: str, data_row: int) -> int:
    # Each row holds one transformed MEV followed by its raw MEVs. Block size
    # leaves a gap so titles, legends and tick labels never overlap.
    chart_block_cols = 9
    chart_block_rows = 22

    for row_index, row_specs in enumerate(_saas_family_chart_rows(specs)):
        anchor_row_num = 1 + row_index * chart_block_rows
        for col_index, spec in enumerate(row_specs):
            anchor_col_index = 1 + col_index * chart_block_cols
            data_row = _add_saas_chart(
                ws_charts, ws_data, spec, scenario_label, data_row, anchor_col_index, anchor_row_num,
            )

    return data_row

def build_saas_excel_workbook(metric_rows: list[dict], chart_specs: list[dict], meta_lines: list[str], scenario_label: str, columns: list[tuple[str, str, str]]) -> Workbook:
    # Group charts by model so each model gets its own worksheet (preserve order).
    specs_by_model: dict[str, list[dict]] = {}
    model_order: list[str] = []
    for spec in chart_specs:
        model_name = spec.get("model") or "Model"
        if model_name not in specs_by_model:
            specs_by_model[model_name] = []
            model_order.append(model_name)
        specs_by_model[model_name].append(spec)

    # Pre-compute the Excel-safe (<= 31 char, unique) tab title for each model,
    # using the model's descriptive name; the README maps each tab back to the
    # full descriptive name and the raw model name.
    used_titles = {"readme", "summary", "chart data"}
    model_tabs = [
        (_excel_safe_sheet_title(_model_descriptive_label(model_name), used_titles), model_name)
        for model_name in model_order
    ]
    title_by_model = {model_name: title for title, model_name in model_tabs}

    workbook = Workbook()
    readme_ws = workbook.active
    readme_ws.title = "README"
    readme_ws.sheet_view.showGridLines = False
    _write_saas_excel_readme(readme_ws, meta_lines, columns, model_tabs, scenario_label)

    # Metrics keeps gridlines; every other tab hides them.
    metrics_ws = workbook.create_sheet("Summary")
    _write_saas_excel_metrics(metrics_ws, metric_rows, columns)

    if not model_order:
        fallback_ws = workbook.create_sheet("Charts")
        fallback_ws.sheet_view.showGridLines = False
        fallback_ws["A1"] = "No charts match the current filters and scenario."
        return workbook

    model_sheets = {}
    for model_name in model_order:
        sheet = workbook.create_sheet(title_by_model[model_name])
        sheet.sheet_view.showGridLines = False
        model_sheets[model_name] = sheet

    chart_data_ws = workbook.create_sheet("Chart Data")
    chart_data_ws.sheet_state = "hidden"
    chart_data_ws.sheet_view.showGridLines = False

    data_row = 1
    for model_name in model_order:
        data_row = _write_saas_model_charts(
            model_sheets[model_name], chart_data_ws, specs_by_model[model_name], scenario_label, data_row,
        )

    return workbook

_SAAS_COMPARE_LINE_COLOR = "C00000"

_SAAS_COMPARE_LINE_DASHES = [None, "dash", "sysDot", "dashDot", "lgDash", "lgDashDot"]

def _style_saas_recon_chart(chart: LineChart) -> None:
    # All cycles share the same red line; the cycles are differentiated by line
    # style (primary = solid, then dashed / dotted / dash-dot for each compare).
    for index, series in enumerate(chart.series):
        dash = _SAAS_COMPARE_LINE_DASHES[index % len(_SAAS_COMPARE_LINE_DASHES)]
        line = LineProperties(w=22225, solidFill=_SAAS_COMPARE_LINE_COLOR)
        if dash:
            line.prstDash = dash
        graphical_properties = GraphicalProperties()
        graphical_properties.line = line
        series.graphicalProperties = graphical_properties
        series.smooth = False

def _excel_set_number(ws, row, col, value, number_format) -> None:
    cell = ws.cell(row=row, column=col)
    if value is None:
        cell.value = None
    else:
        cell.value = float(value)
        cell.number_format = number_format

def _recon_metrics(periods: list[str], primary_vals: list[float], compare_vals: list[float]) -> dict:
    """Summary reconciliation metrics for one MEV between two reporting cycles."""
    n = len(periods)
    signed = [c - p for p, c in zip(primary_vals, compare_vals)]
    magnitude = [abs(d) for d in signed]
    pct_magnitude = [abs(d) / abs(p) if p else None for p, d in zip(primary_vals, signed)]

    max_abs = max(magnitude) if magnitude else None
    idx_max = magnitude.index(max_abs) if max_abs is not None else None
    date_max = periods[idx_max] if idx_max is not None else ""
    mean_abs = statistics.fmean(magnitude) if magnitude else None
    pct_valid = [value for value in pct_magnitude if value is not None]
    max_pct = max(pct_valid) if pct_valid else None
    bias = statistics.fmean(signed) if signed else None
    rmse = math.sqrt(statistics.fmean([d * d for d in signed])) if signed else None
    try:
        corr = statistics.correlation(primary_vals, compare_vals) if n >= 2 else None
    except statistics.StatisticsError:
        corr = None

    top_order = sorted(range(n), key=lambda i: magnitude[i], reverse=True)[:3]
    top_parts = []
    for i in top_order:
        pct_text = f"{pct_magnitude[i] * 100:.2f}%" if pct_magnitude[i] is not None else "n/a"
        top_parts.append(f"{periods[i]}: {signed[i]:+.4f} ({pct_text})")

    return {
        "n": n,
        "max_abs": max_abs,
        "date_max": date_max,
        "mean_abs": mean_abs,
        "max_pct": max_pct,
        "bias": bias,
        "rmse": rmse,
        "corr": corr,
        "top_dates": "; ".join(top_parts),
    }

def _write_recon_data_block(ws_data, block: dict, cycles: list[str], compare_cycles: list[str], data_row: int):
    """Write one MEV's per-date values + |%diff| to the hidden data sheet.
    Returns (next_data_row, refs) where refs lets the Summary/charts reference it."""
    cycle_count = len(cycles)
    header_row = data_row
    ws_data.cell(row=header_row, column=1, value="Period")
    for offset, cycle in enumerate(cycles):
        ws_data.cell(row=header_row, column=2 + offset, value=str(cycle))
    for offset, cycle in enumerate(compare_cycles):
        ws_data.cell(row=header_row, column=2 + cycle_count + offset, value=f"%diff {cycle}")

    first_data_row = header_row + 1
    periods = block["periods"]
    values = block["values"]
    diffs = block["diffs"]
    for index, period in enumerate(periods):
        current = first_data_row + index
        ws_data.cell(row=current, column=1, value=period)
        for offset, cycle in enumerate(cycles):
            ws_data.cell(row=current, column=2 + offset, value=values[cycle][index])
        for offset, cycle in enumerate(compare_cycles):
            pct = diffs[cycle]["pct"][index]
            ws_data.cell(row=current, column=2 + cycle_count + offset, value=(abs(pct) if pct is not None else None))
    last_data_row = first_data_row + len(periods) - 1

    sheet_quoted = f"'{ws_data.title}'"
    pct_a1 = {}
    for offset, cycle in enumerate(compare_cycles):
        letter = get_column_letter(2 + cycle_count + offset)
        pct_a1[cycle] = f"{sheet_quoted}!${letter}${first_data_row}:${letter}${last_data_row}"

    refs = {
        "values": Reference(ws_data, min_col=2, max_col=1 + cycle_count, min_row=header_row, max_row=last_data_row),
        "periods": Reference(ws_data, min_col=1, min_row=first_data_row, max_row=last_data_row),
        "pct_a1": pct_a1,
        "n": len(periods),
    }
    return last_data_row + 2, refs

_RECON_SUMMARY_HEADERS = [
    "Model", "Model Descriptive Name", "MEV Type", "MEV", "Primary Cycle", "Compare Cycle",
    "Overlap N", "Match?", "# > tol", "% > tol", "Max |Diff|", "Date of Max |Diff|", "Mean |Diff|",
    "Max |%Diff|", "Mean Diff (bias)", "RMSE", "Correlation", "Top differing dates",
]

def _write_recon_summary(ws, models, primary, compare_cycles, refs, threshold_fraction) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="center")

    ws["A1"] = "Relative threshold (|%Diff| <=):"
    ws["A1"].font = bold
    threshold_cell = ws["B1"]
    threshold_cell.value = threshold_fraction
    threshold_cell.number_format = "0.000%"
    threshold_cell.font = bold
    ws["A2"] = "Edit cell B1 to recompute Match?, # > tol and % > tol."

    header_row = 4
    for col_index, header in enumerate(_RECON_SUMMARY_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_index, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap

    row = header_row + 1
    wrote_any = False
    for model_block in models:
        model_name = model_block["model"]
        descriptive = _model_descriptive_label(model_name)
        for block in model_block["mevs"]:
            block_refs = refs.get((model_name, block["mev_name"]))
            for cycle in compare_cycles:
                metrics = _recon_metrics(block["periods"], block["values"][primary], block["values"][cycle])
                ws.cell(row=row, column=1, value=model_name)
                ws.cell(row=row, column=2, value=descriptive)
                ws.cell(row=row, column=3, value=block["mev_type"])
                ws.cell(row=row, column=4, value=block["mev_label"])
                ws.cell(row=row, column=5, value=str(primary))
                ws.cell(row=row, column=6, value=str(cycle))
                ws.cell(row=row, column=7, value=metrics["n"])
                # Match? - live vs the editable threshold (N{row} = Max |%Diff|).
                if metrics["max_pct"] is None:
                    ws.cell(row=row, column=8, value="n/a")
                else:
                    ws.cell(row=row, column=8, value=f'=IF(N{row}<=$B$1,"Yes","No")')
                # # > tol / % > tol - live via COUNTIF over the hidden data sheet.
                if block_refs and cycle in block_refs["pct_a1"]:
                    ws.cell(row=row, column=9, value=f'=COUNTIF({block_refs["pct_a1"][cycle]},">"&$B$1)')
                pct_over_cell = ws.cell(row=row, column=10, value=f'=IF(G{row}=0,"",I{row}/G{row})')
                pct_over_cell.number_format = "0.00%"
                _excel_set_number(ws, row, 11, metrics["max_abs"], "0.0000")
                ws.cell(row=row, column=12, value=metrics["date_max"])
                _excel_set_number(ws, row, 13, metrics["mean_abs"], "0.0000")
                _excel_set_number(ws, row, 14, metrics["max_pct"], "0.00%")
                _excel_set_number(ws, row, 15, metrics["bias"], "0.0000")
                _excel_set_number(ws, row, 16, metrics["rmse"], "0.0000")
                _excel_set_number(ws, row, 17, metrics["corr"], "0.0000")
                ws.cell(row=row, column=18, value=metrics["top_dates"]).alignment = wrap
                row += 1
                wrote_any = True

    if not wrote_any:
        ws.cell(
            row=header_row + 1, column=1,
            value="No overlapping historical data. Select at least one Compare To reporting cycle, then export again.",
        )

    ws.freeze_panes = "E5"
    widths = {1: 18, 2: 24, 4: 28, 12: 16, 18: 48}
    for col_index in range(1, len(_RECON_SUMMARY_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_index)].width = widths.get(col_index, 14)

def _bottom_caption_title(caption: str) -> Title:
    """A small italic text, floated at the very bottom of the chart (below the
    legend). Implemented as the x-axis title so it is a separate text element
    from the chart's main title at the top."""
    para = Paragraph(
        pPr=ParagraphProperties(defRPr=CharacterProperties(b=False, i=True, sz=850)),
        r=[RegularTextRun(rPr=CharacterProperties(b=False, i=True, sz=850), t=caption)],
    )
    title = Title()
    title.tx = Text(rich=RichText(p=[para]))
    title.overlay = True
    title.layout = Layout(manualLayout=ManualLayout(xMode="edge", yMode="edge", x=0.04, y=0.88))
    return title

def _build_recon_line_chart(mev_label, mev_type, values_ref, periods_ref, point_count, axis_min, axis_max, title_suffix="historical reconciliation", caption=None) -> LineChart:
    chart = LineChart()
    type_suffix = f" [{mev_type}]" if mev_type and mev_type != "—" else ""
    main_title = f"{mev_label}{type_suffix} - {title_suffix}"
    has_caption = bool(caption)
    chart.title = main_title  # MEV name stays at the top
    chart.style = 2
    chart.height = 9 if has_caption else 8
    chart.width = 15
    chart.y_axis.title = None  # the chart title already names the MEV
    # The caption (description / transformed MEV) floats at the bottom, below the
    # legend - rendered via the x-axis title so it is independent of the top title.
    chart.x_axis.title = _bottom_caption_title(caption) if has_caption else None
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.y_axis.majorGridlines = None

    # Scale the y-axis to the data range (with padding) so trends are visible -
    # the axis does not have to start at 0.
    if axis_min is not None and axis_max is not None and axis_max > axis_min:
        chart.y_axis.scaling.min = axis_min
        chart.y_axis.scaling.max = axis_max
        chart.y_axis.crosses = "min"
        major_unit, axis_number_format = _saas_excel_nice_step(axis_max - axis_min, target_ticks=6)
        chart.y_axis.majorUnit = major_unit
        chart.y_axis.numFmt = axis_number_format
        chart.y_axis.txPr = _saas_excel_axis_text(rotation=0)

    label_skip = max(1, round(point_count / 12))
    chart.x_axis.tickLblSkip = label_skip
    chart.x_axis.tickMarkSkip = label_skip
    chart.x_axis.txPr = _saas_excel_axis_text(rotation=45)
    chart.x_axis.tickLblPos = "low"

    # No y-axis title anymore, so only a thin left margin is needed for the tick
    # labels - the plot uses the full width. When a caption is shown the plot is
    # shorter, the legend is brought up, and the title block sits at the bottom.
    if has_caption:
        # Top title, plot, x labels, legend (lower), then the bottom caption.
        plot_y, plot_h, legend_y = 0.11, 0.49, 0.76
    else:
        plot_y, plot_h, legend_y = 0.08, 0.66, 0.86
    chart.layout = Layout(
        manualLayout=ManualLayout(
            layoutTarget="inner", xMode="edge", yMode="edge",
            x=0.08, y=plot_y, w=0.89, h=plot_h,
        )
    )
    if chart.legend is not None:
        chart.legend.position = "b"
        chart.legend.overlay = False
        chart.legend.layout = Layout(
            manualLayout=ManualLayout(
                xMode="edge", yMode="edge",
                x=0.08, y=legend_y, w=0.89, h=0.09,
            )
        )

    chart.add_data(values_ref, titles_from_data=True)
    chart.set_categories(periods_ref)
    _style_saas_recon_chart(chart)
    return chart

def _projection_chart_caption(model_name: str, mev_name: str, mev_type: str) -> str:
    """Caption shown beneath a projection chart: transformed MEVs show their
    description; raw MEVs show the transformed MEV(s) they feed into."""
    if mev_type == "Transformed":
        description = SAAS_PAGE_DATA.get("mev_description_map", {}).get(mev_name)
        return f"Description: {description}" if description else ""
    family_map = SAAS_PAGE_DATA.get("model_mev_family_map", {}).get(model_name, {})
    parents = [transformed for transformed, raws in family_map.items() if mev_name in raws]
    return f"Transformed MEV: {', '.join(parents)}" if parents else ""

def _write_recon_model_charts(ws, model_block, refs, title_suffix="historical reconciliation", with_captions=False) -> None:
    chart_block_cols = 9
    chart_block_rows = 20

    # Each row: one transformed MEV followed by its raw MEVs (blocks are already
    # family-ordered, so a new row starts at each transformed MEV).
    layout_rows: list[list[dict]] = []
    for block in model_block["mevs"]:
        if block["mev_type"] == "Transformed" or not layout_rows:
            layout_rows.append([block])
        else:
            layout_rows[-1].append(block)

    for row_index, row_blocks in enumerate(layout_rows):
        anchor_row = 1 + row_index * chart_block_rows
        for col_index, block in enumerate(row_blocks):
            block_refs = refs.get((model_block["model"], block["mev_name"]))
            if not block_refs:
                continue
            plotted = [value for series in block["values"].values() for value in series if value is not None]
            if plotted:
                data_min, data_max = min(plotted), max(plotted)
                pad = (data_max - data_min) * 0.08 or (abs(data_max) * 0.08) or 1.0
                axis_min, axis_max = data_min - pad, data_max + pad
            else:
                axis_min = axis_max = None
            anchor_col_index = 1 + col_index * chart_block_cols
            caption = (
                _projection_chart_caption(model_block["model"], block["mev_name"], block["mev_type"])
                if with_captions else None
            )
            chart = _build_recon_line_chart(
                block["mev_label"], block["mev_type"], block_refs["values"], block_refs["periods"],
                block_refs["n"], axis_min, axis_max, title_suffix, caption or None,
            )
            ws.add_chart(chart, f"{get_column_letter(anchor_col_index)}{anchor_row}")

    if not model_block["mevs"]:
        ws["A1"] = "No overlapping historical data for this model."

def _write_recon_readme(ws, meta_lines, model_tabs, primary) -> None:
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    section_font = Font(bold=True, size=12)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws["A1"] = "Scenario Analysis as a Service (SAAS) - Historical Reconciliation"
    ws["A1"].font = title_font

    row = 3
    for line in meta_lines:
        ws.cell(row=row, column=1, value=line)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Summary tab - column definitions").font = section_font
    row += 1
    ws.cell(row=row, column=1, value="Column").font = header_font
    ws.cell(row=row, column=2, value="Description").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=2).fill = header_fill
    row += 1
    definitions = [
        ("Relative threshold (cell B1)", "Editable (default 3.0%). |%Diff| at or below this counts as a match. Change it to recompute Match?, # > tol and % > tol live."),
        ("Model / Model Descriptive Name", "Model name (GMIS) and its descriptive name."),
        ("MEV Type / MEV", "Whether the MEV is Raw or Transformed, and its label."),
        ("Primary Cycle / Compare Cycle", "The two reporting cycles being reconciled (compare vs primary)."),
        ("Overlap N", "Number of historical quarters present in every selected cycle for that MEV."),
        ("Match?", "Yes if Max |%Diff| <= the relative threshold (cell B1); otherwise No."),
        ("# > tol / % > tol", "Count and share of overlapping dates whose |%Diff| exceeds the threshold."),
        ("Max |Diff| / Date of Max |Diff|", "Largest absolute difference between the cycles and the quarter where it occurs."),
        ("Mean |Diff|", "Average absolute difference across the overlapping dates."),
        ("Max |%Diff|", "Largest absolute percentage difference (relative to the primary cycle)."),
        ("Mean Diff (bias)", "Average signed difference (compare - primary); detects a systematic offset."),
        ("RMSE", "Root mean squared difference between the cycles."),
        ("Correlation", "Pearson correlation between the two cycles' series (close to 1 expected)."),
        ("Top differing dates", "Up to three quarters with the largest absolute difference (shown as quarter: diff (%))."),
    ]
    highlight_fill = PatternFill("solid", fgColor="FFE699")
    highlight_font = Font(bold=True, color="7F6000")
    for column_name, description in definitions:
        name_cell = ws.cell(row=row, column=1, value=column_name)
        desc_cell = ws.cell(row=row, column=2, value=description)
        desc_cell.alignment = wrap
        if column_name.startswith("Relative threshold"):
            # Highlight so the user immediately spots the editable threshold.
            name_cell.fill = highlight_fill
            desc_cell.fill = highlight_fill
            name_cell.font = highlight_font
            desc_cell.font = highlight_font
            desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Worksheets in this file").font = section_font
    row += 1
    ws.cell(row=row, column=1, value="Worksheet").font = header_font
    ws.cell(row=row, column=2, value="Contents").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=2).fill = header_fill
    row += 1
    ws.cell(row=row, column=1, value="README")
    ws.cell(row=row, column=2, value="Column definitions (this sheet).").alignment = wrap
    row += 1
    ws.cell(row=row, column=1, value="Summary")
    ws.cell(row=row, column=2, value="One row per Model x MEV x compare cycle, with the reconciliation metrics.").alignment = wrap
    row += 1
    for tab_title, model_name in (model_tabs or []):
        descriptive = _model_descriptive_label(model_name)
        ws.cell(row=row, column=1, value=tab_title)
        ws.cell(row=row, column=2, value=f"Historical line charts for {descriptive} ({model_name}) - each cycle overlaid per MEV.").alignment = wrap
        row += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 92

def build_saas_reconciliation_workbook(recon: dict, meta_lines: list[str], scenario_label: str, threshold_fraction: float) -> Workbook:
    primary = recon.get("primary")
    cycles = recon.get("cycles", [])
    compare_cycles = recon.get("compare_cycles", [])
    models = recon.get("models", [])

    used_titles = {"readme", "summary", "reconciliation data"}
    model_tabs = [
        (_excel_safe_sheet_title(_model_descriptive_label(model_block["model"]), used_titles), model_block["model"])
        for model_block in models
    ]
    title_by_model = {model_name: title for title, model_name in model_tabs}

    workbook = Workbook()
    readme_ws = workbook.active
    readme_ws.title = "README"
    readme_ws.sheet_view.showGridLines = False

    summary_ws = workbook.create_sheet("Summary")  # gridlines kept on the table tab

    model_sheets = {}
    for model_block in models:
        sheet = workbook.create_sheet(title_by_model[model_block["model"]])
        sheet.sheet_view.showGridLines = False
        model_sheets[model_block["model"]] = sheet

    data_ws = workbook.create_sheet("Reconciliation Data")
    data_ws.sheet_state = "hidden"
    data_ws.sheet_view.showGridLines = False

    # Write the hidden per-date data first, collecting references for the Summary
    # formulas and the charts.
    refs: dict = {}
    data_row = 1
    for model_block in models:
        for block in model_block["mevs"]:
            data_row, block_refs = _write_recon_data_block(data_ws, block, cycles, compare_cycles, data_row)
            refs[(model_block["model"], block["mev_name"])] = block_refs

    _write_recon_summary(summary_ws, models, primary, compare_cycles, refs, threshold_fraction)
    for model_block in models:
        _write_recon_model_charts(model_sheets[model_block["model"]], model_block, refs)
    _write_recon_readme(readme_ws, meta_lines, model_tabs, primary)

    return workbook

def _sign(value: float) -> int:
    return (value > 0) - (value < 0)

def _projection_metrics(periods: list[str], primary_vals: list[float], compare_vals: list[float]) -> dict:
    """Divergence + endpoint + severity metrics for one MEV between two cycles' projections."""
    n = len(periods)
    signed = [c - p for p, c in zip(primary_vals, compare_vals)]
    magnitude = [abs(d) for d in signed]
    pct_magnitude = [abs(d) / abs(p) if p else None for p, d in zip(primary_vals, signed)]

    bias = statistics.fmean(signed) if signed else None
    mean_abs = statistics.fmean(magnitude) if magnitude else None
    rmse = math.sqrt(statistics.fmean([d * d for d in signed])) if signed else None
    max_abs = max(magnitude) if magnitude else None
    idx_max = magnitude.index(max_abs) if max_abs is not None else None
    quarter_max = periods[idx_max] if idx_max is not None else ""
    pct_valid = [value for value in pct_magnitude if value is not None]
    max_pct = max(pct_valid) if pct_valid else None
    terminal = signed[-1] if signed else None
    try:
        corr = statistics.correlation(primary_vals, compare_vals) if n >= 2 else None
    except statistics.StatisticsError:
        corr = None

    same, total = 0, 0
    for i in range(1, n):
        total += 1
        if _sign(primary_vals[i] - primary_vals[i - 1]) == _sign(compare_vals[i] - compare_vals[i - 1]):
            same += 1
    dir_agreement = (same / total) if total else None

    delta_drift = (compare_vals[-1] - compare_vals[0]) - (primary_vals[-1] - primary_vals[0]) if n >= 1 else None

    top_order = sorted(range(n), key=lambda i: magnitude[i], reverse=True)[:3]
    top_parts = []
    for i in top_order:
        pct_text = f"{pct_magnitude[i] * 100:.2f}%" if pct_magnitude[i] is not None else "n/a"
        top_parts.append(f"{periods[i]}: {signed[i]:+.4f} ({pct_text})")

    return {
        "n": n,
        "bias": bias,
        "mean_abs": mean_abs,
        "rmse": rmse,
        "max_abs": max_abs,
        "quarter_max": quarter_max,
        "max_pct": max_pct,
        "terminal": terminal,
        "corr": corr,
        "dir_agreement": dir_agreement,
        "delta_drift": delta_drift,
        "primary_max": max(primary_vals) if primary_vals else None,
        "primary_min": min(primary_vals) if primary_vals else None,
        "compare_max": max(compare_vals) if compare_vals else None,
        "compare_min": min(compare_vals) if compare_vals else None,
        "top_quarters": "; ".join(top_parts),
    }

_PROJECTION_SUMMARY_HEADERS = [
    "Model", "Model Descriptive Name", "MEV Type", "MEV", "Primary Cycle", "Compare Cycle",
    "Horizon N", "Mean Diff (bias)", "Mean |Diff|", "RMSE", "Max |Diff|", "Quarter of Max |Diff|",
    "Max |%Diff|", "Terminal Diff", "Correlation", "Directional Agreement", "Delta Total Drift",
    "Primary Max", "Primary Min", "Compare Max", "Compare Min", "Top differing quarters",
]

def _write_projection_summary(ws, models, primary, compare_cycles) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    wrap = Alignment(wrap_text=True, vertical="center")

    for col_index, header in enumerate(_PROJECTION_SUMMARY_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_index, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap

    row = 2
    wrote_any = False
    for model_block in models:
        model_name = model_block["model"]
        descriptive = _model_descriptive_label(model_name)
        for block in model_block["mevs"]:
            for cycle in compare_cycles:
                metrics = _projection_metrics(block["periods"], block["values"][primary], block["values"][cycle])
                ws.cell(row=row, column=1, value=model_name)
                ws.cell(row=row, column=2, value=descriptive)
                ws.cell(row=row, column=3, value=block["mev_type"])
                ws.cell(row=row, column=4, value=block["mev_label"])
                ws.cell(row=row, column=5, value=str(primary))
                ws.cell(row=row, column=6, value=str(cycle))
                ws.cell(row=row, column=7, value=metrics["n"])
                _excel_set_number(ws, row, 8, metrics["bias"], "0.0000")
                _excel_set_number(ws, row, 9, metrics["mean_abs"], "0.0000")
                _excel_set_number(ws, row, 10, metrics["rmse"], "0.0000")
                _excel_set_number(ws, row, 11, metrics["max_abs"], "0.0000")
                ws.cell(row=row, column=12, value=metrics["quarter_max"])
                _excel_set_number(ws, row, 13, metrics["max_pct"], "0.00%")
                _excel_set_number(ws, row, 14, metrics["terminal"], "0.0000")
                _excel_set_number(ws, row, 15, metrics["corr"], "0.0000")
                _excel_set_number(ws, row, 16, metrics["dir_agreement"], "0.0%")
                _excel_set_number(ws, row, 17, metrics["delta_drift"], "0.0000")
                _excel_set_number(ws, row, 18, metrics["primary_max"], "0.0000")
                _excel_set_number(ws, row, 19, metrics["primary_min"], "0.0000")
                _excel_set_number(ws, row, 20, metrics["compare_max"], "0.0000")
                _excel_set_number(ws, row, 21, metrics["compare_min"], "0.0000")
                ws.cell(row=row, column=22, value=metrics["top_quarters"]).alignment = wrap
                row += 1
                wrote_any = True

    if not wrote_any:
        ws.cell(row=2, column=1, value="No overlapping projection data. Select at least one Compare To reporting cycle, then export again.")

    ws.freeze_panes = "E2"
    widths = {1: 18, 2: 24, 4: 28, 12: 16, 22: 48}
    for col_index in range(1, len(_PROJECTION_SUMMARY_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_index)].width = widths.get(col_index, 14)

def _write_projection_readme(ws, meta_lines, model_tabs, primary) -> None:
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    section_font = Font(bold=True, size=12)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws["A1"] = "Scenario Analysis as a Service (SAAS) - Projection Comparison"
    ws["A1"].font = title_font

    row = 3
    for line in meta_lines:
        ws.cell(row=row, column=1, value=line)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Summary tab - column definitions").font = section_font
    row += 1
    ws.cell(row=row, column=1, value="Column").font = header_font
    ws.cell(row=row, column=2, value="Description").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=2).fill = header_fill
    row += 1
    definitions = [
        ("Model / Model Descriptive Name", "Model name (GMIS) and its descriptive name."),
        ("MEV Type / MEV", "Whether the MEV is Raw or Transformed, and its label."),
        ("Primary Cycle / Compare Cycle", "The two reporting cycles being compared (compare vs primary)."),
        ("Horizon N", "Number of quarter offsets (Q0..H) present in every selected cycle for that MEV."),
        ("Mean Diff (bias)", "Average signed difference (compare - primary); detects a systematic offset across the path."),
        ("Mean |Diff| / RMSE", "Average / root-mean-squared magnitude of the per-quarter difference."),
        ("Max |Diff| / Quarter of Max |Diff|", "Largest absolute difference between the paths and the quarter offset where it occurs."),
        ("Max |%Diff|", "Largest absolute percentage difference (relative to the primary cycle)."),
        ("Terminal Diff", "Difference at the last common quarter (compare - primary at end of horizon)."),
        ("Correlation", "Pearson correlation between the two projection paths (shape agreement)."),
        ("Directional Agreement", "Share of quarters where both paths move the same direction quarter-over-quarter."),
        ("Delta Total Drift", "(compare end - compare start) - (primary end - primary start): is the trajectory steeper/flatter."),
        ("Primary/Compare Max & Min", "Peak and trough of each cycle's projection path - compare severity of the excursion."),
        ("Top differing quarters", "Up to three quarter offsets with the largest absolute difference (quarter: diff (%))."),
    ]
    for column_name, description in definitions:
        ws.cell(row=row, column=1, value=column_name)
        ws.cell(row=row, column=2, value=description).alignment = wrap
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Worksheets in this file").font = section_font
    row += 1
    ws.cell(row=row, column=1, value="Worksheet").font = header_font
    ws.cell(row=row, column=2, value="Contents").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=2).fill = header_fill
    row += 1
    ws.cell(row=row, column=1, value="README")
    ws.cell(row=row, column=2, value="Column definitions (this sheet).").alignment = wrap
    row += 1
    ws.cell(row=row, column=1, value="Summary")
    ws.cell(row=row, column=2, value="One row per Model x MEV x compare cycle, with the projection-comparison metrics.").alignment = wrap
    row += 1
    for tab_title, model_name in (model_tabs or []):
        descriptive = _model_descriptive_label(model_name)
        ws.cell(row=row, column=1, value=tab_title)
        ws.cell(row=row, column=2, value=f"Projection line charts for {descriptive} ({model_name}) - each cycle's path overlaid per MEV (by quarter offset).").alignment = wrap
        row += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 92

def build_saas_projection_workbook(comparison: dict, meta_lines: list[str], scenario_label: str) -> Workbook:
    primary = comparison.get("primary")
    cycles = comparison.get("cycles", [])
    compare_cycles = comparison.get("compare_cycles", [])
    models = comparison.get("models", [])

    used_titles = {"readme", "summary", "projection data"}
    model_tabs = [
        (_excel_safe_sheet_title(_model_descriptive_label(model_block["model"]), used_titles), model_block["model"])
        for model_block in models
    ]
    title_by_model = {model_name: title for title, model_name in model_tabs}

    workbook = Workbook()
    readme_ws = workbook.active
    readme_ws.title = "README"
    readme_ws.sheet_view.showGridLines = False

    summary_ws = workbook.create_sheet("Summary")

    model_sheets = {}
    for model_block in models:
        sheet = workbook.create_sheet(title_by_model[model_block["model"]])
        sheet.sheet_view.showGridLines = False
        model_sheets[model_block["model"]] = sheet

    data_ws = workbook.create_sheet("Projection Data")
    data_ws.sheet_state = "hidden"
    data_ws.sheet_view.showGridLines = False

    refs: dict = {}
    data_row = 1
    for model_block in models:
        for block in model_block["mevs"]:
            data_row, block_refs = _write_recon_data_block(data_ws, block, cycles, compare_cycles, data_row)
            refs[(model_block["model"], block["mev_name"])] = block_refs

    _write_projection_summary(summary_ws, models, primary, compare_cycles)
    for model_block in models:
        _write_recon_model_charts(model_sheets[model_block["model"]], model_block, refs, title_suffix="projection comparison", with_captions=True)
    _write_projection_readme(readme_ws, meta_lines, model_tabs, primary)

    return workbook
