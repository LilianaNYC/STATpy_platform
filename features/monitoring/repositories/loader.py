"""Persistence / source-data loading for the PD Performance Dash app.

This is a trimmed port of ``data_manager.py`` and the data-layer helpers in
``callbacks/monitoring_pd_models_callbacks.py`` from the original monitoring
dashboard, keeping only what the PD Performance tab needs:

- the portfolio extract (with ``_quarter`` / ``_snapshot_date`` derived columns)
- the PD/CRR/RAG-assignment threshold tables
- row-level PD performance observations (1y / 2y / NCO 1y)
- worst-grade rating-migration observations
- the dummy MEV catalog and rank-ordering facility data

Functions ``_build_model_rows``, ``_build_model_quarter_breakdown``,
``_build_model_segment_quarter_breakdown`` and ``_build_threshold_summary``
from the original ``build_monitoring_pd_models`` are intentionally NOT
ported -- they feed the model-overview tables on other tabs and are not
referenced by ``renderPdModels()``'s live PD Performance sections. The model
and segment filter option lists are instead derived directly from the
portfolio dataframe.

Feature-private -- only :mod:`features.monitoring.services.data_service`
reads from this module (plus one cross-feature read from
:mod:`features.saas.repositories.loader` as a best-effort MEV catalog
fallback). The ``Filters`` sheet reader lives in ``shared/repositories/`` instead
-- see that package's docstring for why.
"""

from __future__ import annotations

import logging
from typing import Any

import numpy as np
import openpyxl
import pandas as pd

from ....shared.domain import constants as config
from ....config.settings import settings
from ....shared.text import normalize_model_name as _normalize_model_name

log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Small shared helpers
# ---------------------------------------------------------------------------
def _label(ts: pd.Timestamp) -> str:
    """Convert a quarter-end timestamp to a label like '2025Q1'."""
    q = (ts.month - 1) // 3 + 1
    return f"{ts.year}Q{q}"


# ---------------------------------------------------------------------------
# Portfolio loading
# ---------------------------------------------------------------------------
def load_portfolio() -> pd.DataFrame:
    """Load the portfolio Excel file, clean null sentinels, derive quarter labels.

    The facility-level ``Portfolio`` sheet is no longer used by the PD
    Performance tab (its metrics are read directly from
    ``PD_Performance_Metrics``). If the sheet is absent, return an empty frame
    so the app still loads.
    """
    log.info("Loading portfolio from %s [%s]", settings.portfolio_file, config.PORTFOLIO_SHEET_NAME)

    try:
        df = pd.read_excel(settings.portfolio_file, sheet_name=config.PORTFOLIO_SHEET_NAME)
    except (ValueError, KeyError):
        log.info("Portfolio sheet '%s' not found; using empty portfolio.", config.PORTFOLIO_SHEET_NAME)
        return pd.DataFrame(columns=[config.DATE_COLUMN, "_quarter", "_snapshot_date"])

    df.replace(config.NULL_SENTINELS, pd.NA, inplace=True)

    df[config.DATE_COLUMN] = pd.to_datetime(df[config.DATE_COLUMN])
    df["_quarter"] = df[config.DATE_COLUMN].apply(_label)
    df["_snapshot_date"] = df[config.DATE_COLUMN].dt.date

    log.info("Loaded %d records across %d quarters", len(df), df["_quarter"].nunique())
    return df


# ---------------------------------------------------------------------------
# Monitoring thresholds
# ---------------------------------------------------------------------------
def _records(df: pd.DataFrame) -> list[dict[str, Any]]:
    """Convert a DataFrame to a list of JSON-safe records (NaN -> None)."""
    if df.empty:
        return []
    return df.where(pd.notna(df), None).to_dict(orient="records")


def load_monitoring_thresholds() -> dict[str, list[dict[str, Any]]]:
    """Load the PD / CRR-master-scale / RAG-assignment sheets.

    The original ``load_monitoring_thresholds`` only loads sheets whose
    config key ends with ``_thresholds_sheet_name`` or equals
    ``crr_master_scale_sheet_name`` -- which means ``RAG_Assignment_PD`` is
    never loaded by the existing pipeline. We load all three sheets the PD
    Performance tab needs explicitly here.
    """
    thresholds: dict[str, list[dict[str, Any]]] = {}

    for key, sheet_name in (
        ("pd_thresholds", config.PD_THRESHOLDS_SHEET_NAME),
        ("crr_master_scale", config.CRR_MASTER_SCALE_SHEET_NAME),
        ("rag_assignment_pd", config.RAG_ASSIGNMENT_PD_SHEET_NAME),
        ("lgd_thresholds", config.LGD_THRESHOLDS_SHEET_NAME),
        ("ead_thresholds", config.EAD_THRESHOLDS_SHEET_NAME),
        ("loss_thresholds", config.LOSS_THRESHOLDS_SHEET_NAME),
        ("scenario_test_thresholds", "Scenario_Test_Thresholds"),
        # Chapter-1 RAG-Assignment fallback rules, keyed by
        # (Model Type, Component, Test), giving each test's behaviour by
        # default count (< 15 vs >= 15): "Applicable" / "Non-Applicable" /
        # "Fallback Amber". Loaded dynamically so the rules can change in the
        # workbook without a code change (see resolve_pd_fallback_rule).
        ("fallback_amber_rules", "fallback_amber_rules"),
    ):
        try:
            df = pd.read_excel(settings.monitoring_thresholds_file, sheet_name=sheet_name)
            thresholds[key] = _records(df)
            log.info("Loaded thresholds sheet '%s' as %s (%d rows)", sheet_name, key, len(df))
        except Exception as exc:  # noqa: BLE001 - mirror original best-effort loading
            log.warning("Unable to load sheet '%s' for %s: %s", sheet_name, key, exc)
            thresholds[key] = []

    return thresholds


# ---------------------------------------------------------------------------
# MEV catalog
# ---------------------------------------------------------------------------
def _date_to_quarter_label(date_str: str) -> str:
    """Convert ``'MM/DD/YYYY'`` or a pandas Timestamp to ``'YYYY-QN'``."""
    ts = pd.Timestamp(date_str)
    return f"{ts.year}-Q{(ts.month - 1) // 3 + 1}"


def _compute_dev_range(
    values: list[float], development_date: str,
) -> dict[str, Any]:
    """Compute development-period statistics from *values* (all baseline
    observations up to and including *development_date*).
    """
    if not values:
        return {
            "min": None, "max": None, "mean": None,
            "2std_lower": None, "2std_upper": None,
            "development_date": development_date,
        }
    arr = np.array(values, dtype=float)
    mean = round(float(np.mean(arr)), 6)
    std = float(np.std(arr, ddof=1)) if len(arr) > 1 else 0.0
    return {
        "min": round(float(np.min(arr)), 6),
        "max": round(float(np.max(arr)), 6),
        "mean": mean,
        "2std_lower": round(mean - 2 * std, 6),
        "2std_upper": round(mean + 2 * std, 6),
        "development_date": development_date,
    }


def load_pd_mev_catalog() -> dict[str, Any]:
    """Load the MEV catalog from ``dummy_mev_data.xlsx``.

    Reads the model characteristics, transformed MEV descriptions, and
    baseline time-series data from the Excel workbook. Development-range
    statistics (``dev_range``) are computed from baseline observations up
    to each model's development date.
    """
    path = settings.dummy_mev_data_file
    try:
        xls = pd.ExcelFile(path)
    except FileNotFoundError:
        log.warning("MEV workbook not found: %s", path)
        return {}, {}, {}, {}

    # -- model_names: development dates, segments, descriptive names
    mc_df = pd.read_excel(
        xls,
        sheet_name=config.DUMMY_MEV_MODEL_CHARACTERISTIC_SHEET_NAME,
    ).dropna(how="all")

    dev_dates: dict[str, str] = {}
    descriptive_names: dict[str, str] = {}
    model_types: dict[str, str] = {}
    regions: dict[str, str] = {}
    portfolios: dict[str, str] = {}
    for _, row in mc_df.iterrows():
        model_key = str(row.get("Model Name", "")).strip()
        if not model_key:
            continue
        date_val = row.get("Development Date", "")
        if date_val:
            dev_dates[model_key] = _date_to_quarter_label(str(date_val))
        desc = row.get("Model Descriptive Name", "")
        if desc:
            descriptive_names[model_key] = str(desc).strip()
        model_type = str(row.get("Model Type", "")).strip().upper()
        if model_type:
            model_types[model_key] = model_type
        region_val = row.get("Region")
        if pd.notna(region_val) and str(region_val).strip():
            regions[model_key] = str(region_val).strip()
        portfolio_val = row.get("Portfolio")
        if pd.notna(portfolio_val) and str(portfolio_val).strip():
            portfolios[model_key] = str(portfolio_val).strip()

    # -- mev_transformed: model→segment and model→MEV mapping
    desc_df = pd.read_excel(
        xls,
        sheet_name=config.DUMMY_MEV_TRANSFORMED_DESCRIPTION_SHEET_NAME,
    ).dropna(how="all")

    model_segments: dict[str, list[str]] = {}
    mev_long_names: dict[str, str] = {}
    mev_descriptions: dict[str, str] = {}
    model_transformed_mevs: dict[str, set[str]] = {}
    # A model can own different MEVs per segment (e.g. PD Model B's Defensive
    # rows use HPI/UNEMP while its Cyclical rows use CRE_PRICE/OILPRICE/GDP) --
    # tracked per (model, mnemonic) so the catalog can tag each MEV with the
    # segment(s) it actually belongs to, instead of flattening every segment's
    # MEVs into one undifferentiated set per model.
    model_mev_segments: dict[str, dict[str, set[str]]] = {}
    model_mev_contributions: dict[str, dict[str, float]] = {}
    # Per-segment contributions, keyed by (model, segment, mnemonic) so the
    # Post-Scenario MEV Summary can show a specific segment's own weights instead
    # of the segment-collapsed, last-write-wins ``model_mev_contributions``.
    model_mev_segment_contributions: dict[str, dict[str, dict[str, float]]] = {}
    for _, row in desc_df.iterrows():
        model_key = str(row.get("Model Name", "")).strip()
        segment = str(row.get("Segment", "")).strip()
        mnemonic = str(row.get("US Mnemonic", "")).strip()
        long_name = str(row.get("Long Name", "")).strip()
        description = str(row.get("Description", "")).strip()
        contribution = row.get("Model Contribution")
        if model_key and segment:
            model_segments.setdefault(model_key, [])
            if segment not in model_segments[model_key]:
                model_segments[model_key].append(segment)
        if mnemonic and long_name:
            mev_long_names[mnemonic] = long_name
        if mnemonic and description:
            mev_descriptions[mnemonic] = description
        if model_key and mnemonic:
            model_transformed_mevs.setdefault(model_key, set()).add(mnemonic)
        if model_key and mnemonic and segment:
            model_mev_segments.setdefault(model_key, {}).setdefault(mnemonic, set()).add(segment)
        if model_key and mnemonic and contribution is not None:
            try:
                contribution_value = float(contribution)
            except (TypeError, ValueError):
                contribution_value = None
            if contribution_value is not None:
                model_mev_contributions.setdefault(model_key, {})[mnemonic] = contribution_value
                if segment:
                    model_mev_segment_contributions.setdefault(model_key, {}).setdefault(segment, {})[mnemonic] = contribution_value

    # -- scenario (all scenarios): time series per model+MEV+scenario
    ts_df = pd.read_excel(
        xls,
        sheet_name=config.DUMMY_MEV_TIME_SERIES_SHEET_NAME,
    ).dropna(how="all")
    ts_df["Date"] = pd.to_datetime(ts_df["Date"], errors="coerce")
    ts_df["Quarter"] = pd.to_numeric(ts_df.get("Quarter"), errors="coerce")
    ts_df["Run For"] = ts_df.get("Run For").map(lambda value: str(value).strip() if value is not None else "")
    ts_df["Scenario"] = ts_df.get("Scenario").map(lambda value: str(value).strip() if value is not None else "")
    ts_df["_quarter_label"] = ts_df["Date"].astype(str).map(_date_to_quarter_label)
    ts_df["MEV Value"] = pd.to_numeric(ts_df["MEV Value"], errors="coerce")
    ts_df = ts_df.dropna(subset=["Date", "MEV Value"])

    xls.close()

    # -- assemble the catalog keyed by descriptive model name
    all_model_keys = sorted(
        set(ts_df["Model Name"].dropna().unique())
        | set(desc_df["Model Name"].dropna().unique()),
    )

    mev_mnemonic_map: dict[str, str] = {
        long_name: mnemonic for mnemonic, long_name in mev_long_names.items()
    }
    mev_description_map: dict[str, str] = {
        mev_long_names.get(mnemonic, mnemonic): desc
        for mnemonic, desc in mev_descriptions.items()
        if mnemonic in mev_long_names
    }

    catalog: dict[str, Any] = {}
    for model_key in all_model_keys:
        model_name = _normalize_model_name(
            descriptive_names.get(model_key, model_key),
        )
        if not model_name:
            continue

        model_ts = ts_df[ts_df["Model Name"] == model_key]
        allowed = model_transformed_mevs.get(model_key, set())
        mev_names_for_model = sorted(
            name for name in model_ts["MEV Name"].unique() if name in allowed
        )
        dev_date = dev_dates.get(model_key, "")

        mevs: dict[str, Any] = {}
        for mev_name in mev_names_for_model:
            series_rows = model_ts[model_ts["MEV Name"] == mev_name].sort_values("_quarter_label")
            baseline_rows = series_rows[series_rows["Scenario"] == "baseline"]
            time_series = {
                row["_quarter_label"]: round(float(row["MEV Value"]), 6)
                for _, row in baseline_rows.iterrows()
            }
            scenario_series: dict[str, dict[str, float]] = {}
            for scenario in series_rows["Scenario"].dropna().unique():
                sc_rows = series_rows[series_rows["Scenario"] == scenario]
                scenario_series[str(scenario).strip()] = {
                    row["_quarter_label"]: round(float(row["MEV Value"]), 6)
                    for _, row in sc_rows.iterrows()
                }
            scenario_series_by_cycle: dict[str, dict[str, dict[str, float]]] = {}
            scenario_quarter_zero_by_cycle: dict[str, dict[str, str]] = {}
            for run_for in series_rows["Run For"].dropna().unique():
                cycle_rows = series_rows[series_rows["Run For"] == run_for]
                cycle_key = str(run_for).strip()
                if not cycle_key:
                    continue
                for scenario in cycle_rows["Scenario"].dropna().unique():
                    scenario_key = str(scenario).strip()
                    sc_rows = cycle_rows[cycle_rows["Scenario"] == scenario]
                    scenario_series_by_cycle.setdefault(cycle_key, {})[scenario_key] = {
                        row["_quarter_label"]: round(float(row["MEV Value"]), 6)
                        for _, row in sc_rows.iterrows()
                    }
                    q0_rows = sc_rows[sc_rows["Quarter"] == 0].sort_values("Date")
                    if not q0_rows.empty:
                        scenario_quarter_zero_by_cycle.setdefault(cycle_key, {})[scenario_key] = q0_rows.iloc[0]["_quarter_label"]
            dev_values = [
                v for q, v in time_series.items()
                if not dev_date or q <= dev_date
            ]
            display_name = mev_long_names.get(mev_name, mev_name)
            mevs[display_name] = {
                "dev_range": _compute_dev_range(dev_values, dev_date),
                "time_series": time_series,
                "scenario_series": scenario_series,
                "scenario_series_by_cycle": scenario_series_by_cycle,
                "scenario_quarter_zero_by_cycle": scenario_quarter_zero_by_cycle,
                "segments": sorted(model_mev_segments.get(model_key, {}).get(mev_name, set())),
            }

        contributions = {}
        raw_contribs = model_mev_contributions.get(model_key, {})
        for mnemonic, value in raw_contribs.items():
            display_name = mev_long_names.get(mnemonic, mnemonic)
            contributions[display_name] = value

        contributions_by_segment: dict[str, dict[str, float]] = {}
        for segment_key, seg_contribs in model_mev_segment_contributions.get(model_key, {}).items():
            seg_map = {}
            for mnemonic, value in seg_contribs.items():
                seg_map[mev_long_names.get(mnemonic, mnemonic)] = value
            contributions_by_segment[segment_key] = seg_map

        catalog[model_name] = {
            "model_type": model_types.get(model_key, ""),
            "region": regions.get(model_key, ""),
            "portfolio": portfolios.get(model_key, ""),
            "segments": model_segments.get(model_key, []),
            "severe_scenario_date": "",
            "mevs": mevs,
            "contributions": contributions,
            "contributions_by_segment": contributions_by_segment,
        }

    # -- Model Use Case / Cycle -> Scenario: the Scenario filter's options are
    # the distinct "Scenario" values available for the selected "Run For"
    # cycle in this same scenario sheet (not from any config list).
    scenarios_by_cycle: dict[str, list[str]] = {
        str(run_for).strip(): sorted({
            str(scenario).strip()
            for scenario in group["Scenario"].dropna().unique()
            if str(scenario).strip()
        })
        for run_for, group in ts_df.groupby("Run For")
        if str(run_for).strip()
    }

    return catalog, mev_mnemonic_map, mev_description_map, scenarios_by_cycle


# ---------------------------------------------------------------------------
# Aggregated-sheet loader
# ---------------------------------------------------------------------------

PD_AGGREGATED_SHEET_NAME = "PD_Performance_Metrics"
LGD_AGGREGATED_SHEET_NAME = "LGD_Performance_Metrics"
EAD_AGGREGATED_SHEET_NAME = "EAD_Performance_Metrics"
LOSS_AGGREGATED_SHEET_NAME = "Loss_Performance_Metrics"
PD_SENSITIVITY_SHEET_NAME = "PD_Sensitivity_Projections"
LGD_SENSITIVITY_SHEET_NAME = "LGD_Sensitivity_Projections"
EAD_SENSITIVITY_SHEET_NAME = "EAD_Sensitivity_Projections"

# Sheet column -> metric-row key consumed by each performance page/data module.
_LGD_METRIC_COLUMN_MAP = {
    "me": "ME",
    "rmse": "RMSE",
    "kendall_tau": "Kendall's Tau",
    "predicted_lgd": "Predicted LGD",
    "actual_lgd": "Actual LGD",
    "recovery_rate": "Recovery Rate",
    "population_stability_index": "Population Stability Index",
    "observations": "Observations",
    "defaults": "Defaults",
}
_EAD_METRIC_COLUMN_MAP = {
    "me": "ME",
    "rmse": "RMSE",
    "kendall_tau": "Kendall's Tau",
    "predicted_ead": "Predicted EAD",
    "actual_ead": "Actual EAD",
    "population_stability_index": "Population Stability Index",
    "observations": "Observations",
    "defaults": "Defaults",
}
_LOSS_METRIC_COLUMN_MAP = {
    "me": "ME",
    "me_pct": "ME %",
    "predicted_loss": "Predicted Loss",
    "actual_loss": "Actual Loss",
    "defaults": "Defaults",
    "balance": "Balance",
    "observations": "Observations",
    "nco_me": "NCO ME",
    "nco_me_pct": "NCO ME %",
    "nco_predicted": "Predicted NCO",
    "nco_actual": "Actual NCO",
    "acl_me": "ACL ME",
    "acl_me_pct": "ACL ME %",
    "acl_predicted": "Predicted ACL",
    "acl_actual": "Actual ACL",
}
_OPTIONAL_REVIEW_TEXT_COLUMNS = (
    "rag_post_sr",
    "rag_pre_mitig",
    "rag_post_mitig",
    "reviewer_commentary",
    "compensating_controls",  # see the matching comment on _PD_TEXT_COLUMNS in this module
    "scenario",  # see the matching comment on _PD_TEXT_COLUMNS in this module
)


def _build_metric_rows_store(sheet_name: str, column_map: dict[str, str]) -> dict[str, Any]:
    """Load a precomputed performance sheet into a per-cycle metric-row store.

    Returns ``{cycle: {"quarters": [...], "metrics_store": {(model, segment): [rows]}}}``
    where each row matches the metric-row shape the matching page expects.
    Values are taken verbatim from the sheet — no metric is recomputed.
    """
    try:
        df = pd.read_excel(settings.portfolio_file, sheet_name=sheet_name)
    except (FileNotFoundError, ValueError, KeyError):
        return {}
    df = df.dropna(how="all")
    if df.empty or "reporting_cycle" not in df.columns:
        return {}

    def _num(value):
        return float(value) if pd.notna(value) else None

    by_cycle: dict[str, Any] = {}
    for cycle in sorted(df["reporting_cycle"].dropna().unique()):
        cycle_df = df[df["reporting_cycle"] == cycle]
        store: dict = {}
        for _, row in cycle_df.iterrows():
            model = str(row.get("model", "")).strip()
            segment = str(row.get("segment", "")).strip()
            quarter = str(row.get("quarter", "")).strip()
            if not model or not segment or not quarter:
                continue
            metric_row = {"Monitoring Period": quarter}
            for col, key in column_map.items():
                metric_row[key] = _num(row.get(col)) if col in cycle_df.columns else None
            for col in _OPTIONAL_REVIEW_TEXT_COLUMNS:
                if col in cycle_df.columns and pd.notna(row.get(col)) and str(row.get(col)).strip():
                    metric_row[col] = str(row.get(col)).strip()
            for count_key in ("Observations", "Defaults"):
                if metric_row.get(count_key) is not None:
                    metric_row[count_key] = int(metric_row[count_key])
            store.setdefault((model, segment), []).append(metric_row)
        for key in store:
            store[key].sort(key=lambda r: r["Monitoring Period"])
        quarters = sorted(cycle_df["quarter"].dropna().astype(str).unique())
        by_cycle[cycle] = {"quarters": quarters, "metrics_store": store}
    return by_cycle


def load_lgd_performance_metrics() -> dict[str, Any]:
    """Load LGD metrics per reporting cycle from ``LGD_Performance_Metrics``."""
    return _build_metric_rows_store(LGD_AGGREGATED_SHEET_NAME, _LGD_METRIC_COLUMN_MAP)


def load_ead_performance_metrics() -> dict[str, Any]:
    """Load EAD metrics per reporting cycle from ``EAD_Performance_Metrics``."""
    return _build_metric_rows_store(EAD_AGGREGATED_SHEET_NAME, _EAD_METRIC_COLUMN_MAP)


def load_loss_performance_metrics() -> dict[str, Any]:
    """Load Loss metrics per reporting cycle from ``Loss_Performance_Metrics``."""
    return _build_metric_rows_store(LOSS_AGGREGATED_SHEET_NAME, _LOSS_METRIC_COLUMN_MAP)


def load_sensitivity_projections(sheet_name: str, value_col: str) -> list[dict[str, Any]]:
    """Load projected sensitivity rows from a ``*_Sensitivity_Projections`` sheet.

    ``value_col`` is the tab's projected-value column (``projected_pd`` /
    ``projected_lgd`` / ``projected_ead``). It is exposed verbatim under that key
    and also under the generic ``projected_value`` key, so the shared chart
    builders and Post-Subjective section helpers stay tab-agnostic.
    """
    try:
        df = pd.read_excel(settings.portfolio_file, sheet_name=sheet_name)
    except (FileNotFoundError, ValueError, KeyError):
        return []

    df = df.dropna(how="all")
    required = {
        "reporting_cycle",
        "model",
        "segment",
        "projection_quarter",
        "scenario_variant",
        value_col,
    }
    if "quarter" not in df.columns and "projection_offset" in df.columns:
        df = df.rename(columns={"projection_offset": "quarter"})
    required.add("quarter")
    if df.empty or not required.issubset(df.columns):
        return []

    df = df.copy()
    df["quarter"] = pd.to_numeric(df["quarter"], errors="coerce")
    df["projection_quarter"] = pd.to_datetime(df["projection_quarter"], errors="coerce")
    df[value_col] = pd.to_numeric(df[value_col], errors="coerce")
    df = df.dropna(subset=["quarter", "projection_quarter", value_col])

    # MM_P0 / MM_Pm are scenario-independent margins read verbatim from the
    # sheet: MM_P0 is a single value per entity, MM_Pm varies by projection
    # quarter. Both repeat across the scenario rows.
    def _opt_num(row, column):
        if column not in df.columns or pd.isna(row.get(column)):
            return None
        return round(float(row[column]), 8)

    records = []
    for _, row in df.iterrows():
        value = round(float(row[value_col]), 8)
        records.append({
            "reporting_cycle": str(row["reporting_cycle"]).strip(),
            "model": str(row["model"]).strip(),
            "segment": str(row["segment"]).strip(),
            "quarter": int(row["quarter"]),
            "projection_quarter": row["projection_quarter"].date().isoformat(),
            "scenario_variant": str(row["scenario_variant"]).strip(),
            "base_scenario": str(row.get("base_scenario") or "").strip(),
            "shock_std": float(row["shock_std"]) if "shock_std" in df.columns and pd.notna(row.get("shock_std")) else None,
            "shock_direction": str(row.get("shock_direction") or "").strip(),
            value_col: value,
            "projected_value": value,
            "mm_p0": _opt_num(row, "MM_P0"),
            "mm_pm": _opt_num(row, "MM_Pm"),
        })

    return records


def load_pd_sensitivity_projections() -> list[dict[str, Any]]:
    """Load projected PD sensitivity rows from ``PD_Sensitivity_Projections``."""
    return load_sensitivity_projections(PD_SENSITIVITY_SHEET_NAME, "projected_pd")


MONITORING_ACTIONS_SHEET_NAME = "monitoring_actions"

# Workbook column -> normalized record key for the governance action playbook.
_MONITORING_ACTIONS_COLUMN_MAP = {
    "Stage": "stage",
    "Trigger": "trigger",
    "RAG": "rag",
    "Description": "description",
    "Required Action": "required_action",
    "Additional Requirements": "additional_requirements",
    "Escalation / Discussion": "escalation",
    "Sponsor Approval Required": "sponsor_approval",
    "Deep Dive Required": "deep_dive",
    "Redevelopment Considered": "redevelopment",
    "Action Owner": "owner",
    "Due in Report": "due_in_report",
}


def load_monitoring_actions() -> list[dict[str, Any]]:
    """Load the governance action playbook from the ``monitoring_actions`` sheet
    of the monitoring-thresholds workbook.

    One row per ``(Stage, Trigger, RAG)``; blank cells become empty strings so
    the domain/UI layers never see NaN. Missing file or sheet -> ``[]`` (the
    Conclusion section then simply renders without a Required Actions panel).
    """
    try:
        df = pd.read_excel(settings.monitoring_thresholds_file, sheet_name=MONITORING_ACTIONS_SHEET_NAME)
    except (FileNotFoundError, ValueError, KeyError):
        log.warning(
            "Monitoring actions sheet '%s' not readable in %s; playbook disabled",
            MONITORING_ACTIONS_SHEET_NAME, settings.monitoring_thresholds_file,
        )
        return []

    df = df.dropna(how="all")
    records = []
    for _, row in df.iterrows():
        record = {
            key: str(row[col]).strip() if col in df.columns and pd.notna(row.get(col)) else ""
            for col, key in _MONITORING_ACTIONS_COLUMN_MAP.items()
        }
        if record["stage"] and record["rag"]:
            records.append(record)
    return records


def _build_observations_from_aggregated(agg_df: pd.DataFrame) -> tuple[dict, list[dict], list[str], list[dict]]:
    """Deprecated facility synthesis.

    PD Performance metrics are now read verbatim from the
    ``PD_Performance_Metrics`` tab via :func:`_build_metrics_store`; no
    facility-level data is synthesized or recomputed. Retained as a no-op so
    existing call sites keep working.
    """
    return config.PD_HORIZON_COLUMNS, [], [], []


# Metric columns read verbatim from the PD_Performance_Metrics tab.
_PD_METRIC_COLUMNS = (
    "confidence_interval_test",
    "notching_test_signed",
    "notching_test_abs",
    "observed_default_rate",
    "predicted_default_rate",
    "actual_expected_ratio",
    "accuracy_ratio",
    "go_live_accuracy_ratio",
    "delta_accuracy_ratio",
    "gini_coefficient",
    "ks_statistic",
    "kendall_tau",
    "brier_score",
    "population_stability_index",
    "rating_migration_index",
    "ead",
    "ead_share",
    "total_defaults",
    "default_count_1y",
)

_PD_TEXT_COLUMNS = (
    "rag_post_sr",
    "rag_pre_mitig",
    "rag_post_mitig",
    "reviewer_commentary",
    # The reviewer's compensating-controls justification for the Post
    # Mitigation RAG (the judgement/controls that move Pre-Mitigation ->
    # Post-Mitigation). Free text, saved/read exactly like reviewer_commentary
    # (self-healing column, see _update_review_flow_field). PD only for now;
    # LGD/EAD would add it to their own *_TEXT_COLUMNS when rolled out there.
    "compensating_controls",
    # Not a reviewer-facing RAG/commentary field -- the Scenario filter value
    # in effect the last time this row's review flow was saved, so Overview's
    # MEV Range can look up the scenario each row was actually reviewed under
    # instead of assuming a single portfolio-wide default (see
    # save_pd_review_flow_rag_changes).
    "scenario",
)

# The horizons each per-horizon row is replicated to when its ``horizon`` cell
# is left blank (i.e. the metric is horizon-agnostic, e.g. discrimination).
_PD_HORIZONS = ("1y", "2y", "nco_1y")


def _build_metrics_store(cycle_df: pd.DataFrame) -> dict:
    """Build the precomputed-metrics lookup the calculation engine reads from.

    The ``PD_Performance_Metrics`` tab is keyed by ``reporting_cycle × model ×
    segment × quarter × horizon``. The store is keyed by
    ``(model, segment, quarter, horizon)`` and every value is taken
    verbatim from the sheet. Rows whose ``horizon`` is blank carry
    horizon-agnostic metrics (discrimination, PSI, rating migration); they are
    merged into every horizon for that ``(model, segment, quarter)`` so the engine
    finds them regardless of which horizon it queries.
    """
    go_live_quarter = config.PD_GO_LIVE_QUARTER_END

    def _num(value):
        return float(value) if pd.notna(value) else None

    # Collect, per (model, segment, quarter), the blank-horizon agnostic metrics
    # and each specific horizon's metrics.
    grouped: dict = {}
    for _, row in cycle_df.iterrows():
        quarter = str(row["quarter"]).strip()
        model = str(row.get("model", "")).strip()
        segment = str(row.get("segment", "")).strip()
        horizon = str(row.get("horizon", "")).strip()
        if not quarter or not model or not segment:
            continue
        metrics = {col: _num(row.get(col)) for col in _PD_METRIC_COLUMNS if col in cycle_df.columns}
        total_defaults = metrics.get("total_defaults")
        legacy_default_count_1y = metrics.get("default_count_1y")
        if total_defaults is None and legacy_default_count_1y is not None:
            total_defaults = legacy_default_count_1y
            metrics["total_defaults"] = legacy_default_count_1y
        if horizon == "1y" and metrics.get("default_count_1y") is None and total_defaults is not None:
            metrics["default_count_1y"] = total_defaults
        metrics.update({
            col: str(row.get(col)).strip()
            for col in _PD_TEXT_COLUMNS
            if col in cycle_df.columns and pd.notna(row.get(col)) and str(row.get(col)).strip()
        })
        bucket = grouped.setdefault((model, segment, quarter), {"shared": {}, "horizons": {}})
        if horizon in ("", "nan", "all"):
            bucket["shared"].update({k: v for k, v in metrics.items() if v is not None})
        else:
            bucket["horizons"][horizon] = metrics

    store: dict = {}
    for (model, segment, quarter), bucket in grouped.items():
        shared = bucket["shared"]
        horizons = bucket["horizons"] or {h: {} for h in _PD_HORIZONS}
        for horizon, specific in horizons.items():
            merged = {**shared, **{k: v for k, v in specific.items() if v is not None}}
            merged.setdefault("go_live_quarter", go_live_quarter)
            store[(model, segment, quarter, horizon)] = merged

    return store


_LGD_TEXT_COLUMNS = (
    "rag_post_sr",
    "rag_pre_mitig",
    "rag_post_mitig",
    "reviewer_commentary",
    "compensating_controls",  # see the matching comment on _PD_TEXT_COLUMNS
    "scenario",  # see the matching comment on _PD_TEXT_COLUMNS
)

_EAD_TEXT_COLUMNS = (
    "rag_post_sr",
    "rag_pre_mitig",
    "rag_post_mitig",
    "reviewer_commentary",
    "compensating_controls",  # see the matching comment on _PD_TEXT_COLUMNS
    "scenario",  # see the matching comment on _PD_TEXT_COLUMNS
)


def _update_review_flow_field(
    sheet_name: str,
    valid_fields: tuple[str, ...],
    reporting_cycle: str, model: str, segment: str, quarter: str, field: str, new_value: str,
) -> int:
    """Write ``new_value`` for ``field`` into ``sheet_name`` of the portfolio file, in place.

    Shared by :func:`update_pd_review_flow_rag` and :func:`update_lgd_review_flow_rag`. Every row for a
    given ``(reporting_cycle, model, segment, quarter)`` carries the same value in these sheets
    (PD duplicates it across horizon rows; LGD has exactly one row per key), so all matching rows are
    updated together to keep the file internally consistent. Returns the number of rows written; 0 means
    no matching rows were found (nothing was written, so the file is untouched).
    """
    if field not in valid_fields:
        raise ValueError(f"Unknown review-flow RAG field: {field!r}")

    workbook = openpyxl.load_workbook(settings.portfolio_file)
    sheet = workbook[sheet_name]

    header_row = next(sheet.iter_rows(min_row=1, max_row=1))
    col_index = {cell.value: cell.column for cell in header_row if cell.value}
    base_columns = ("reporting_cycle", "quarter", "model", "segment")
    if any(name not in col_index for name in base_columns):
        log.warning("%s is missing a required column for the RAG write-back; skipped.", sheet_name)
        return 0

    if field not in col_index:
        # Self-healing schema: e.g. reviewer_commentary may not exist in the sheet yet on first use --
        # add the column header once instead of silently failing every save attempt.
        new_col = sheet.max_column + 1
        sheet.cell(row=1, column=new_col, value=field)
        col_index[field] = new_col
        log.info("Added missing column %r to %s in %s", field, sheet_name, settings.portfolio_file)

    def _cell_text(row, name: str) -> str:
        value = row[col_index[name] - 1].value
        return str(value).strip() if value is not None else ""

    updated = 0
    for row in sheet.iter_rows(min_row=2):
        if (
            _cell_text(row, "reporting_cycle") == reporting_cycle
            and _cell_text(row, "quarter") == quarter
            and _cell_text(row, "model") == model
            and _cell_text(row, "segment") == segment
        ):
            row[col_index[field] - 1].value = new_value
            updated += 1

    if updated:
        workbook.save(settings.portfolio_file)
        log.info(
            "Updated %s to %r for %s %s %s %s (%d row(s)) in %s [%s]",
            field, new_value, reporting_cycle, model, segment, quarter, updated, settings.portfolio_file, sheet_name,
        )
    return updated


def update_pd_review_flow_rag(
    reporting_cycle: str, model: str, segment: str, quarter: str, field: str, new_value: str,
) -> int:
    """Write ``new_value`` for ``field`` into the ``PD_Performance_Metrics`` sheet, in place.

    ``field`` is one of ``rag_post_sr`` / ``rag_pre_mitig`` / ``rag_post_mitig`` / ``reviewer_commentary`` / ``scenario``.
    """
    return _update_review_flow_field(
        PD_AGGREGATED_SHEET_NAME, _PD_TEXT_COLUMNS,
        reporting_cycle, model, segment, quarter, field, new_value,
    )


def update_lgd_review_flow_rag(
    reporting_cycle: str, model: str, segment: str, quarter: str, field: str, new_value: str,
) -> int:
    """Write ``new_value`` for ``field`` into the ``LGD_Performance_Metrics`` sheet, in place.

    ``field`` is one of ``rag_post_sr`` / ``rag_pre_mitig`` / ``rag_post_mitig`` / ``reviewer_commentary`` / ``scenario``.
    """
    return _update_review_flow_field(
        LGD_AGGREGATED_SHEET_NAME, _LGD_TEXT_COLUMNS,
        reporting_cycle, model, segment, quarter, field, new_value,
    )


def update_ead_review_flow_rag(
    reporting_cycle: str, model: str, segment: str, quarter: str, field: str, new_value: str,
) -> int:
    """Write ``new_value`` for ``field`` into the ``EAD_Performance_Metrics`` sheet, in place.

    ``field`` is one of ``rag_post_sr`` / ``rag_pre_mitig`` / ``rag_post_mitig`` / ``reviewer_commentary`` / ``scenario``.
    """
    return _update_review_flow_field(
        EAD_AGGREGATED_SHEET_NAME, _EAD_TEXT_COLUMNS,
        reporting_cycle, model, segment, quarter, field, new_value,
    )


def _build_model_segment_cycle_map(agg_df: pd.DataFrame) -> dict[tuple[str, str], list[str]]:
    """Which reporting cycles have real data for each (model, segment) pair.

    Some models have less cycle history for specific segments than for their
    own "All" aggregate row (e.g. PD Model B's Cyclical/Defensive breakdown
    only starts at CCAR 2025, while its All row also has BAU 2025Q1) -- this
    lets the Model Use Case / Cycle filter narrow to what the currently
    selected model/segment population actually has, instead of always
    showing the full global cycle list.
    """
    result: dict[tuple[str, str], list[str]] = {}
    grouped = agg_df.groupby([
        agg_df["model"].map(_normalize_model_name),
        agg_df["segment"].astype(str).str.strip(),
    ])
    for (model, segment), group in grouped:
        if not model or not segment:
            continue
        result[(model, segment)] = sorted(group["reporting_cycle"].dropna().unique())
    return result


def _build_model_segment_cycle_map_from_sheet(sheet_name: str) -> dict[tuple[str, str], list[str]]:
    """Same as :func:`_build_model_segment_cycle_map`, reading ``sheet_name`` directly.

    Used for LGD/EAD, whose sheet rows use model names verbatim (no
    normalization needed, unlike PD's).
    """
    try:
        df = pd.read_excel(settings.portfolio_file, sheet_name=sheet_name)
    except (FileNotFoundError, ValueError, KeyError):
        return {}
    df = df.dropna(how="all")
    if df.empty or "reporting_cycle" not in df.columns:
        return {}
    result: dict[tuple[str, str], list[str]] = {}
    grouped = df.groupby([
        df["model"].astype(str).str.strip(),
        df["segment"].astype(str).str.strip(),
    ])
    for (model, segment), group in grouped:
        if not model or not segment:
            continue
        result[(model, segment)] = sorted(group["reporting_cycle"].dropna().unique())
    return result


def _build_model_segment_map(agg_df: pd.DataFrame) -> dict[str, list[str]]:
    """Which real segments each model actually owns.

    More than one model can cover the same segment name (e.g. both PD Model A
    and PD Model D have Cyclical rows) -- this lets the PD Performance tab
    narrow its Segment filter to whatever the currently-selected model has,
    instead of offering every segment in the portfolio regardless of model.
    """
    return {
        model: sorted({
            s for s in group["segment"].dropna().astype(str).str.strip().unique()
            if s and s.lower() != "all"
        })
        for model, group in agg_df.groupby(agg_df["model"].map(_normalize_model_name))
    }


def _build_model_segment_map_from_sheet(sheet_name: str) -> dict[str, list[str]]:
    """Same as :func:`_build_model_segment_map`, reading ``sheet_name`` directly.

    Used for LGD/EAD, whose sheet rows use model names verbatim (no
    normalization needed, unlike PD's).
    """
    try:
        df = pd.read_excel(settings.portfolio_file, sheet_name=sheet_name)
    except (FileNotFoundError, ValueError, KeyError):
        return {}
    df = df.dropna(how="all")
    if df.empty or "model" not in df.columns or "segment" not in df.columns:
        return {}
    return {
        model: sorted({
            s for s in group["segment"].dropna().astype(str).str.strip().unique()
            if s and s.lower() != "all"
        })
        for model, group in df.groupby(df["model"].astype(str).str.strip())
    }


def _invert_str_list_map(mapping: dict[str, list[str]]) -> dict[str, list[str]]:
    """Invert a ``key -> [values]`` map into ``value -> [keys]``.

    Used to derive a segment -> models map from each tab's model -> segments
    map, so the Segment filter can narrow the Model filter's options.
    """
    result: dict[str, list[str]] = {}
    for key, values in mapping.items():
        for value in values:
            result.setdefault(value, []).append(key)
    return result


def load_pd_performance_data_from_aggregated() -> dict[str, Any]:
    """Load from the PD_Performance_Metrics sheet instead of facility-level data."""
    log.info("Loading PD aggregated metrics from %s [%s]", settings.portfolio_file, PD_AGGREGATED_SHEET_NAME)

    agg_df = pd.read_excel(settings.portfolio_file, sheet_name=PD_AGGREGATED_SHEET_NAME)
    agg_df = agg_df.dropna(how="all")

    portfolio = load_portfolio()

    quarters = sorted(agg_df["quarter"].dropna().unique())
    latest_quarter = quarters[-1] if quarters else ""
    previous_quarter = quarters[-2] if len(quarters) > 1 else ""

    # Every row now names both a model and a segment ("All" for the model's
    # aggregate row). We still ignore any legacy "All Models" rows if they
    # exist in older files.
    data_model_names = sorted({
        m for m in agg_df["model"].dropna().map(_normalize_model_name).unique()
        if m and m != "All Models"
    })
    data_segment_values = sorted({
        s for s in agg_df["segment"].dropna().astype(str).str.strip().unique()
        if s and s.lower() != "all"
    })
    model_names = data_model_names or ["PD Model A", "PD Model B"]
    segment_values = data_segment_values or ["Cyclical", "Defensive", "O&M", "LoL", "IVB"]
    pd_model_segments = _build_model_segment_map(agg_df)
    pd_model_segment_cycles = _build_model_segment_cycle_map(agg_df)
    pd_segment_models = _invert_str_list_map(pd_model_segments)
    lgd_model_segments = _build_model_segment_map_from_sheet(LGD_AGGREGATED_SHEET_NAME)
    ead_model_segments = _build_model_segment_map_from_sheet(EAD_AGGREGATED_SHEET_NAME)
    loss_model_segments = _build_model_segment_map_from_sheet(LOSS_AGGREGATED_SHEET_NAME)

    monitoring_thresholds = load_monitoring_thresholds()

    # Build observations per reporting cycle
    reporting_cycles = sorted(agg_df["reporting_cycle"].dropna().unique()) if "reporting_cycle" in agg_df.columns else []
    observations_by_cycle = {}
    for cycle in reporting_cycles:
        cycle_df = agg_df[agg_df["reporting_cycle"] == cycle]
        obs = _build_observations_from_aggregated(cycle_df)
        cycle_quarters = sorted(cycle_df["quarter"].dropna().unique())
        observations_by_cycle[cycle] = {
            "performance_horizons": obs[0],
            "performance_observations": obs[1],
            "rating_values": obs[2],
            "rating_migration_observations": obs[3],
            "quarters": cycle_quarters,
            "metrics_store": _build_metrics_store(cycle_df),
        }

    # Default to first available cycle for backwards compatibility
    default_cycle = reporting_cycles[0] if reporting_cycles else None
    if default_cycle and default_cycle in observations_by_cycle:
        default_obs = observations_by_cycle[default_cycle]
        performance_horizons = default_obs["performance_horizons"]
        performance_observations = default_obs["performance_observations"]
        rating_values = default_obs["rating_values"]
        rating_migration_observations = default_obs["rating_migration_observations"]
    else:
        performance_horizons, performance_observations, rating_values, rating_migration_observations = (
            _build_observations_from_aggregated(agg_df)
        )

    mev_catalog, mev_mnemonic_map, mev_description_map, mev_scenarios_by_cycle = load_pd_mev_catalog()

    return {
        "portfolio": portfolio,
        "quarters": quarters,
        "latest_quarter": latest_quarter,
        "previous_quarter": previous_quarter,
        "latest_snapshot_date": latest_quarter,
        "previous_snapshot_date": previous_quarter,
        "source_file": settings.portfolio_file.name,
        "model_names": model_names,
        "segment_values": segment_values,
        "pd_model_segments": pd_model_segments,
        "pd_model_segment_cycles": pd_model_segment_cycles,
        "pd_segment_models": pd_segment_models,
        "lgd_model_segments": lgd_model_segments,
        "lgd_segment_models": _invert_str_list_map(lgd_model_segments),
        "ead_model_segments": ead_model_segments,
        "ead_segment_models": _invert_str_list_map(ead_model_segments),
        "loss_model_segments": loss_model_segments,
        "loss_segment_models": _invert_str_list_map(loss_model_segments),
        "lgd_model_segment_cycles": _build_model_segment_cycle_map_from_sheet(LGD_AGGREGATED_SHEET_NAME),
        "ead_model_segment_cycles": _build_model_segment_cycle_map_from_sheet(EAD_AGGREGATED_SHEET_NAME),
        "mev_scenarios_by_cycle": mev_scenarios_by_cycle,
        "monitoring_thresholds": monitoring_thresholds,
        "performance_horizons": performance_horizons,
        "performance_observations": performance_observations,
        "rating_values": rating_values,
        "rating_migration_observations": rating_migration_observations,
        "observations_by_cycle": observations_by_cycle,
        "reporting_cycles": reporting_cycles,
        "mev_catalog": mev_catalog,
        "mev_mnemonic_map": mev_mnemonic_map,
        "mev_description_map": mev_description_map,
        "sensitivity_projections": load_pd_sensitivity_projections(),
        "monitoring_actions": load_monitoring_actions(),
        "lgd_sensitivity_projections": load_sensitivity_projections(LGD_SENSITIVITY_SHEET_NAME, "projected_lgd"),
        "ead_sensitivity_projections": load_sensitivity_projections(EAD_SENSITIVITY_SHEET_NAME, "projected_ead"),
        "rank_ordering_facilities": {},
        "lgd_observations_by_cycle": load_lgd_performance_metrics(),
        "ead_observations_by_cycle": load_ead_performance_metrics(),
        "loss_observations_by_cycle": load_loss_performance_metrics(),
    }
