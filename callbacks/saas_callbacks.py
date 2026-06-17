"""Callbacks for the SAAS workspace."""

from __future__ import annotations

from datetime import datetime
from html import escape as html_escape
import math
import re
import statistics

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import CharacterProperties, Paragraph, ParagraphProperties, RichTextProperties
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from dash import ALL, MATCH, Input, Output, State, ctx, dcc, html, no_update

from ..components import filters as shared_filters
from ..components.charts import (
    SAAS_SCENARIO_COLOR_MAP,
    SAAS_SCENARIO_LABEL_MAP,
    build_saas_mev_time_series_figure,
    compute_saas_monitoring_band_spec,
)
from ..data.mev import format_pd_mev_value
from ..data.transformations import _finite
from ..data_store import SAAS_PAGE_DATA
from ..pages import saas_layout as layout

_GRAPH_CONFIG = {"displayModeBar": False, "responsive": True}
_RANGE_PRESET_COUNTS = {"last-4": 4, "last-8": 8, "last-12": 12}
_RAW_MEV_NAMES = set(SAAS_PAGE_DATA.get("raw_mev_names") or set())
_TRANSFORMED_MEV_NAMES = set(SAAS_PAGE_DATA.get("transformed_mev_names") or set())


def _is_segment_active(segment: str | None) -> bool:
    return bool(segment) and segment != layout.SEGMENT_ALL_VALUE


def _normalize_selected_run_fors(value) -> list[str]:
    selected_values = _normalize_multi_values(value)
    valid_values = [option["value"] for option in layout.RUN_FOR_OPTIONS]
    return [value for value in valid_values if value in selected_values]


def _normalize_compare_against_values(value, selected_run_for: str | None = None) -> list[str]:
    selected_values = _normalize_multi_values(value)
    valid_values = [
        option["value"]
        for option in layout.RUN_FOR_OPTIONS
        if option["value"] and option["value"] != selected_run_for
    ]
    compare_values = [item for item in selected_values if item in valid_values]
    if not compare_values:
        return [layout.COMPARE_AGAINST_NONE_VALUE]
    return compare_values


def _scoped_run_for_values(run_for, compare_against) -> list[str]:
    primary_values = _normalize_selected_run_fors(run_for)
    primary_value = primary_values[0] if primary_values else None
    compare_values = _normalize_compare_against_values(compare_against, primary_value)
    scoped_values = list(primary_values)
    for compare_value in compare_values:
        if compare_value == layout.COMPARE_AGAINST_NONE_VALUE or compare_value in scoped_values:
            continue
        scoped_values.append(compare_value)
    return scoped_values


def _model_names_for_filters(segment: str | None) -> list[str]:
    base_models = list(SAAS_PAGE_DATA.get("model_names", []))
    model_segments = SAAS_PAGE_DATA.get("model_segments", {})
    if _is_segment_active(segment):
        base_models = [model_name for model_name in base_models if model_segments.get(model_name) == segment]
    return base_models


def _model_options_for_filters(segment: str | None, *, disabled: bool = False) -> list[dict]:
    values = _model_names_for_filters(segment)
    seen: set[str] = set()
    options: list[dict] = []
    for value in values:
        if not value or value in seen:
            continue
        seen.add(value)
        option = {"label": value, "value": value}
        if disabled:
            option["disabled"] = True
        options.append(option)
    return options


def _normalize_selected_models(value) -> list[str]:
    if isinstance(value, str):
        return [value] if value else []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    return []


def _normalize_multi_values(value) -> list[str]:
    if isinstance(value, str):
        return [value] if value else []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    return []


def _normalize_selected_mevs(value) -> list[str]:
    return _normalize_multi_values(value)


def _normalize_selected_mev_mode(value: str | None, options: list[dict] | None = None) -> str:
    raw_value = str(value or "").strip().lower()
    valid_values = {
        str(option.get("value") or "").strip().lower()
        for option in (options or layout.MEV_TYPE_OPTIONS)
        if str(option.get("value") or "").strip()
    }
    if raw_value in valid_values:
        return raw_value
    return layout.DEFAULT_MEV_TYPE


def _normalize_selected_scenarios(value, options: list[dict] | None = None) -> list[str]:
    if isinstance(value, str):
        selected_values = [value] if value else []
    elif isinstance(value, list):
        selected_values = [str(item).strip().lower() for item in value if str(item).strip()]
    else:
        selected_values = []
    if options is None:
        return selected_values
    valid_values = {
        str(option.get("value") or "").strip().lower()
        for option in options
        if str(option.get("value") or "").strip()
    }
    return [selected_value for selected_value in selected_values if selected_value in valid_values]


def _coerce_quarter(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _normalize_snapshot_period(value: str | None) -> str:
    valid_values = {option["value"] for option in layout.SUBNAV_VIEW_OPTIONS}
    if value in valid_values:
        return value
    return layout.DEFAULT_SUBNAV_VIEW


def _normalize_mev_label_mode(value: str | None) -> str:
    valid_values = {option["value"] for option in layout.MEV_LABEL_MODE_OPTIONS}
    if value in valid_values:
        return value
    return layout.DEFAULT_MEV_LABEL_MODE


def _date_period_key(value) -> str:
    if value is None:
        return ""
    if hasattr(value, "date"):
        try:
            return value.date().isoformat()
        except TypeError:
            pass
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except TypeError:
            pass
    return str(value)


def _available_date_periods(records: list[dict]) -> list[str]:
    return sorted({
        _date_period_key(row.get("Date"))
        for row in records
        if _date_period_key(row.get("Date"))
    })


def _normalize_date_range(range_value, periods: list[str]) -> dict[str, str]:
    range_value = range_value or {}
    range_from = range_value.get("from", "")
    range_to = range_value.get("to", "")
    return {
        "from": range_from if range_from in periods else "",
        "to": range_to if range_to in periods else "",
    }


def _filter_records_by_snapshot_period(records: list[dict], snapshot_period: str | None) -> list[dict]:
    snapshot_period_value = _normalize_snapshot_period(snapshot_period)
    if snapshot_period_value == layout.DEFAULT_SUBNAV_VIEW:
        return list(records)

    filtered_records: list[dict] = []
    for row in records:
        quarter_value = _coerce_quarter(row.get("Quarter"))
        if quarter_value is None:
            continue
        if snapshot_period_value == "history" and quarter_value <= 0:
            filtered_records.append(row)
        elif snapshot_period_value == "projection" and quarter_value >= 0:
            filtered_records.append(row)
    return filtered_records


def _filter_records_by_date_range(records: list[dict], range_value) -> list[dict]:
    periods = _available_date_periods(records)
    selection = _normalize_date_range(range_value, periods)
    if not selection["from"] and not selection["to"]:
        return list(records)

    filtered_records: list[dict] = []
    for row in records:
        period = _date_period_key(row.get("Date"))
        if not period:
            continue
        if selection["from"] and period < selection["from"]:
            continue
        if selection["to"] and period > selection["to"]:
            continue
        filtered_records.append(row)
    return filtered_records


def _resolve_date_range_selection(periods: list[str], window_value: str | None, from_value: str | None, to_value: str | None, triggered_id) -> dict[str, str]:
    current = _normalize_date_range({"from": from_value or "", "to": to_value or ""}, periods)

    if isinstance(triggered_id, dict) and triggered_id.get("type") == layout.MODEL_DATE_RANGE_WINDOW_TYPE:
        if window_value == "all":
            return {"from": "", "to": ""}
        count = _RANGE_PRESET_COUNTS.get(window_value or "")
        if not count or not periods:
            return current
        return {"from": periods[max(0, len(periods) - count)], "to": periods[-1]}

    if current["from"] and current["to"] and current["from"] > current["to"]:
        if isinstance(triggered_id, dict) and triggered_id.get("type") == layout.MODEL_DATE_RANGE_FROM_TYPE:
            current["to"] = current["from"]
        elif isinstance(triggered_id, dict) and triggered_id.get("type") == layout.MODEL_DATE_RANGE_TO_TYPE:
            current["from"] = current["to"]
    return current


def _effective_model_names(segment: str | None, selected_models) -> list[str]:
    selected_model_values = _normalize_selected_models(selected_models)
    all_model_values = [option["value"] for option in _model_options_for_filters(None)]
    all_models_selected = bool(all_model_values) and set(selected_model_values) == set(all_model_values)

    if _is_segment_active(segment):
        return _model_names_for_filters(segment)
    if selected_model_values and not all_models_selected:
        selected_value_set = set(selected_model_values)
        return [value for value in all_model_values if value in selected_value_set]
    if all_models_selected:
        return all_model_values
    return []


def _run_for_meta_label(selected_run_fors) -> str:
    normalized_values = _normalize_selected_run_fors(selected_run_fors)
    if not normalized_values:
        return "No Reporting Cycle selected"
    all_values = [option["value"] for option in layout.RUN_FOR_OPTIONS]
    if all_values and set(normalized_values) == set(all_values):
        return "All Reporting Cycle values"
    if len(normalized_values) == 1:
        return normalized_values[0]
    return ", ".join(normalized_values)


def _build_compare_against_options(selected_run_for: str | None) -> list[dict]:
    options = [{"label": "None", "value": layout.COMPARE_AGAINST_NONE_VALUE}]
    for option in layout.RUN_FOR_OPTIONS:
        value = option.get("value")
        if not value or value == selected_run_for:
            continue
        options.append({"label": option.get("label", value), "value": value})
    return options


def _compare_against_toggle_label(selected_values, selected_run_for: str | None) -> str:
    normalized_values = _normalize_compare_against_values(selected_values, selected_run_for)
    compare_values = [value for value in normalized_values if value != layout.COMPARE_AGAINST_NONE_VALUE]
    if not compare_values:
        return "None"
    if len(compare_values) == 1:
        return compare_values[0]
    return f"{len(compare_values)} Reporting Cycle values selected"


def _primary_run_for_value(run_for) -> str | None:
    selected_values = _normalize_selected_run_fors(run_for)
    return selected_values[0] if selected_values else None


def _model_development_date(model_name: str, run_for) -> datetime | None:
    primary_run_for = _primary_run_for_value(run_for)
    if not primary_run_for:
        return None
    return SAAS_PAGE_DATA.get("model_development_dates", {}).get(primary_run_for, {}).get(model_name)


def _current_date_for_run_for(run_for) -> datetime | None:
    primary_run_for = _primary_run_for_value(run_for)
    if not primary_run_for:
        return None
    match = re.search(r"(\d{4})$", primary_run_for)
    if not match:
        return None
    return datetime(int(match.group(1)) - 1, 12, 31)


def _projection_start_date_for_run_for(run_for):
    """Return the date where Quarter == 0 for the primary Reporting Cycle."""
    primary_run_for = _primary_run_for_value(run_for)
    if not primary_run_for:
        return None
    return SAAS_PAGE_DATA.get("run_for_quarter_zero_dates", {}).get(primary_run_for)


def _format_monitoring_date(date_value) -> str:
    if date_value is None:
        return "—"
    if hasattr(date_value, "year") and hasattr(date_value, "month"):
        year = int(date_value.year)
        month = int(date_value.month)
    else:
        text = str(date_value).strip()
        parsed_date = None
        for candidate in (text[:10], text):
            try:
                parsed_date = datetime.fromisoformat(candidate)
                break
            except ValueError:
                continue
        if parsed_date is None:
            return text
        year = parsed_date.year
        month = parsed_date.month
    quarter = ((month - 1) // 3) + 1
    return f"{year}Q{quarter}"


def _build_monitoring_threshold_chips(band_spec: dict | None) -> list[html.Span]:
    if not band_spec:
        return []

    def chip(label, value, tone):
        return html.Span(
            [html.Strong(label), value],
            className=f"pd-mev-threshold-chip pd-mev-threshold-chip-{tone}",
        )

    return [
        chip(
            "Green",
            f"{format_pd_mev_value(band_spec['green_low'])} to {format_pd_mev_value(band_spec['green_high'])}",
            "green",
        ),
        chip(
            "Amber low",
            f"{format_pd_mev_value(band_spec['amber_low_low'])} to {format_pd_mev_value(band_spec['amber_low_high'])}",
            "amber",
        ),
        chip(
            "Amber high",
            f"{format_pd_mev_value(band_spec['amber_high_low'])} to {format_pd_mev_value(band_spec['amber_high_high'])}",
            "amber",
        ),
        chip(
            "Red",
            f"< {format_pd_mev_value(band_spec['red_low_cutoff'])} or > {format_pd_mev_value(band_spec['red_high_cutoff'])}",
            "red",
        ),
    ]


def _monitoring_marker_legend_item(
    label: str,
    value_text: str,
    tone: str,
    *,
    line_color: str | None = None,
    line_dash: str | None = None,
) -> html.Div:
    line_style = {}
    if line_color:
        line_style["borderTopColor"] = line_color
    if line_dash:
        line_style["borderTopStyle"] = line_dash
    return html.Div(
        className=f"pd-mev-marker-legend-item pd-mev-marker-legend-item-{tone}",
        children=[
            html.Span(
                className=f"pd-mev-marker-legend-line pd-mev-marker-legend-line-{tone}",
                style=line_style or None,
                **{"aria-hidden": "true"},
            ),
            html.Span(
                className="pd-mev-marker-legend-copy",
                children=[
                    html.Span(label, className="pd-mev-marker-legend-label"),
                    html.Span(value_text, className="pd-mev-marker-legend-date"),
                ],
            ),
        ],
    )


def _monitoring_scenario_legend_item(selected_scenarios) -> html.Div | None:
    normalized_scenarios = _normalize_selected_scenarios(selected_scenarios)
    if not normalized_scenarios:
        return None
    scenario_value = normalized_scenarios[0]
    scenario_label = SAAS_SCENARIO_LABEL_MAP.get(scenario_value, scenario_value.replace("_", " ").title())
    scenario_color = SAAS_SCENARIO_COLOR_MAP.get(scenario_value, "#475569")
    return _monitoring_marker_legend_item(
        "Scenario",
        scenario_label,
        "series",
        line_color=scenario_color,
        line_dash="solid" if scenario_value == "baseline" else "dashed",
    )


def _build_monitoring_summary(
    records: list[dict],
    selected_mev_mode,
    selected_scenarios,
    selected_mevs,
    primary_run_for: str | None,
    development_date,
    current_date,
):
    monitoring_records = _filter_records_by_mevs(
        _filter_records_by_scenarios(
            records,
            selected_scenarios,
        ),
        selected_mevs,
    )
    band_spec = compute_saas_monitoring_band_spec(
        monitoring_records,
        primary_run_for=primary_run_for,
        development_date=development_date,
    )

    threshold_items = _build_monitoring_threshold_chips(band_spec)
    marker_items = []
    scenario_item = _monitoring_scenario_legend_item(selected_scenarios)
    if scenario_item is not None:
        marker_items.append(scenario_item)
    if development_date is not None:
        marker_items.append(
            _monitoring_marker_legend_item(
                "Development",
                _format_monitoring_date(development_date),
                "development",
            )
        )
    if current_date is not None:
        marker_items.append(
            _monitoring_marker_legend_item(
                "Current date",
                _format_monitoring_date(current_date),
                "current",
            )
        )

    summary_rows = []
    if threshold_items:
        summary_rows.append(html.Div(threshold_items, className="pd-mev-monitoring-summary-row pd-mev-monitoring-summary-row-thresholds"))
    if marker_items:
        summary_rows.append(html.Div(marker_items, className="pd-mev-monitoring-summary-row pd-mev-monitoring-summary-row-markers"))

    if not summary_rows:
        return None
    return html.Div(summary_rows, className="pd-mev-monitoring-summary", **{"aria-label": "Monitoring summary"})


def _model_toggle_label(selected_models: list[str], all_options: list[dict], segment_active: bool) -> str:
    if segment_active:
        return "Disabled while Segment is selected"
    if not selected_models:
        return "Select models"
    all_values = [option["value"] for option in all_options]
    if all_values and set(selected_models) == set(all_values):
        return "All"
    if len(selected_models) == 1:
        return selected_models[0]
    return f"{len(selected_models)} models selected"


def _single_select_option_classes(value: str | None, option_ids: list[dict]) -> list[str]:
    return [
        "single-select-option is-selected" if option_id["value"] == value else "single-select-option"
        for option_id in option_ids
    ]


def _toggle_menu_class(current_class: str | None, *, base_class: str) -> str:
    if "open" in (current_class or "").split():
        return base_class
    return f"{base_class} open"


def _pluralize(count: int, label: str) -> str:
    return f"{count} {label}" if count == 1 else f"{count} {label}s"


def _format_month_year(value) -> str | None:
    if value is None or not hasattr(value, "strftime"):
        return None
    return value.strftime("%b %Y")




def _build_empty_state(title: str, copy: str):
    return html.Div(
        className="section-card pd-mev-empty-state",
        children=[
            html.Div(title, className="pd-mev-chart-title"),
            html.P(copy, className="pd-section-subtitle"),
        ],
    )


def _model_panel_id(model_name: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", model_name.strip().lower()).strip("-")
    return f"saas-model-panel-{slug}"


def _build_subnav_models(segment: str | None, selected_models):
    effective_models = _effective_model_names(segment, selected_models)

    if not effective_models:
        return []

    return [
        html.Div(
            [
                html.Button(
                    model_name,
                    type="button",
                    className="saas-subnav-model-chip",
                    **{"data-saas-scroll-target": _model_panel_id(model_name)},
                )
                for model_name in effective_models
            ],
            className="saas-subnav-model-list",
        ),
    ]


def _filter_records_by_scenarios(records: list[dict], selected_scenarios) -> list[dict]:
    normalized_scenarios = _normalize_selected_scenarios(selected_scenarios)
    if not normalized_scenarios:
        return []
    if layout.DEFAULT_SCENARIO_FILTER in normalized_scenarios:
        return records
    selected_scenario_set = set(normalized_scenarios)
    return [row for row in records if str(row.get("Scenario") or "").strip().lower() in selected_scenario_set]


def _mev_types_for_name(mev_name: str) -> set[str]:
    normalized_name = str(mev_name or "").strip()
    mev_types: set[str] = set()
    if normalized_name in _TRANSFORMED_MEV_NAMES:
        mev_types.add("transformed")
    if normalized_name in _RAW_MEV_NAMES:
        mev_types.add("raw")
    return mev_types


def _mev_picker_label(selected_mev_mode: str | None) -> str:
    normalized_mode = _normalize_selected_mev_mode(selected_mev_mode)
    if normalized_mode == "family":
        return "MEV Family"
    if normalized_mode == "transformed_only":
        return "Transformed MEVs"
    return "Raw MEVs"


def _mev_picker_empty_label(selected_mev_mode: str | None) -> str:
    normalized_mode = _normalize_selected_mev_mode(selected_mev_mode)
    if normalized_mode == "family":
        return "Select transformed MEV"
    if normalized_mode == "transformed_only":
        return "Select transformed MEVs"
    return "Select raw MEVs"


def _mev_picker_section_class(*, visible: bool) -> str:
    return "saas-mev-picker-section" if visible else "saas-mev-picker-section is-hidden"


def _family_map_for_model(model_name: str) -> dict[str, list[str]]:
    return SAAS_PAGE_DATA.get("model_mev_family_map", {}).get(model_name, {})


def _allowed_mev_names_for_model(model_name: str, selected_mev_mode: str | None = None) -> set[str]:
    model_mev_map = SAAS_PAGE_DATA.get("model_mev_map", {})
    model_entry = model_mev_map.get(model_name, {})
    normalized_mode = _normalize_selected_mev_mode(selected_mev_mode)
    if normalized_mode in {"family", "transformed_only"}:
        return set(model_entry.get("transformed", []))
    return set(model_entry.get("raw", []))


def _filter_records_by_model_mevs(records: list[dict], model_name: str, selected_mev_mode: str | None = None) -> list[dict]:
    allowed_mev_names = _allowed_mev_names_for_model(model_name, selected_mev_mode)
    if not allowed_mev_names:
        return []
    return [
        row for row in records
        if str(row.get("MEV Name") or "").strip() in allowed_mev_names
    ]


def _format_scenario_label(value: str) -> str:
    if value == layout.DEFAULT_SCENARIO_FILTER:
        return "All"
    return value.replace("_", " ").title()


def _build_scenario_options(records: list[dict]) -> list[dict]:
    scenario_values: list[str] = []
    seen: set[str] = set()
    for row in records:
        scenario_value = str(row.get("Scenario") or "").strip().lower()
        if not scenario_value or scenario_value in seen:
            continue
        seen.add(scenario_value)
        scenario_values.append(scenario_value)
    return [
        {"label": _format_scenario_label(value), "value": value}
        for value in scenario_values
    ]


def _active_mev_label_map(label_mode: str | None) -> dict[str, str]:
    normalized_mode = _normalize_mev_label_mode(label_mode)
    if normalized_mode == "long_name":
        return SAAS_PAGE_DATA.get("mev_label_map", {})
    if normalized_mode == "group_mnemonic":
        return SAAS_PAGE_DATA.get("mev_group_label_map", {})
    return {}


def _resolve_mev_label(mev_name: str, label_mode: str | None) -> str:
    return _active_mev_label_map(label_mode).get(mev_name) or mev_name


def _resolve_mev_description(mev_name: str) -> str | None:
    description = SAAS_PAGE_DATA.get("mev_description_map", {}).get(mev_name)
    return description or None


def _mev_description_label(mev_name: str) -> str:
    mev_types = _mev_types_for_name(mev_name)
    if mev_types == {"raw"}:
        return "Raw MEV description"
    if mev_types == {"transformed"}:
        return "Transformed MEV description"
    if mev_types:
        return "Raw / Transformed MEV description"
    return "MEV description"


def _show_historical_statistics(selected_value) -> bool:
    return str(selected_value or "").strip().lower() == "on"


def _compute_historical_dispersion_stats(records: list[dict], selected_scenarios) -> dict | None:
    scoped_records = _filter_records_by_scenarios(records, selected_scenarios)
    values_by_date: dict[object, dict[tuple[str, str], float]] = {}
    visible_lines: set[tuple[str, str]] = set()

    for row in scoped_records:
        date_value = row.get("Date")
        scenario_value = str(row.get("Scenario") or "").strip().lower()
        run_for_value = str(row.get("Run For") or "").strip()
        numeric_value = row.get("MEV Value")
        if date_value is None or not run_for_value or not _finite(numeric_value):
            continue
        line_key = (run_for_value, scenario_value)
        visible_lines.add(line_key)
        values_by_date.setdefault(date_value, {})[line_key] = float(numeric_value)

    if len(visible_lines) < 2:
        return {
            "visible_lines": len(visible_lines),
            "matched_quarters": 0,
            "has_dispersion": False,
        }

    per_date_ranges: list[float] = []
    per_date_stddevs: list[float] = []
    max_range_date = None
    max_range_value = None

    for date_value in sorted(values_by_date):
        line_values = list(values_by_date[date_value].values())
        if len(line_values) < 2:
            continue
        range_value = max(line_values) - min(line_values)
        stddev_value = statistics.pstdev(line_values) if len(line_values) > 1 else 0.0
        per_date_ranges.append(range_value)
        per_date_stddevs.append(stddev_value)
        if max_range_value is None or range_value > max_range_value:
            max_range_value = range_value
            max_range_date = date_value

    if not per_date_ranges:
        return {
            "visible_lines": len(visible_lines),
            "matched_quarters": 0,
            "has_dispersion": False,
        }

    return {
        "visible_lines": len(visible_lines),
        "matched_quarters": len(per_date_ranges),
        "avg_range": sum(per_date_ranges) / len(per_date_ranges),
        "max_range": max(per_date_ranges),
        "avg_stddev": sum(per_date_stddevs) / len(per_date_stddevs),
        "max_stddev": max(per_date_stddevs),
        "max_range_date": max_range_date,
        "has_dispersion": True,
    }


def _format_historical_dispersion_date(date_value) -> str:
    if date_value is None:
        return "—"
    if hasattr(date_value, "strftime"):
        return date_value.strftime("%b %d, %Y").replace(" 0", " ")
    return str(date_value)


def _build_historical_dispersion_summary(records: list[dict], selected_scenarios):
    stats = _compute_historical_dispersion_stats(records, selected_scenarios)
    if not stats:
        return None

    if not stats.get("has_dispersion"):
        return html.Div(
            className="saas-historical-dispersion saas-historical-dispersion-empty",
            children=[
                html.Div("Historical Dispersion", className="saas-historical-dispersion-title"),
                html.Div(
                    "Add another visible line through Scenario or Compare To to compute historical dispersion statistics.",
                    className="saas-historical-dispersion-note",
                ),
            ],
        )

    def stat_card(label: str, value: str, tooltip: str, tone: str = "default"):
        return html.Div(
            className=f"saas-historical-dispersion-card is-{tone}",
            children=[
                html.Div(
                    className="saas-historical-dispersion-card-label-row",
                    children=[
                        html.Div(label, className="saas-historical-dispersion-card-label"),
                        html.Span(
                            "i",
                            className="pd-info-chip saas-historical-dispersion-info",
                            title=tooltip,
                            **{"aria-label": f"{label}. {tooltip}"},
                        ),
                    ],
                ),
                html.Div(value, className="saas-historical-dispersion-card-value"),
            ],
        )

    max_range_caption = _format_historical_dispersion_date(stats.get("max_range_date"))
    return html.Div(
        className="saas-historical-dispersion",
        children=[
            html.Div(
                className="saas-historical-dispersion-header",
                children=[
                    html.Div("Historical Dispersion", className="saas-historical-dispersion-title"),
                    html.Div("Computed from the exact lines visible in this chart.", className="saas-historical-dispersion-note"),
                ],
            ),
            html.Div(
                className="saas-historical-dispersion-grid",
                children=[
                    stat_card(
                        "Visible lines",
                        str(stats["visible_lines"]),
                        "Number of distinct chart lines included in this historical dispersion calculation.",
                        "neutral",
                    ),
                    stat_card(
                        "Common quarters",
                        str(stats["matched_quarters"]),
                        "Number of historical dates with at least two visible line values, so dispersion can be computed.",
                        "neutral",
                    ),
                    stat_card(
                        "Avg range",
                        format_pd_mev_value(stats["avg_range"]),
                        "Average across common historical quarters of max(value) minus min(value) across the visible lines.",
                        "blue",
                    ),
                    stat_card(
                        "Max range",
                        format_pd_mev_value(stats["max_range"]),
                        "Largest single-quarter spread between the highest and lowest visible line values.",
                        "amber",
                    ),
                    stat_card(
                        "Avg stdev",
                        format_pd_mev_value(stats["avg_stddev"]),
                        "Average across common historical quarters of the population standard deviation across the visible lines.",
                        "blue",
                    ),
                    stat_card(
                        "Max stdev",
                        format_pd_mev_value(stats["max_stddev"]),
                        "Largest single-quarter population standard deviation across the visible line values.",
                        "amber",
                    ),
                ],
            ),
            html.Div(f"Worst quarter by range: {max_range_caption}", className="saas-historical-dispersion-footer"),
        ],
    )


def _build_model_mev_options(records: list[dict], label_mode: str | None) -> list[dict]:
    mev_names = sorted(
        {str(row.get("MEV Name") or "").strip() for row in records if str(row.get("MEV Name") or "").strip()},
        key=lambda value: _resolve_mev_label(value, label_mode).lower(),
    )
    options: list[dict] = []
    for mev_name in mev_names:
        options.append({"label": _resolve_mev_label(mev_name, label_mode), "value": mev_name})
    return options


def _records_for_model_scope(model_name: str, run_for, snapshot_period: str | None, compare_against=None) -> list[dict]:
    time_series_df = SAAS_PAGE_DATA.get("mev_time_series")
    if time_series_df is None or time_series_df.empty:
        return []

    selected_run_fors = _scoped_run_for_values(run_for, compare_against)
    filtered_df = time_series_df[time_series_df["Model Name"] == model_name]
    if selected_run_fors:
        filtered_df = filtered_df[filtered_df["Run For"].isin(selected_run_fors)]
    else:
        filtered_df = filtered_df.iloc[0:0]

    return _filter_records_by_snapshot_period(
        filtered_df.to_dict(orient="records"),
        _normalize_snapshot_period(snapshot_period),
    )


def _build_model_mev_options_for_mode(
    model_name: str,
    run_for,
    snapshot_period: str | None,
    mev_label_mode: str | None,
    selected_mev_mode: str | None,
    compare_against=None,
) -> list[dict]:
    base_records = _records_for_model_scope(model_name, run_for, snapshot_period, compare_against)
    return _build_model_mev_options(
        _filter_records_by_model_mevs(base_records, model_name, selected_mev_mode),
        mev_label_mode,
    )


def _scenario_toggle_label(selected_scenarios, scenario_options: list[dict]) -> str:
    normalized_scenarios = _normalize_selected_scenarios(selected_scenarios, scenario_options)
    if not normalized_scenarios:
        return "Select scenarios"

    all_values = [option["value"] for option in scenario_options if option.get("value")]
    if all_values and set(normalized_scenarios) == set(all_values):
        return "All"

    label_by_value = {option["value"]: option["label"] for option in scenario_options if option.get("value")}
    if len(normalized_scenarios) == 1:
        return label_by_value.get(normalized_scenarios[0], normalized_scenarios[0])
    return f"{len(normalized_scenarios)} scenarios selected"


def _single_selected_scenario(selected_scenarios, scenario_options: list[dict]) -> str:
    normalized_scenarios = _normalize_selected_scenarios(selected_scenarios, scenario_options)
    if normalized_scenarios:
        return normalized_scenarios[0]
    all_values = [option["value"] for option in scenario_options if option.get("value")]
    return all_values[0] if all_values else ""


def _build_model_scenario_dropdown(
    model_name: str,
    selected_scenarios,
    scenario_options: list[dict],
    *,
    single_select: bool = False,
) -> html.Div:
    effective_scenarios = _normalize_selected_scenarios(selected_scenarios, scenario_options)
    all_values = [option["value"] for option in scenario_options if option.get("value")]
    if single_select:
        selected_value = _single_selected_scenario(selected_scenarios, scenario_options)
        return html.Div(
            className="checkbox-dropdown single-select-dropdown",
            children=[
                dcc.Dropdown(
                    id={"type": layout.MODEL_SCENARIO_FILTER_TYPE, "model": model_name},
                    options=scenario_options,
                    value=selected_value,
                    clearable=False,
                    searchable=False,
                    style={"display": "none"},
                ),
                dcc.Checklist(
                    id={"type": layout.MODEL_SCENARIO_SELECT_ALL_TYPE, "model": model_name},
                    options=[],
                    value=[],
                    style={"display": "none"},
                ),
                html.Button(
                    _scenario_toggle_label(selected_value, scenario_options),
                    id={"type": layout.MODEL_SCENARIO_TOGGLE_TYPE, "model": model_name},
                    type="button",
                    n_clicks=0,
                    className="checkbox-dropdown-toggle",
                ),
                html.Div(
                    id={"type": layout.MODEL_SCENARIO_MENU_TYPE, "model": model_name},
                    className="checkbox-dropdown-menu",
                    children=[
                        html.Button(
                            [
                                html.Span(option["label"], className="single-select-option-label"),
                                html.Span("✓", className="single-select-option-check"),
                            ],
                            id={"type": layout.MODEL_SCENARIO_OPTION_TYPE, "model": model_name, "value": option["value"]},
                            type="button",
                            n_clicks=0,
                            className="single-select-option is-selected" if option["value"] == selected_value else "single-select-option",
                        )
                        for option in scenario_options
                    ],
                ),
            ],
        )
    return html.Div(
        className="checkbox-dropdown",
        children=[
            html.Button(
                _scenario_toggle_label(effective_scenarios, scenario_options),
                id={"type": layout.MODEL_SCENARIO_TOGGLE_TYPE, "model": model_name},
                type="button",
                n_clicks=0,
                className="checkbox-dropdown-toggle",
            ),
            html.Div(
                id={"type": layout.MODEL_SCENARIO_MENU_TYPE, "model": model_name},
                className="checkbox-dropdown-menu",
                children=[
                    dcc.Checklist(
                        id={"type": layout.MODEL_SCENARIO_SELECT_ALL_TYPE, "model": model_name},
                        options=[{"label": "All", "value": "all"}],
                        value=["all"] if all_values and set(effective_scenarios) == set(all_values) else [],
                        className="pd-models-select-all",
                    ),
                    dcc.Checklist(
                        id={"type": layout.MODEL_SCENARIO_FILTER_TYPE, "model": model_name},
                        options=scenario_options,
                        value=effective_scenarios,
                        className="pd-models-checklist",
                    ),
                ],
            ),
        ],
    )


def _mev_type_toggle_label(selected_mev_mode: str | None, mev_type_options: list[dict]) -> str:
    normalized_mode = _normalize_selected_mev_mode(selected_mev_mode, mev_type_options)
    label_by_value = {option["value"]: option["label"] for option in mev_type_options if option.get("value")}
    return label_by_value.get(normalized_mode, normalized_mode or "Select MEV View")


def _build_model_mev_type_dropdown(model_name: str, selected_mev_mode: str | None, mev_type_options: list[dict]) -> html.Div:
    effective_mev_mode = _normalize_selected_mev_mode(selected_mev_mode, mev_type_options)
    return html.Div(
        className="checkbox-dropdown single-select-dropdown",
        children=[
            dcc.Dropdown(
                id={"type": layout.MODEL_MEV_TYPE_FILTER_TYPE, "model": model_name},
                options=mev_type_options,
                value=effective_mev_mode,
                clearable=False,
                searchable=False,
                style={"display": "none"},
            ),
            html.Button(
                _mev_type_toggle_label(effective_mev_mode, mev_type_options),
                id={"type": layout.MODEL_MEV_TYPE_TOGGLE_TYPE, "model": model_name},
                type="button",
                n_clicks=0,
                className="checkbox-dropdown-toggle",
            ),
            html.Div(
                id={"type": layout.MODEL_MEV_TYPE_MENU_TYPE, "model": model_name},
                className="checkbox-dropdown-menu single-select-menu",
                children=[
                    html.Button(
                        [
                            html.Span(option["label"], className="single-select-option-label"),
                            html.Span("✓", className="single-select-option-check", **{"aria-hidden": "true"}),
                        ],
                        id={"type": layout.MODEL_MEV_TYPE_OPTION_TYPE, "model": model_name, "value": option["value"]},
                        type="button",
                        n_clicks=0,
                        className="single-select-option is-selected" if option["value"] == effective_mev_mode else "single-select-option",
                    )
                    for option in mev_type_options
                ],
            ),
        ],
    )


def _filter_records_by_mevs(records: list[dict], selected_mevs) -> list[dict]:
    selected_mev_values = _normalize_selected_mevs(selected_mevs)
    if not selected_mev_values:
        return []
    selected_mev_set = set(selected_mev_values)
    return [row for row in records if str(row.get("MEV Name") or "").strip() in selected_mev_set]


def _build_single_mev_option_buttons(options: list[dict], selected_value: str | None, model_name: str):
    return [
        html.Button(
            [
                html.Span(option["label"], className="single-select-option-label"),
                html.Span("✓", className="single-select-option-check", **{"aria-hidden": "true"}),
            ],
            id={"type": layout.MODEL_MEV_SINGLE_OPTION_TYPE, "model": model_name, "value": option["value"]},
            type="button",
            n_clicks=0,
            className="single-select-option is-selected" if option["value"] == selected_value else "single-select-option",
        )
        for option in options
    ]


def _family_display_mevs(model_name: str, selected_mev: str | None, records: list[dict] | None = None) -> list[str]:
    available_names = {
        str(row.get("MEV Name") or "").strip()
        for row in (records or [])
        if str(row.get("MEV Name") or "").strip()
    }
    family_map = _family_map_for_model(model_name)
    normalized_selected = str(selected_mev or "").strip()
    if not normalized_selected:
        return []

    family_values = [normalized_selected]
    family_values.extend(family_map.get(normalized_selected, []))
    ordered_values = list(dict.fromkeys(value for value in family_values if value))
    if not available_names:
        return ordered_values
    return [value for value in ordered_values if value in available_names]


def _active_selected_mevs(
    model_name: str,
    selected_mev_mode: str | None,
    selected_mev_single,
    selected_mevs_multi,
    records: list[dict] | None = None,
) -> list[str]:
    normalized_mode = _normalize_selected_mev_mode(selected_mev_mode)
    if normalized_mode == "family":
        selected_value = _normalize_selected_mevs(selected_mev_single)
        selected_mev = selected_value[0] if selected_value else None
        return _family_display_mevs(model_name, selected_mev, records)
    return _normalize_selected_mevs(selected_mevs_multi)


def _mev_toggle_label(
    selected_mev_mode: str | None,
    selected_mev_single,
    selected_mevs_multi,
    single_mev_options: list[dict],
    multi_mev_options: list[dict],
) -> str:
    normalized_mode = _normalize_selected_mev_mode(selected_mev_mode)
    if normalized_mode == "family":
        selected_mev_values = _normalize_selected_mevs(selected_mev_single)
        if not selected_mev_values:
            return _mev_picker_empty_label(normalized_mode)
        label_by_value = {option["value"]: option["label"] for option in single_mev_options if option.get("value")}
        return label_by_value.get(selected_mev_values[0], selected_mev_values[0])

    selected_mev_values = _normalize_selected_mevs(selected_mevs_multi)
    if not selected_mev_values:
        return _mev_picker_empty_label(normalized_mode)

    all_values = [option["value"] for option in multi_mev_options if option.get("value")]
    if all_values and set(selected_mev_values) == set(all_values):
        if normalized_mode == "transformed_only":
            return "All"
        return "All raw MEVs"

    label_by_value = {option["value"]: option["label"] for option in multi_mev_options if option.get("value")}
    if len(selected_mev_values) == 1:
        return label_by_value.get(selected_mev_values[0], selected_mev_values[0])
    return f"{len(selected_mev_values)} MEVs selected"


def _build_model_mev_dropdown(
    model_name: str,
    selected_mev_mode: str | None,
    single_mev_options: list[dict],
    selected_mev_single: str | None,
    multi_mev_options: list[dict],
    selected_mevs_multi: list[str],
) -> html.Div:
    normalized_mode = _normalize_selected_mev_mode(selected_mev_mode)
    show_single_picker = normalized_mode == "family"
    return html.Div(
        className="checkbox-dropdown pd-mev-filter-dropdown",
        children=[
            html.Button(
                _mev_toggle_label(
                    normalized_mode,
                    selected_mev_single,
                    selected_mevs_multi,
                    single_mev_options,
                    multi_mev_options,
                ),
                id={"type": layout.MODEL_MEV_TOGGLE_TYPE, "model": model_name},
                type="button",
                n_clicks=0,
                className="checkbox-dropdown-toggle",
            ),
            html.Div(
                id={"type": layout.MODEL_MEV_MENU_TYPE, "model": model_name},
                className="checkbox-dropdown-menu",
                children=[
                    html.Div(
                        id={"type": layout.MODEL_MEV_SINGLE_SECTION_TYPE, "model": model_name},
                        className=_mev_picker_section_class(visible=show_single_picker),
                        children=[
                            dcc.Dropdown(
                                id={"type": layout.MODEL_MEV_SINGLE_VALUE_TYPE, "model": model_name},
                                options=single_mev_options,
                                value=selected_mev_single,
                                clearable=False,
                                searchable=False,
                                style={"display": "none"},
                            ),
                            html.Div(
                                id={"type": layout.MODEL_MEV_SINGLE_OPTIONS_TYPE, "model": model_name},
                                className="saas-mev-picker-single-options",
                                children=_build_single_mev_option_buttons(single_mev_options, selected_mev_single, model_name),
                            ),
                        ],
                    ),
                    html.Div(
                        id={"type": layout.MODEL_MEV_MULTI_SECTION_TYPE, "model": model_name},
                        className=_mev_picker_section_class(visible=not show_single_picker),
                        children=[
                            dcc.Checklist(
                                id={"type": layout.MODEL_MEV_SELECT_ALL_TYPE, "model": model_name},
                                options=[{"label": "All", "value": "all"}],
                                value=["all"] if multi_mev_options and set(selected_mevs_multi) == {option["value"] for option in multi_mev_options if option.get("value")} else [],
                                className="pd-models-select-all",
                            ),
                            dcc.Checklist(
                                id={"type": layout.MODEL_MEV_FILTER_TYPE, "model": model_name},
                                options=multi_mev_options,
                                value=selected_mevs_multi,
                                className="pd-models-checklist",
                            ),
                        ],
                    ),
                ],
            ),
        ],
    )


def _build_model_figure(
    model_name: str,
    records: list[dict],
    reference_records: list[dict],
    selected_mev_mode,
    selected_scenarios,
    snapshot_period: str | None,
    mev_label_mode: str | None,
    range_value,
    reference_lines: str | None,
    primary_run_for: str | None = None,
    development_date=None,
    current_date=None,
    selected_mevs=None,
):
    monitoring_reference_records = _filter_records_by_mevs(
        _filter_records_by_scenarios(
            reference_records,
            selected_scenarios,
        ),
        selected_mevs,
    )
    scoped_reference_records = _filter_records_by_date_range(
        monitoring_reference_records,
        range_value,
    )
    y_axis_title = None
    if selected_mevs:
        selected_mev_name = _normalize_selected_mevs(selected_mevs)
        if selected_mev_name:
            y_axis_title = _resolve_mev_label(selected_mev_name[0], mev_label_mode)
    return build_saas_mev_time_series_figure(
        _filter_records_by_date_range(
            _filter_records_by_mevs(
                _filter_records_by_scenarios(
                    records,
                    selected_scenarios,
                ),
                selected_mevs,
            ),
            range_value,
        ),
        mev_label_map=_active_mev_label_map(mev_label_mode),
        model_label_map={model_name: model_name},
        y_axis_title=y_axis_title,
        snapshot_period=_normalize_snapshot_period(snapshot_period),
        historical_reference_records=scoped_reference_records,
        monitoring_reference_records=monitoring_reference_records,
        reference_lines=reference_lines or layout.DEFAULT_REFERENCE_LINES,
        empty_message=f"No MEV time-series data matches the active filters for {model_name}.",
        primary_run_for=primary_run_for,
        development_date=development_date,
        current_date=current_date,
        projection_start_date=_projection_start_date_for_run_for(primary_run_for),
    )


def _build_model_chart_cards(
    model_name: str,
    records: list[dict],
    reference_records: list[dict],
    selected_mev_mode,
    selected_scenarios,
    snapshot_period: str | None,
    mev_label_mode: str | None,
    range_value,
    reference_lines: str | None,
    selected_mevs,
    shared_meta_parts: list[str],
    primary_run_for: str | None = None,
    show_historical_statistics=False,
):
    normalized_mev_mode = _normalize_selected_mev_mode(selected_mev_mode)
    scenario_options = _build_scenario_options(records)
    visible_mev_names = {
        str(row.get("MEV Name") or "").strip()
        for row in records
        if str(row.get("MEV Name") or "").strip()
    }
    effective_mev_values = [
        value for value in _normalize_selected_mevs(selected_mevs)
        if value in visible_mev_names
    ]
    effective_scenarios = _normalize_selected_scenarios(selected_scenarios, scenario_options)

    if not effective_mev_values:
        empty_mev_message = "Choose one transformed MEV in the model header to render the charts."
        if normalized_mev_mode != "family":
            empty_mev_message = "Choose one or more MEVs in the model header to render the charts."
        return [
            html.Article(
                className="pd-mev-chart-card",
                children=[
                    html.Div(
                        className="pd-mev-chart-header",
                        children=[
                            html.Div(
                                [
                                    html.Div("No MEVs selected", className="pd-mev-chart-title"),
                                    html.Div(
                                        empty_mev_message,
                                        className="pd-mev-chart-meta",
                                    ),
                                ]
                            ),
                        ],
                    ),
                ],
            ),
        ]

    if not effective_scenarios:
        return [
            html.Article(
                className="pd-mev-chart-card",
                children=[
                    html.Div(
                        className="pd-mev-chart-header",
                        children=[
                            html.Div(
                                [
                                    html.Div("No scenarios selected", className="pd-mev-chart-title"),
                                    html.Div(
                                        "Choose one or more scenarios in the model header to render the charts.",
                                        className="pd-mev-chart-meta",
                                    ),
                                ]
                            ),
                        ],
                    ),
                ],
            ),
        ]

    if reference_lines == "monitoring" and len(effective_scenarios) != 1:
        return [
            html.Article(
                className="pd-mev-chart-card",
                children=[
                    html.Div(
                        className="pd-mev-chart-header",
                        children=[
                            html.Div(
                                [
                                    html.Div("Choose one scenario", className="pd-mev-chart-title"),
                                    html.Div(
                                        "Monitoring view requires exactly one scenario to be selected.",
                                        className="pd-mev-chart-meta",
                                    ),
                                ]
                            ),
                        ],
                    ),
                ],
            ),
        ]

    cards = []
    development_date = _model_development_date(model_name, primary_run_for)
    current_date = _current_date_for_run_for(primary_run_for)

    for mev_name in effective_mev_values:
        mev_records = [row for row in records if str(row.get("MEV Name") or "").strip() == mev_name]
        mev_reference_records = [row for row in reference_records if str(row.get("MEV Name") or "").strip() == mev_name]
        mev_label = _resolve_mev_label(mev_name, mev_label_mode)
        mev_description = _resolve_mev_description(mev_name)
        meta_text = f"{_mev_description_label(mev_name)}: {mev_description}" if mev_description else ""
        monitoring_summary = None
        historical_dispersion_summary = None
        if reference_lines == "monitoring":
            monitoring_summary = _build_monitoring_summary(
                mev_reference_records,
                normalized_mev_mode,
                effective_scenarios,
                [mev_name],
                primary_run_for,
                development_date,
                current_date,
            )
        if _normalize_snapshot_period(snapshot_period) == "history" and show_historical_statistics:
            historical_dispersion_summary = _build_historical_dispersion_summary(
                mev_records,
                effective_scenarios,
            )
        fig = _build_model_figure(
            model_name,
            mev_records,
            mev_reference_records,
            normalized_mev_mode,
            effective_scenarios,
            snapshot_period,
            mev_label_mode,
            range_value,
            reference_lines,
            primary_run_for,
            development_date,
            current_date,
            [mev_name],
        )

        cards.append(
            html.Article(
                className="pd-mev-chart-card",
                children=[
                    html.Div(
                        className="pd-mev-chart-header",
                        children=[
                            html.Div(
                                [
                                    html.Div(mev_label, className="pd-mev-chart-title"),
                                    html.Div(meta_text, className="pd-mev-chart-meta"),
                                ]
                            ),
                        ],
                    ),
                    monitoring_summary,
                    historical_dispersion_summary,
                    dcc.Graph(
                        figure=fig,
                        config=_GRAPH_CONFIG,
                        className="pd-mev-chart",
                    ),
                ],
            )
        )

    return cards


def _build_model_panel(
    panel_index: int,
    model_name: str,
    records: list[dict],
    run_for,
    compare_against,
    snapshot_period: str | None,
    mev_label_mode: str | None,
    reference_lines: str | None,
    show_historical_statistics=False,
):
    snapshot_period_value = _normalize_snapshot_period(snapshot_period)
    visible_records = _filter_records_by_snapshot_period(records, snapshot_period_value)
    scenario_options = _build_scenario_options(visible_records)
    mev_type_options = list(layout.MEV_TYPE_OPTIONS)
    default_model_mev_mode = layout.DEFAULT_MEV_TYPE
    default_model_scenarios = (
        [scenario_options[0]["value"]] if scenario_options else []
    ) if reference_lines == "monitoring" else [option["value"] for option in scenario_options if option.get("value")]
    family_mev_options = _build_model_mev_options(
        _filter_records_by_model_mevs(visible_records, model_name, "family"),
        mev_label_mode,
    )
    transformed_mev_options = _build_model_mev_options(
        _filter_records_by_model_mevs(visible_records, model_name, "transformed_only"),
        mev_label_mode,
    )
    default_family_mev = family_mev_options[0]["value"] if family_mev_options else ""
    default_model_mevs = [option["value"] for option in transformed_mev_options]
    default_display_mevs = _active_selected_mevs(
        model_name,
        default_model_mev_mode,
        default_family_mev,
        default_model_mevs,
        visible_records,
    )
    date_periods = _available_date_periods(visible_records)
    mev_names = sorted({str(row.get("MEV Name") or "").strip() for row in visible_records if str(row.get("MEV Name") or "").strip()})
    scenario_names = sorted({str(row.get("Scenario") or "").strip().lower() for row in visible_records if str(row.get("Scenario") or "").strip()})
    date_values = sorted({row.get("Date") for row in visible_records if row.get("Date") is not None})

    model_segments = SAAS_PAGE_DATA.get("model_segments", {})
    segment_name = layout.format_segment_label(model_segments.get(model_name))
    range_start = _format_month_year(date_values[0]) if date_values else None
    range_end = _format_month_year(date_values[-1]) if date_values else None

    meta_parts = [f"Reporting Cycle: {_run_for_meta_label(run_for)}"]
    snapshot_label = next(
        (option["label"] for option in layout.SUBNAV_VIEW_OPTIONS if option["value"] == snapshot_period_value),
        None,
    )
    if snapshot_label:
        meta_parts.append(snapshot_label)
    if mev_names:
        meta_parts.append(_pluralize(len(mev_names), "MEV"))
    if scenario_names:
        meta_parts.append(_pluralize(len(scenario_names), "scenario"))
    if range_start and range_end:
        meta_parts.append(f"{range_start} to {range_end}")

    chart_cards = _build_model_chart_cards(
        model_name,
        visible_records,
        records,
        default_model_mev_mode,
        default_model_scenarios,
        snapshot_period_value,
        mev_label_mode,
        None,
        reference_lines,
        default_display_mevs,
        meta_parts,
        _normalize_selected_run_fors(run_for)[0] if _normalize_selected_run_fors(run_for) else None,
        show_historical_statistics,
    )

    return html.Div(
        id=_model_panel_id(model_name),
        className="section-card pd-mev-model-panel",
        children=[
            html.Div(
                className="pd-mev-model-heading",
                children=[
                    html.Div(
                        className="pd-mev-model-copy",
                        children=[
                            html.Div(f"{panel_index}. Model Profile", className="pd-content-kicker"),
                            html.H4(model_name),
                            html.P(f"Segment: {segment_name}"),
                        ],
                    ),
                    html.Div(
                        className="pd-mev-model-heading-actions",
                        children=[
                            html.Div(
                                className="pd-mev-filter-group saas-model-mev-type-filter",
                                children=[
                                    html.Label("MEV View"),
                                    _build_model_mev_type_dropdown(
                                        model_name,
                                        default_model_mev_mode,
                                        mev_type_options,
                                    ),
                                ],
                            ),
                            html.Div(
                                className="pd-mev-filter-group saas-model-scenario-filter",
                                children=[
                                    html.Label("Scenario"),
                                    _build_model_scenario_dropdown(
                                        model_name,
                                        default_model_scenarios,
                                        scenario_options,
                                        single_select=reference_lines == "monitoring",
                                    ),
                                ],
                            ),
                            html.Div(
                                className="pd-mev-filter-group saas-model-mev-filter",
                                children=[
                                    html.Label(
                                        _mev_picker_label(default_model_mev_mode),
                                        id={"type": layout.MODEL_MEV_LABEL_TYPE, "model": model_name},
                                    ),
                                    _build_model_mev_dropdown(
                                        model_name,
                                        default_model_mev_mode,
                                        family_mev_options,
                                        default_family_mev,
                                        transformed_mev_options,
                                        default_model_mevs,
                                    ),
                                ],
                            ),
                            html.Div(
                                id={"type": layout.MODEL_DATE_RANGE_CONTROLS_TYPE, "model": model_name},
                                className="saas-date-range-controls",
                                children=[
                                    layout.build_model_date_range_controls(
                                        model_name,
                                        date_periods,
                                        None,
                                        disabled=snapshot_period_value == "projection",
                                    )
                                ],
                            ),
                        ],
                    ),
                ],
            ),
            html.Div(
                id={"type": layout.MODEL_MEV_GRID_TYPE, "model": model_name},
                className="pd-mev-chart-grid",
                children=chart_cards,
            ),
        ],
    )


def _build_model_report_figures(
    model_name: str,
    records: list[dict],
    reference_records: list[dict],
    selected_mev_mode,
    selected_scenarios,
    snapshot_period: str | None,
    mev_label_mode: str | None,
    reference_lines: str | None,
    selected_mevs,
    primary_run_for: str | None = None,
) -> list[tuple[str, object]]:
    """Build (title, figure) pairs for a model using the default chart selections."""
    normalized_mev_mode = _normalize_selected_mev_mode(selected_mev_mode)
    scenario_options = _build_scenario_options(records)
    visible_mev_names = {
        str(row.get("MEV Name") or "").strip()
        for row in records
        if str(row.get("MEV Name") or "").strip()
    }
    effective_mev_values = [
        value for value in _normalize_selected_mevs(selected_mevs)
        if value in visible_mev_names
    ]
    effective_scenarios = _normalize_selected_scenarios(selected_scenarios, scenario_options)

    if not effective_mev_values or not effective_scenarios:
        return []
    if reference_lines == "monitoring" and len(effective_scenarios) != 1:
        return []

    development_date = _model_development_date(model_name, primary_run_for)
    current_date = _current_date_for_run_for(primary_run_for)

    figures: list[tuple[str, object]] = []
    for mev_name in effective_mev_values:
        mev_records = [row for row in records if str(row.get("MEV Name") or "").strip() == mev_name]
        mev_reference_records = [row for row in reference_records if str(row.get("MEV Name") or "").strip() == mev_name]
        mev_label = _resolve_mev_label(mev_name, mev_label_mode)
        fig = _build_model_figure(
            model_name,
            mev_records,
            mev_reference_records,
            normalized_mev_mode,
            effective_scenarios,
            snapshot_period,
            mev_label_mode,
            None,
            reference_lines,
            primary_run_for,
            development_date,
            current_date,
            [mev_name],
        )
        figures.append((f"{model_name} — {mev_label}", fig))
    return figures


def _build_model_report_sections(
    model_name: str,
    records: list[dict],
    run_for,
    snapshot_period: str | None,
    mev_label_mode: str | None,
    reference_lines: str | None,
) -> list[tuple[str, object]]:
    """Build (title, figure) pairs for a model panel's default view."""
    snapshot_period_value = _normalize_snapshot_period(snapshot_period)
    visible_records = _filter_records_by_snapshot_period(records, snapshot_period_value)
    scenario_options = _build_scenario_options(visible_records)
    default_model_mev_mode = layout.DEFAULT_MEV_TYPE
    default_model_scenarios = (
        [scenario_options[0]["value"]] if scenario_options else []
    ) if reference_lines == "monitoring" else [option["value"] for option in scenario_options if option.get("value")]
    family_mev_options = _build_model_mev_options(
        _filter_records_by_model_mevs(visible_records, model_name, "family"),
        mev_label_mode,
    )
    transformed_mev_options = _build_model_mev_options(
        _filter_records_by_model_mevs(visible_records, model_name, "transformed_only"),
        mev_label_mode,
    )
    default_family_mev = family_mev_options[0]["value"] if family_mev_options else ""
    default_model_mevs = [option["value"] for option in transformed_mev_options]
    default_display_mevs = _active_selected_mevs(
        model_name,
        default_model_mev_mode,
        default_family_mev,
        default_model_mevs,
        visible_records,
    )
    selected_run_fors = _normalize_selected_run_fors(run_for)
    primary_run_for = selected_run_fors[0] if selected_run_fors else None

    return _build_model_report_figures(
        model_name,
        visible_records,
        records,
        default_model_mev_mode,
        default_model_scenarios,
        snapshot_period_value,
        mev_label_mode,
        reference_lines,
        default_display_mevs,
        primary_run_for,
    )


def _build_report_figures(run_for, compare_against, segment, selected_models, snapshot_period, reference_lines, mev_label_mode) -> list[tuple[str, object]]:
    selected_run_fors = _normalize_selected_run_fors(run_for)
    scoped_run_fors = _scoped_run_for_values(run_for, compare_against)
    effective_models = _effective_model_names(segment, selected_models)

    if not selected_run_fors or not effective_models:
        return []

    time_series_df = SAAS_PAGE_DATA.get("mev_time_series")
    if time_series_df is None or time_series_df.empty:
        return []

    filtered_df = time_series_df[time_series_df["Model Name"].isin(effective_models)]
    filtered_df = filtered_df[filtered_df["Run For"].isin(scoped_run_fors)]
    records = filtered_df.to_dict(orient="records")

    sections: list[tuple[str, object]] = []
    for model_name in effective_models:
        model_records = [row for row in records if row.get("Model Name") == model_name]
        sections.extend(
            _build_model_report_sections(
                model_name,
                model_records,
                run_for,
                snapshot_period,
                mev_label_mode,
                reference_lines,
            )
        )
    return sections


def _build_saas_report_html(sections: list[tuple[str, object]], meta_lines: list[str]) -> str:
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    meta_items = "".join(f"<li>{html_escape(line)}</li>" for line in meta_lines)

    if sections:
        chart_blocks = []
        for index, (title, fig) in enumerate(sections):
            # Fixed pixel size keeps the chart from being laid out at the on-screen
            # viewport width and then clipped when the browser paginates for print.
            fig = fig.update_layout(autosize=False, width=900, height=420)
            chart_html = fig.to_html(
                full_html=False,
                include_plotlyjs="cdn" if index == 0 else False,
                config={"responsive": False},
            )
            chart_blocks.append(
                f'<section class="saas-report-chart"><h2>{html_escape(title)}</h2>{chart_html}</section>'
            )
        body = "\n".join(chart_blocks)
    else:
        body = "<p>No charts match the current filters.</p>"

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>SAAS MEV Report</title>
<style>
  body {{ font-family: -apple-system, Helvetica, Arial, sans-serif; margin: 24px; color: #1f2933; }}
  h1 {{ font-size: 20px; margin-bottom: 4px; }}
  .saas-report-meta {{ font-size: 13px; color: #52606d; margin-bottom: 24px; }}
  .saas-report-meta ul {{ margin: 4px 0 0; padding-left: 18px; }}
  .saas-report-chart {{ margin-bottom: 32px; page-break-inside: avoid; }}
  .saas-report-chart h2 {{ font-size: 15px; margin-bottom: 8px; }}
  .saas-report-chart .plotly-graph-div {{ margin: 0; }}
  @media print {{
    @page {{ size: landscape; margin: 12mm; }}
    .saas-report-chart {{ page-break-after: always; }}
  }}
</style>
</head>
<body>
  <h1>Scenario Analysis as a Service (SAAS) &mdash; MEV Report</h1>
  <div class="saas-report-meta">
    Generated {generated_at}
    <ul>{meta_items}</ul>
  </div>
  {body}
  <p class="saas-report-print-hint">To save this report as a PDF, open this file in a browser and use Print &rarr; Save as PDF.</p>
</body>
</html>"""


# ---------------------------------------------------------------------------
# Excel export (metrics workbook)
# ---------------------------------------------------------------------------

# (key, header, description). The README tab is generated from this list so the
# column descriptions always match the Metrics tab.
SAAS_METRIC_COLUMNS: list[tuple[str, str, str]] = [
    ("model", "Model", "Model name."),
    ("mev_type", "MEV Type", "Whether the MEV is a Raw or a Transformed variable."),
    ("mev", "MEV", "MEV name, shown using the MEV Label mode selected on the dashboard."),
    ("history_min", "History Min", "Minimum of the historical (Quarter <= 0) MEV values - the data behind the Min-Max lines."),
    ("history_max", "History Max", "Maximum of the historical (Quarter <= 0) MEV values - the data behind the Min-Max lines."),
    ("history_mean", "History Mean", "Mean of the historical (Quarter <= 0) MEV values."),
    ("history_std", "History STD", "Population standard deviation (ddof=0) of the historical MEV values."),
    ("sevadv_min", "Severely Adverse Min", "Minimum of the projection (Quarter > 0) MEV values for the chosen scenario."),
    ("sevadv_max", "Severely Adverse Max", "Maximum of the projection (Quarter > 0) MEV values for the chosen scenario."),
    ("is_min_beyond", "Is Severely Adverse Min Beyond?", "Yes if the projection min is lower than the historical min."),
    ("is_max_beyond", "Is Severely Adverse Max Beyond?", "Yes if the projection max is greater than the historical max."),
    ("minmax_daterange", "Historical Min-Max Date Range", "Date range spanned by the historical (Quarter <= 0) data: earliest historical date to latest historical date."),
    ("time_of_min", "Date of Min", "Date on which the historical minimum value occurs."),
    ("time_of_max", "Date of Max", "Date on which the historical maximum value occurs."),
    ("history_min_2std", "History Min - 2 STD", "History Min - 2 x History STD."),
    ("history_max_2std", "History Max + 2 STD", "History Max + 2 x History STD."),
    ("is_min_lt_2std", "Is Severely Adverse Min < History Min - 2 STD?", "Yes if the projection min is less than History Min - 2 x History STD."),
    ("is_max_gt_2std", "Is Severely Adverse Max > History Max + 2 STD?", "Yes if the projection max is greater than History Max + 2 x History STD."),
]

_BOOLEAN_METRIC_KEYS = {"is_min_beyond", "is_max_beyond", "is_min_lt_2std", "is_max_gt_2std"}
_DATE_METRIC_KEYS = {"time_of_min", "time_of_max"}
_FLOAT_METRIC_KEYS = {
    "history_min", "history_max", "history_mean", "history_std",
    "sevadv_min", "sevadv_max",
    "history_min_2std", "history_max_2std",
}


def _run_for_filename_prefix(run_for) -> str:
    """Filesystem-safe prefix derived from the selected Reporting Cycle."""
    primary_run_for = _primary_run_for_value(run_for)
    slug = re.sub(r"[^A-Za-z0-9]+", "-", str(primary_run_for or "").strip()).strip("-")
    return slug or "SAAS"


def _excel_to_py_date(value):
    if value is None:
        return None
    if hasattr(value, "to_pydatetime"):
        try:
            return value.to_pydatetime()
        except (TypeError, ValueError):
            return None
    return value


def _excel_quarter_label(value) -> str:
    parsed = _excel_to_py_date(value)
    if parsed is None:
        return ""
    quarter = ((parsed.month - 1) // 3) + 1
    return f"{parsed.year}Q{quarter}"


def _excel_mev_type_label(mev_name: str) -> str:
    mev_types = _mev_types_for_name(mev_name)
    if mev_types == {"raw"}:
        return "Raw"
    if mev_types == {"transformed"}:
        return "Transformed"
    if mev_types:
        return "Raw / Transformed"
    return "—"


def _compute_saas_metric_record(model_name: str, mev_label: str, subset_rows: list[dict]) -> dict | None:
    history = [
        (row.get("Date"), float(row.get("MEV Value")))
        for row in subset_rows
        if _finite(row.get("MEV Value")) and _finite(row.get("Quarter")) and float(row.get("Quarter")) <= 0
    ]
    projection = [
        float(row.get("MEV Value"))
        for row in subset_rows
        if _finite(row.get("MEV Value")) and _finite(row.get("Quarter")) and float(row.get("Quarter")) > 0
    ]
    if not history:
        return None

    history_values = [value for _, value in history]
    history_min = min(history_values)
    history_max = max(history_values)
    history_mean = statistics.fmean(history_values)
    history_std = statistics.pstdev(history_values) if len(history_values) > 1 else 0.0

    def _extreme_date(target: float):
        candidates = sorted(
            (date_value for date_value, value in history if value == target and date_value is not None),
            key=lambda value: (getattr(value, "year", 0), getattr(value, "month", 0), getattr(value, "day", 0)),
        )
        return _excel_to_py_date(candidates[0]) if candidates else None

    sevadv_min = min(projection) if projection else None
    sevadv_max = max(projection) if projection else None
    history_min_2std = history_min - 2 * history_std
    history_max_2std = history_max + 2 * history_std

    time_of_min = _extreme_date(history_min)
    time_of_max = _extreme_date(history_max)

    # Full date span of the historical (Quarter <= 0) data: earliest to latest date.
    history_dates = sorted(
        _excel_to_py_date(date_value) for date_value, _ in history if date_value is not None
    )
    if len(history_dates) >= 2:
        minmax_daterange = f"{history_dates[0]:%Y-%m-%d} to {history_dates[-1]:%Y-%m-%d}"
    elif len(history_dates) == 1:
        minmax_daterange = f"{history_dates[0]:%Y-%m-%d}"
    else:
        minmax_daterange = ""

    return {
        "model": model_name,
        "mev": mev_label,
        "history_min": history_min,
        "history_max": history_max,
        "history_mean": history_mean,
        "history_std": history_std,
        "sevadv_min": sevadv_min,
        "sevadv_max": sevadv_max,
        "is_min_beyond": sevadv_min is not None and sevadv_min < history_min,
        "is_max_beyond": sevadv_max is not None and sevadv_max > history_max,
        "minmax_daterange": minmax_daterange,
        "time_of_min": time_of_min,
        "time_of_max": time_of_max,
        "history_min_2std": history_min_2std,
        "history_max_2std": history_max_2std,
        "is_min_lt_2std": sevadv_min is not None and sevadv_min < history_min_2std,
        "is_max_gt_2std": sevadv_max is not None and sevadv_max > history_max_2std,
    }


def _build_saas_chart_spec(model_name: str, mev_label: str, subset_rows: list[dict], *, mev_type: str | None = None) -> dict:
    points = sorted(
        (
            (float(row.get("Quarter")), row.get("Date"), float(row.get("MEV Value")))
            for row in subset_rows
            if _finite(row.get("MEV Value")) and _finite(row.get("Quarter"))
        ),
        key=lambda item: item[0],
    )
    labels = [_excel_quarter_label(date_value) for _, date_value, _ in points]
    history = [value if quarter <= 0 else None for quarter, _, value in points]
    projection = [value if quarter > 0 else None for quarter, _, value in points]
    history_values = [value for quarter, _, value in points if quarter <= 0]
    type_suffix = f" [{mev_type}]" if mev_type and mev_type != "—" else ""
    return {
        "model": model_name,
        "mev_type": mev_type or "",
        "title": f"{mev_label}{type_suffix}",
        "y_title": mev_label,
        "labels": labels,
        "history": history,
        "projection": projection,
        "hist_min": min(history_values) if history_values else None,
        "hist_max": max(history_values) if history_values else None,
    }


def _compute_saas_metrics(run_for, segment, selected_models, mev_label_mode, scenario):
    """Return (metric_rows, chart_specs, primary_run_for) for the chosen scenario."""
    time_series_df = SAAS_PAGE_DATA.get("mev_time_series")
    selected_run_fors = _normalize_selected_run_fors(run_for)
    primary_run_for = selected_run_fors[0] if selected_run_fors else None
    effective_models = _effective_model_names(segment, selected_models)
    scenario_value = str(scenario or "").strip().lower()

    if (
        time_series_df is None
        or time_series_df.empty
        or not primary_run_for
        or not effective_models
        or not scenario_value
    ):
        return [], [], primary_run_for

    scoped_df = time_series_df[
        (time_series_df["Run For"] == primary_run_for)
        & (time_series_df["Scenario"].astype(str).str.strip().str.lower() == scenario_value)
        & (time_series_df["Model Name"].isin(effective_models))
    ]
    records = scoped_df.to_dict(orient="records")

    records_by_model: dict[str, list[dict]] = {}
    for row in records:
        records_by_model.setdefault(row.get("Model Name"), []).append(row)

    metric_rows: list[dict] = []
    chart_specs: list[dict] = []
    for model_name in effective_models:
        model_records = records_by_model.get(model_name, [])
        if not model_records:
            continue
        present_names = {str(row.get("MEV Name") or "").strip() for row in model_records if str(row.get("MEV Name") or "").strip()}
        # Include every MEV available for the model (both Raw and Transformed),
        # grouped by type then ordered by label.
        mev_names = sorted(
            present_names,
            key=lambda name: (_excel_mev_type_label(name), _resolve_mev_label(name, mev_label_mode).lower()),
        )
        for mev_name in mev_names:
            subset = [row for row in model_records if str(row.get("MEV Name") or "").strip() == mev_name]
            mev_label = _resolve_mev_label(mev_name, mev_label_mode)
            mev_type = _excel_mev_type_label(mev_name)
            record = _compute_saas_metric_record(model_name, mev_label, subset)
            if record is None:
                continue
            record["mev_type"] = mev_type
            metric_rows.append(record)
            chart_specs.append(_build_saas_chart_spec(model_name, mev_label, subset, mev_type=mev_type))

    return metric_rows, chart_specs, primary_run_for


def _excel_format_bool(value) -> str:
    if value is None:
        return ""
    return "Yes" if value else "No"


def _write_saas_excel_readme(ws, meta_lines: list[str]) -> None:
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    wrap = Alignment(wrap_text=True, vertical="top")

    ws["A1"] = "Scenario Analysis as a Service (SAAS) - MEV Metrics"
    ws["A1"].font = title_font

    row = 3
    for line in meta_lines:
        ws.cell(row=row, column=1, value=line)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Column").font = header_font
    ws.cell(row=row, column=2, value="Description").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=2).fill = header_fill
    row += 1
    for _key, header, description in SAAS_METRIC_COLUMNS:
        ws.cell(row=row, column=1, value=header)
        cell = ws.cell(row=row, column=2, value=description)
        cell.alignment = wrap
        row += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 90


def _write_saas_excel_metrics(ws, metric_rows: list[dict]) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D4ED8")

    if not metric_rows:
        ws["A1"] = "No metrics match the current filters and scenario."
        return

    for col_index, (_key, header, _description) in enumerate(SAAS_METRIC_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_index, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for row_index, record in enumerate(metric_rows, start=2):
        for col_index, (key, _header, _description) in enumerate(SAAS_METRIC_COLUMNS, start=1):
            value = record.get(key)
            cell = ws.cell(row=row_index, column=col_index)
            if key in _BOOLEAN_METRIC_KEYS:
                cell.value = _excel_format_bool(value)
            elif key in _DATE_METRIC_KEYS:
                cell.value = value
                if value is not None:
                    cell.number_format = "yyyy-mm-dd"
            elif key in _FLOAT_METRIC_KEYS:
                cell.value = None if value is None else float(value)
                cell.number_format = "0.0000"
            else:
                cell.value = value

    ws.freeze_panes = "D2"
    for col_index, (key, header, _description) in enumerate(SAAS_METRIC_COLUMNS, start=1):
        width = 16
        if key == "model":
            width = 18
        elif key == "mev_type":
            width = 16
        elif key == "mev":
            width = 30
        elif key == "minmax_daterange":
            width = 26
        elif len(header) > 22:
            width = min(46, len(header) + 2)
        ws.column_dimensions[get_column_letter(col_index)].width = width


# Series styling to mirror the dashboard's "Min-Max lines" presentation:
# History (solid navy), Projection (dashed faded blue), Historical Min / Max
# (dashed teal). Order matches the data columns added to each chart.
_SAAS_EXCEL_SERIES_STYLES = [
    ("1E3A8A", 28575, None),      # History - solid navy, ~2.25pt
    ("93C5FD", 22225, "dash"),    # Projection - dashed faded blue, ~1.75pt
    ("0F766E", 19050, "dash"),    # Historical Min - dashed teal, ~1.5pt
    ("0F766E", 19050, "dash"),    # Historical Max - dashed teal, ~1.5pt
]

_EXCEL_INVALID_SHEET_CHARS = set("[]:*?/\\")


def _excel_safe_sheet_title(name: str, used: set[str]) -> str:
    cleaned = "".join("_" if ch in _EXCEL_INVALID_SHEET_CHARS else ch for ch in str(name or "")).strip()
    cleaned = cleaned[:31] or "Model"
    candidate = cleaned
    counter = 1
    while candidate.lower() in used:
        suffix = f"_{counter}"
        candidate = cleaned[: 31 - len(suffix)] + suffix
        counter += 1
    used.add(candidate.lower())
    return candidate


def _style_saas_excel_chart(chart: LineChart) -> None:
    for series, (color, width, dash) in zip(chart.series, _SAAS_EXCEL_SERIES_STYLES):
        line = LineProperties(w=width, solidFill=color)
        if dash:
            line.prstDash = dash
        graphical_properties = GraphicalProperties()
        graphical_properties.line = line
        series.graphicalProperties = graphical_properties
        series.smooth = False


def _saas_excel_axis_text(rotation: int = 0, size: int = 900) -> RichText:
    """Axis text properties: small font, optional rotation (degrees)."""
    return RichText(
        bodyPr=RichTextProperties(rot=int(rotation * -60000), vert="horz"),
        p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=size)), endParaRPr=CharacterProperties(sz=size))],
    )


def _saas_excel_nice_step(value_range: float, target_ticks: int = 5) -> tuple[float, str]:
    """Pick a clean y-axis major unit (~target_ticks intervals) and a matching
    number format, using the 1/2/5 x 10^k "nice numbers" rule."""
    if value_range is None or value_range <= 0:
        return 1.0, "0.00"
    raw = value_range / target_ticks
    exponent = math.floor(math.log10(raw))
    normalized = raw / (10 ** exponent)
    if normalized < 1.5:
        nice, step_exponent = 1, exponent
    elif normalized < 3:
        nice, step_exponent = 2, exponent
    elif normalized < 7:
        nice, step_exponent = 5, exponent
    else:
        nice, step_exponent = 1, exponent + 1  # 10 x 10^exponent
    step = nice * (10 ** step_exponent)
    decimals = max(0, -step_exponent)
    number_format = "0" if decimals == 0 else "0." + "0" * decimals
    return step, number_format


def _write_saas_model_charts(ws_charts, ws_data, specs: list[dict], scenario_label: str, data_row: int) -> int:
    # Three charts per row; block size leaves a gap so titles, legends and tick
    # labels of adjacent charts never overlap.
    charts_per_row = 3
    chart_block_cols = 9
    chart_block_rows = 22

    for index, spec in enumerate(specs):
        history = spec["history"]
        projection = spec["projection"]

        mev_name = spec.get("y_title") or spec.get("title") or ""
        headers = ["MEV", "Period", "History", f"{scenario_label} (Projection)", "History Min", "History Max"]
        for col_index, header in enumerate(headers, start=1):
            ws_data.cell(row=data_row, column=col_index, value=header)
        first_data_row = data_row + 1
        for offset, label in enumerate(spec["labels"]):
            current = first_data_row + offset
            ws_data.cell(row=current, column=1, value=mev_name)
            ws_data.cell(row=current, column=2, value=label)
            ws_data.cell(row=current, column=3, value=history[offset])
            ws_data.cell(row=current, column=4, value=projection[offset])
            ws_data.cell(row=current, column=5, value=spec["hist_min"])
            ws_data.cell(row=current, column=6, value=spec["hist_max"])
        last_data_row = max(first_data_row, first_data_row + len(spec["labels"]) - 1)

        chart = LineChart()
        chart.title = spec["title"]
        chart.style = 2
        chart.height = 9
        chart.width = 15
        chart.y_axis.title = spec.get("y_title") or "MEV Value"
        chart.x_axis.title = None  # quarter labels are self-explanatory; frees space for the bottom legend
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        chart.y_axis.majorGridlines = None  # remove horizontal gridlines

        # Scale the y-axis so the historical Min/Max band fills the central half of
        # the plot: the Historical Min line sits at ~25% of the height and the
        # Historical Max at ~75% (min - 0.5*band .. max + 0.5*band). The axis is
        # widened to include any projection excursions so nothing is clipped; the
        # baseline (x-axis) sits at the axis minimum.
        plotted = [value for value in history if value is not None]
        plotted += [value for value in projection if value is not None]
        hist_low = spec.get("hist_min")
        hist_high = spec.get("hist_max")
        if hist_low is not None:
            plotted.append(hist_low)
        if hist_high is not None:
            plotted.append(hist_high)
        if plotted:
            data_min = min(plotted)
            data_max = max(plotted)
            if hist_low is not None and hist_high is not None and hist_high > hist_low:
                band = hist_high - hist_low
                axis_min = min(hist_low - 0.5 * band, data_min)
                axis_max = max(hist_high + 0.5 * band, data_max)
            else:
                pad = (data_max - data_min) * 0.1 or 1.0
                axis_min = data_min - pad
                axis_max = data_max + pad
            chart.y_axis.scaling.min = axis_min
            chart.y_axis.scaling.max = axis_max
            chart.y_axis.crosses = "min"

            # Clean tick frequency (~8 intervals) and matching rounding for labels.
            major_unit, axis_number_format = _saas_excel_nice_step(axis_max - axis_min, target_ticks=8)
            chart.y_axis.majorUnit = major_unit
            chart.y_axis.numFmt = axis_number_format
            chart.y_axis.txPr = _saas_excel_axis_text(rotation=0)

        # Thin out and rotate the period labels so they don't crowd each other.
        point_count = len(spec["labels"])
        label_skip = max(1, round(point_count / 12))
        chart.x_axis.tickLblSkip = label_skip
        chart.x_axis.tickMarkSkip = label_skip
        chart.x_axis.txPr = _saas_excel_axis_text(rotation=45)
        # Keep the period labels at the bottom of the plot rather than at the y=0
        # crossing (where they would sit in the middle of the data for MEVs that go
        # negative).
        chart.x_axis.tickLblPos = "low"

        # Pin the plot to the upper area and the legend just below it, with only a
        # small gap for the rotated x-axis labels. (chart.layout drives the inner
        # plot area; openpyxl copies it onto plot_area on write.)
        chart.layout = Layout(
            manualLayout=ManualLayout(
                layoutTarget="inner", xMode="edge", yMode="edge",
                x=0.10, y=0.08, w=0.86, h=0.64,
            )
        )
        if chart.legend is not None:
            chart.legend.position = "b"
            chart.legend.overlay = False
            chart.legend.layout = Layout(
                manualLayout=ManualLayout(
                    xMode="edge", yMode="edge",
                    x=0.10, y=0.84, w=0.86, h=0.10,
                )
            )

        data_ref = Reference(ws_data, min_col=3, max_col=6, min_row=data_row, max_row=last_data_row)
        cats_ref = Reference(ws_data, min_col=2, min_row=first_data_row, max_row=last_data_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        _style_saas_excel_chart(chart)

        grid_row = index // charts_per_row
        grid_col = index % charts_per_row
        anchor_col = get_column_letter(1 + grid_col * chart_block_cols)
        anchor_row = 1 + grid_row * chart_block_rows
        ws_charts.add_chart(chart, f"{anchor_col}{anchor_row}")

        data_row = last_data_row + 2

    return data_row


def _build_saas_excel_workbook(metric_rows: list[dict], chart_specs: list[dict], meta_lines: list[str], scenario_label: str) -> Workbook:
    workbook = Workbook()
    readme_ws = workbook.active
    readme_ws.title = "README"
    readme_ws.sheet_view.showGridLines = False
    _write_saas_excel_readme(readme_ws, meta_lines)

    # Metrics keeps gridlines; every other tab hides them.
    metrics_ws = workbook.create_sheet("Metrics")
    _write_saas_excel_metrics(metrics_ws, metric_rows)

    # Group charts by model so each model gets its own worksheet (preserve order).
    specs_by_model: dict[str, list[dict]] = {}
    model_order: list[str] = []
    for spec in chart_specs:
        model_name = spec.get("model") or "Model"
        if model_name not in specs_by_model:
            specs_by_model[model_name] = []
            model_order.append(model_name)
        specs_by_model[model_name].append(spec)

    if not model_order:
        fallback_ws = workbook.create_sheet("Charts")
        fallback_ws.sheet_view.showGridLines = False
        fallback_ws["A1"] = "No charts match the current filters and scenario."
        return workbook

    used_titles = {"readme", "metrics", "chart data"}
    model_sheets = {}
    for model_name in model_order:
        sheet = workbook.create_sheet(_excel_safe_sheet_title(model_name, used_titles))
        sheet.sheet_view.showGridLines = False
        model_sheets[model_name] = sheet

    chart_data_ws = workbook.create_sheet("Chart Data")
    chart_data_ws.sheet_state = "hidden"
    chart_data_ws.sheet_view.showGridLines = False

    data_row = 1
    for model_name in model_order:
        data_row = _write_saas_model_charts(
            model_sheets[model_name], chart_data_ws, specs_by_model[model_name], scenario_label, data_row,
        )

    return workbook


def register_callbacks(app) -> None:
    """Register SAAS top-bar and chart callbacks."""

    segment_labels = {option["value"]: option["label"] for option in layout.SEGMENT_NAME_OPTIONS}
    subnav_view_labels = {option["value"]: option["label"] for option in layout.SUBNAV_VIEW_OPTIONS}
    reference_line_labels = {option["value"]: option["label"] for option in layout.REFERENCE_LINES_OPTIONS}
    mev_label_mode_labels = {option["value"]: option["label"] for option in layout.MEV_LABEL_MODE_OPTIONS}
    run_for_labels = {option["value"]: option["label"] for option in layout.RUN_FOR_OPTIONS}
    run_for_values = [option["value"] for option in layout.RUN_FOR_OPTIONS]

    @app.callback(
        Output(layout.MODEL_NAME_ID, "value", allow_duplicate=True),
        Output(layout.MODEL_NAME_SELECT_ALL_ID, "value"),
        Input(layout.MODEL_NAME_SELECT_ALL_ID, "value"),
        Input(layout.MODEL_NAME_ID, "value"),
        State(layout.MODEL_NAME_ID, "options"),
        prevent_initial_call=True,
    )
    def sync_saas_model_selection(select_all_value, models_value, models_options):
        all_names = [option["value"] for option in models_options if option.get("value")]
        if ctx.triggered_id == layout.MODEL_NAME_SELECT_ALL_ID:
            if "all" in (select_all_value or []):
                return all_names, no_update
            return [], no_update
        select_all = ["all"] if all_names and set(models_value or []) == set(all_names) else []
        return no_update, select_all

    @app.callback(
        Output(layout.SEGMENT_NAME_ID, "disabled"),
        Output(layout.SEGMENT_TOGGLE_ID, "disabled"),
        Output(layout.MODEL_NAME_SELECT_ALL_ID, "options"),
        Output(layout.MODEL_NAME_ID, "options"),
        Output(layout.MODEL_NAME_ID, "value"),
        Output(layout.MODEL_NAME_TOGGLE_ID, "children"),
        Output(layout.MODEL_NAME_TOGGLE_ID, "disabled"),
        Output(layout.FILTER_HELP_ID, "children"),
        Input(layout.SEGMENT_NAME_ID, "value"),
        Input(layout.MODEL_NAME_ID, "value"),
    )
    def sync_saas_filter_controls(segment, selected_models):
        segment_active = _is_segment_active(segment)
        all_options = _model_options_for_filters(None)
        all_option_values = [option["value"] for option in all_options]
        current_values = _normalize_selected_models(selected_models)

        option_values = set(all_option_values)
        current_values = [value for value in current_values if value in option_values]
        has_specific_model_selection = 0 < len(current_values) < len(all_option_values)
        select_all_options = [{"label": "All", "value": "all", "disabled": segment_active}]

        if segment_active:
            model_options = _model_options_for_filters(None, disabled=True)
            toggle_label = "Disabled while Segment is selected"
            help_text = "Segment filtering is active. Reset Segment to All to select specific models."
        else:
            model_options = _model_options_for_filters(None, disabled=False)
            toggle_label = _model_toggle_label(current_values, all_options, False)
            if has_specific_model_selection:
                help_text = "Specific Models filtering is active. Clear the model selection to use the Segment filter."
            else:
                help_text = ""

        return (
            has_specific_model_selection,
            has_specific_model_selection,
            select_all_options,
            model_options,
            current_values,
            toggle_label,
            segment_active,
            help_text,
        )

    @app.callback(
        Output(layout.RUN_FOR_MENU_ID, "className"),
        Input(layout.RUN_FOR_TOGGLE_ID, "n_clicks"),
        State(layout.RUN_FOR_MENU_ID, "className"),
        prevent_initial_call=True,
    )
    def toggle_run_for_menu(_n_clicks, current_class):
        return _toggle_menu_class(current_class, base_class="checkbox-dropdown-menu single-select-menu")

    @app.callback(
        Output(layout.COMPARE_AGAINST_MENU_ID, "className"),
        Input(layout.COMPARE_AGAINST_TOGGLE_ID, "n_clicks"),
        State(layout.COMPARE_AGAINST_MENU_ID, "className"),
        prevent_initial_call=True,
    )
    def toggle_compare_against_menu(_n_clicks, current_class):
        return _toggle_menu_class(current_class, base_class="checkbox-dropdown-menu")

    @app.callback(
        Output(layout.SEGMENT_MENU_ID, "className"),
        Input(layout.SEGMENT_TOGGLE_ID, "n_clicks"),
        State(layout.SEGMENT_MENU_ID, "className"),
        State(layout.SEGMENT_TOGGLE_ID, "disabled"),
        prevent_initial_call=True,
    )
    def toggle_segment_menu(_n_clicks, current_class, is_disabled):
        if is_disabled:
            return "checkbox-dropdown-menu single-select-menu"
        return _toggle_menu_class(current_class, base_class="checkbox-dropdown-menu single-select-menu")

    @app.callback(
        Output(layout.SUBNAV_VIEW_MENU_ID, "className"),
        Input(layout.SUBNAV_VIEW_TOGGLE_ID, "n_clicks"),
        State(layout.SUBNAV_VIEW_MENU_ID, "className"),
        prevent_initial_call=True,
    )
    def toggle_subnav_view_menu(_n_clicks, current_class):
        return _toggle_menu_class(current_class, base_class="checkbox-dropdown-menu single-select-menu")

    @app.callback(
        Output(layout.REFERENCE_LINES_MENU_ID, "className"),
        Input(layout.REFERENCE_LINES_TOGGLE_ID, "n_clicks"),
        State(layout.REFERENCE_LINES_MENU_ID, "className"),
        prevent_initial_call=True,
    )
    def toggle_reference_lines_menu(_n_clicks, current_class):
        return _toggle_menu_class(current_class, base_class="checkbox-dropdown-menu single-select-menu")

    @app.callback(
        Output(layout.MEV_LABEL_MODE_MENU_ID, "className"),
        Input(layout.MEV_LABEL_MODE_TOGGLE_ID, "n_clicks"),
        State(layout.MEV_LABEL_MODE_MENU_ID, "className"),
        prevent_initial_call=True,
    )
    def toggle_mev_label_mode_menu(_n_clicks, current_class):
        return _toggle_menu_class(current_class, base_class="checkbox-dropdown-menu single-select-menu")

    @app.callback(
        Output({"type": layout.MODEL_SCENARIO_MENU_TYPE, "model": MATCH}, "className"),
        Input({"type": layout.MODEL_SCENARIO_TOGGLE_TYPE, "model": MATCH}, "n_clicks"),
        State({"type": layout.MODEL_SCENARIO_MENU_TYPE, "model": MATCH}, "className"),
        prevent_initial_call=True,
    )
    def toggle_model_scenario_menu(_n_clicks, current_class):
        return _toggle_menu_class(current_class, base_class="checkbox-dropdown-menu")

    @app.callback(
        Output({"type": layout.MODEL_MEV_TYPE_MENU_TYPE, "model": MATCH}, "className"),
        Input({"type": layout.MODEL_MEV_TYPE_TOGGLE_TYPE, "model": MATCH}, "n_clicks"),
        State({"type": layout.MODEL_MEV_TYPE_MENU_TYPE, "model": MATCH}, "className"),
        prevent_initial_call=True,
    )
    def toggle_model_mev_type_menu(_n_clicks, current_class):
        return _toggle_menu_class(current_class, base_class="checkbox-dropdown-menu single-select-menu")

    @app.callback(
        Output(layout.MODEL_NAME_MENU_ID, "className"),
        Input(layout.MODEL_NAME_TOGGLE_ID, "n_clicks"),
        State(layout.MODEL_NAME_MENU_ID, "className"),
        State(layout.MODEL_NAME_TOGGLE_ID, "disabled"),
        prevent_initial_call=True,
    )
    def toggle_model_menu(_n_clicks, current_class, is_disabled):
        if is_disabled:
            return "checkbox-dropdown-menu"
        return _toggle_menu_class(current_class, base_class="checkbox-dropdown-menu")

    @app.callback(
        Output({"type": layout.MODEL_MEV_MENU_TYPE, "model": MATCH}, "className"),
        Input({"type": layout.MODEL_MEV_TOGGLE_TYPE, "model": MATCH}, "n_clicks"),
        State({"type": layout.MODEL_MEV_MENU_TYPE, "model": MATCH}, "className"),
        prevent_initial_call=True,
    )
    def toggle_model_mev_menu(_n_clicks, current_class):
        return _toggle_menu_class(current_class, base_class="checkbox-dropdown-menu")

    @app.callback(
        Output(layout.RUN_FOR_TOGGLE_ID, "children"),
        Output({"type": shared_filters.SINGLE_SELECT_OPTION_ID, "filter": layout.RUN_FOR_FILTER_KEY, "value": ALL}, "className"),
        Input(layout.RUN_FOR_ID, "value"),
        State({"type": shared_filters.SINGLE_SELECT_OPTION_ID, "filter": layout.RUN_FOR_FILTER_KEY, "value": ALL}, "id"),
    )
    def sync_run_for_shell(value, option_ids):
        label = run_for_labels.get(value, value or "Select")
        return label, _single_select_option_classes(value, option_ids)

    @app.callback(
        Output(layout.COMPARE_AGAINST_ID, "options"),
        Output(layout.COMPARE_AGAINST_ID, "value", allow_duplicate=True),
        Output(layout.COMPARE_AGAINST_PREV_STORE_ID, "data"),
        Input(layout.RUN_FOR_ID, "value"),
        Input(layout.COMPARE_AGAINST_ID, "value"),
        Input(layout.REFERENCE_LINES_ID, "value"),
        State(layout.COMPARE_AGAINST_PREV_STORE_ID, "data"),
        prevent_initial_call=True,
    )
    def sync_compare_against_selection(run_for_value, compare_against_values, reference_lines, prev_values):
        selected_run_fors = _normalize_selected_run_fors(run_for_value)
        selected_run_for = selected_run_fors[0] if selected_run_fors else None
        options = _build_compare_against_options(selected_run_for)
        option_values = {option["value"] for option in options}
        triggered_ids = {triggered["prop_id"].split(".")[0] for triggered in ctx.triggered}

        if layout.REFERENCE_LINES_ID in triggered_ids and reference_lines == "monitoring":
            result = [layout.COMPARE_AGAINST_NONE_VALUE]
            return options, result, result

        raw_values = [value for value in _normalize_multi_values(compare_against_values) if value in option_values]
        prev_values = [value for value in _normalize_multi_values(prev_values) if value in option_values]

        if layout.COMPARE_AGAINST_ID in triggered_ids:
            # Selecting "None" or a reporting cycle should deselect the other -
            # keep only whichever option was just checked.
            added = [value for value in raw_values if value not in prev_values]
            if added:
                result = [added[-1]]
            elif raw_values:
                result = raw_values
            else:
                result = [layout.COMPARE_AGAINST_NONE_VALUE]
        else:
            result = prev_values if prev_values else [layout.COMPARE_AGAINST_NONE_VALUE]

        return options, result, result

    @app.callback(
        Output(layout.COMPARE_AGAINST_TOGGLE_ID, "children"),
        Input(layout.COMPARE_AGAINST_ID, "value"),
        State(layout.RUN_FOR_ID, "value"),
    )
    def sync_compare_against_toggle(compare_against_values, run_for_value):
        selected_run_fors = _normalize_selected_run_fors(run_for_value)
        selected_run_for = selected_run_fors[0] if selected_run_fors else None
        return _compare_against_toggle_label(compare_against_values, selected_run_for)

    @app.callback(
        Output(layout.SEGMENT_TOGGLE_ID, "children"),
        Output({"type": shared_filters.SINGLE_SELECT_OPTION_ID, "filter": layout.SEGMENT_FILTER_KEY, "value": ALL}, "className"),
        Input(layout.SEGMENT_NAME_ID, "value"),
        State({"type": shared_filters.SINGLE_SELECT_OPTION_ID, "filter": layout.SEGMENT_FILTER_KEY, "value": ALL}, "id"),
    )
    def sync_segment_shell(value, option_ids):
        label = segment_labels.get(value, value or "Select")
        return label, _single_select_option_classes(value, option_ids)

    @app.callback(
        Output(layout.SUBNAV_VIEW_TOGGLE_ID, "children"),
        Output({"type": shared_filters.SINGLE_SELECT_OPTION_ID, "filter": layout.SUBNAV_VIEW_FILTER_KEY, "value": ALL}, "className"),
        Input(layout.SUBNAV_VIEW_ID, "value"),
        State({"type": shared_filters.SINGLE_SELECT_OPTION_ID, "filter": layout.SUBNAV_VIEW_FILTER_KEY, "value": ALL}, "id"),
    )
    def sync_subnav_view_shell(value, option_ids):
        label = subnav_view_labels.get(value, value or "Select")
        return label, _single_select_option_classes(value, option_ids)

    @app.callback(
        Output(layout.REFERENCE_LINES_TOGGLE_ID, "children"),
        Output({"type": shared_filters.SINGLE_SELECT_OPTION_ID, "filter": layout.REFERENCE_LINES_FILTER_KEY, "value": ALL}, "className"),
        Input(layout.REFERENCE_LINES_ID, "value"),
        State({"type": shared_filters.SINGLE_SELECT_OPTION_ID, "filter": layout.REFERENCE_LINES_FILTER_KEY, "value": ALL}, "id"),
    )
    def sync_reference_lines_shell(value, option_ids):
        label = reference_line_labels.get(value, value or "Select")
        return label, _single_select_option_classes(value, option_ids)

    @app.callback(
        Output(layout.MEV_LABEL_MODE_TOGGLE_ID, "children"),
        Output({"type": shared_filters.SINGLE_SELECT_OPTION_ID, "filter": layout.MEV_LABEL_MODE_FILTER_KEY, "value": ALL}, "className"),
        Input(layout.MEV_LABEL_MODE_ID, "value"),
        State({"type": shared_filters.SINGLE_SELECT_OPTION_ID, "filter": layout.MEV_LABEL_MODE_FILTER_KEY, "value": ALL}, "id"),
    )
    def sync_mev_label_mode_shell(value, option_ids):
        label_mode_value = _normalize_mev_label_mode(value)
        label = mev_label_mode_labels.get(label_mode_value, label_mode_value or "Select")
        return label, _single_select_option_classes(label_mode_value, option_ids)

    @app.callback(
        Output({"type": shared_filters.SINGLE_SELECT_OPTION_ID, "filter": layout.HISTORICAL_STATS_FILTER_KEY, "value": ALL}, "className"),
        Input(layout.HISTORICAL_STATS_ID, "value"),
        State({"type": shared_filters.SINGLE_SELECT_OPTION_ID, "filter": layout.HISTORICAL_STATS_FILTER_KEY, "value": ALL}, "id"),
    )
    def sync_historical_stats_toggle(value, option_ids):
        return _single_select_option_classes(value, option_ids)

    @app.callback(
        Output({"type": layout.MODEL_SCENARIO_FILTER_TYPE, "model": MATCH}, "value", allow_duplicate=True),
        Output({"type": layout.MODEL_SCENARIO_SELECT_ALL_TYPE, "model": MATCH}, "value"),
        Input({"type": layout.MODEL_SCENARIO_SELECT_ALL_TYPE, "model": MATCH}, "value"),
        Input({"type": layout.MODEL_SCENARIO_FILTER_TYPE, "model": MATCH}, "value"),
        State({"type": layout.MODEL_SCENARIO_FILTER_TYPE, "model": MATCH}, "options"),
        State(layout.REFERENCE_LINES_ID, "value"),
        prevent_initial_call=True,
    )
    def sync_model_scenario_selection(select_all_value, selected_scenarios, scenario_options, reference_lines):
        all_scenario_values = [option["value"] for option in scenario_options if option.get("value")]
        if (reference_lines or layout.DEFAULT_REFERENCE_LINES) == "monitoring":
            selected_value = _single_selected_scenario(selected_scenarios, scenario_options or [])
            return selected_value, []
        if ctx.triggered_id and ctx.triggered_id.get("type") == layout.MODEL_SCENARIO_SELECT_ALL_TYPE:
            if "all" in (select_all_value or []):
                return all_scenario_values, no_update
            return [], no_update
        select_all = ["all"] if all_scenario_values and set(selected_scenarios or []) == set(all_scenario_values) else []
        return no_update, select_all

    @app.callback(
        Output({"type": layout.MODEL_SCENARIO_TOGGLE_TYPE, "model": MATCH}, "children"),
        Input({"type": layout.MODEL_SCENARIO_FILTER_TYPE, "model": MATCH}, "value"),
        State({"type": layout.MODEL_SCENARIO_FILTER_TYPE, "model": MATCH}, "options"),
    )
    def sync_model_scenario_toggle(selected_scenarios, scenario_options):
        return _scenario_toggle_label(selected_scenarios, scenario_options or [])

    @app.callback(
        Output({"type": layout.MODEL_SCENARIO_FILTER_TYPE, "model": MATCH}, "value", allow_duplicate=True),
        Output({"type": layout.MODEL_SCENARIO_MENU_TYPE, "model": MATCH}, "className", allow_duplicate=True),
        Input({"type": layout.MODEL_SCENARIO_OPTION_TYPE, "model": MATCH, "value": ALL}, "n_clicks"),
        prevent_initial_call=True,
    )
    def select_model_scenario_option(_clicks):
        triggered = ctx.triggered_id
        if not isinstance(triggered, dict):
            return no_update, no_update
        return triggered.get("value"), "checkbox-dropdown-menu"

    @app.callback(
        Output({"type": layout.MODEL_SCENARIO_OPTION_TYPE, "model": MATCH, "value": ALL}, "className"),
        Input({"type": layout.MODEL_SCENARIO_FILTER_TYPE, "model": MATCH}, "value"),
        State({"type": layout.MODEL_SCENARIO_OPTION_TYPE, "model": MATCH, "value": ALL}, "id"),
    )
    def sync_model_scenario_option_classes(selected_scenarios, option_ids):
        selected_value = _single_selected_scenario(selected_scenarios, [
            {"value": option_id.get("value")}
            for option_id in (option_ids or [])
        ])
        return _single_select_option_classes(selected_value, option_ids or [])

    @app.callback(
        Output({"type": layout.MODEL_MEV_TYPE_TOGGLE_TYPE, "model": MATCH}, "children"),
        Output({"type": layout.MODEL_MEV_TYPE_OPTION_TYPE, "model": MATCH, "value": ALL}, "className"),
        Input({"type": layout.MODEL_MEV_TYPE_FILTER_TYPE, "model": MATCH}, "value"),
        State({"type": layout.MODEL_MEV_TYPE_OPTION_TYPE, "model": MATCH, "value": ALL}, "id"),
    )
    def sync_model_mev_type_toggle(selected_mev_mode, option_ids):
        normalized_mode = _normalize_selected_mev_mode(selected_mev_mode)
        return _mev_type_toggle_label(normalized_mode, layout.MEV_TYPE_OPTIONS), _single_select_option_classes(normalized_mode, option_ids)

    @app.callback(
        Output({"type": layout.MODEL_MEV_TYPE_FILTER_TYPE, "model": MATCH}, "value", allow_duplicate=True),
        Output({"type": layout.MODEL_MEV_TYPE_MENU_TYPE, "model": MATCH}, "className", allow_duplicate=True),
        Input({"type": layout.MODEL_MEV_TYPE_OPTION_TYPE, "model": MATCH, "value": ALL}, "n_clicks"),
        prevent_initial_call=True,
    )
    def select_model_mev_type(_clicks):
        triggered = ctx.triggered_id
        if not triggered:
            return no_update, no_update
        return triggered["value"], "checkbox-dropdown-menu single-select-menu"

    @app.callback(
        Output({"type": layout.MODEL_MEV_LABEL_TYPE, "model": MATCH}, "children"),
        Output({"type": layout.MODEL_MEV_SINGLE_VALUE_TYPE, "model": MATCH}, "options"),
        Output({"type": layout.MODEL_MEV_SINGLE_VALUE_TYPE, "model": MATCH}, "value", allow_duplicate=True),
        Output({"type": layout.MODEL_MEV_SINGLE_OPTIONS_TYPE, "model": MATCH}, "children"),
        Output({"type": layout.MODEL_MEV_FILTER_TYPE, "model": MATCH}, "options"),
        Output({"type": layout.MODEL_MEV_FILTER_TYPE, "model": MATCH}, "value", allow_duplicate=True),
        Output({"type": layout.MODEL_MEV_SELECT_ALL_TYPE, "model": MATCH}, "value"),
        Output({"type": layout.MODEL_MEV_SINGLE_SECTION_TYPE, "model": MATCH}, "className"),
        Output({"type": layout.MODEL_MEV_MULTI_SECTION_TYPE, "model": MATCH}, "className"),
        Input({"type": layout.MODEL_MEV_SELECT_ALL_TYPE, "model": MATCH}, "value"),
        Input({"type": layout.MODEL_MEV_FILTER_TYPE, "model": MATCH}, "value"),
        Input({"type": layout.MODEL_MEV_TYPE_FILTER_TYPE, "model": MATCH}, "value"),
        Input({"type": layout.MODEL_MEV_SINGLE_VALUE_TYPE, "model": MATCH}, "value"),
        State(layout.RUN_FOR_ID, "value"),
        State(layout.COMPARE_AGAINST_ID, "value"),
        State(layout.SUBNAV_VIEW_ID, "value"),
        State(layout.MEV_LABEL_MODE_ID, "value"),
        prevent_initial_call=True,
    )
    def sync_model_mev_selection(
        select_all_value,
        selected_mevs_multi,
        selected_mev_mode,
        selected_mev_single,
        run_for,
        compare_against,
        snapshot_period,
        mev_label_mode,
    ):
        model_name = (ctx.triggered_id or {}).get("model")
        normalized_mode = _normalize_selected_mev_mode(selected_mev_mode)
        single_mev_options = _build_model_mev_options_for_mode(
            model_name,
            run_for,
            snapshot_period,
            mev_label_mode,
            "family",
            compare_against,
        )
        multi_mev_options = _build_model_mev_options_for_mode(
            model_name,
            run_for,
            snapshot_period,
            mev_label_mode,
            normalized_mode,
            compare_against,
        )

        single_option_values = [option["value"] for option in single_mev_options if option.get("value")]
        next_single_value = next(
            (value for value in _normalize_selected_mevs(selected_mev_single) if value in single_option_values),
            single_option_values[0] if single_option_values else "",
        )
        multi_option_values = [option["value"] for option in multi_mev_options if option.get("value")]
        valid_selected_multi = [value for value in _normalize_selected_mevs(selected_mevs_multi) if value in multi_option_values]

        triggered_type = (ctx.triggered_id or {}).get("type")
        if normalized_mode == "family":
            next_multi_value = valid_selected_multi
            next_select_all = []
        elif triggered_type == layout.MODEL_MEV_SELECT_ALL_TYPE:
            if "all" in (select_all_value or []):
                next_multi_value = multi_option_values
            else:
                next_multi_value = []
            next_select_all = no_update
        elif triggered_type == layout.MODEL_MEV_FILTER_TYPE:
            next_multi_value = valid_selected_multi
            next_select_all = ["all"] if multi_option_values and set(next_multi_value) == set(multi_option_values) else []
        else:
            next_multi_value = valid_selected_multi or multi_option_values
            next_select_all = ["all"] if multi_option_values and set(next_multi_value) == set(multi_option_values) else []

        return (
            _mev_picker_label(normalized_mode),
            single_mev_options,
            next_single_value,
            _build_single_mev_option_buttons(single_mev_options, next_single_value, model_name),
            multi_mev_options,
            next_multi_value,
            next_select_all,
            _mev_picker_section_class(visible=normalized_mode == "family"),
            _mev_picker_section_class(visible=normalized_mode != "family"),
        )

    @app.callback(
        Output({"type": layout.MODEL_MEV_TOGGLE_TYPE, "model": MATCH}, "children"),
        Input({"type": layout.MODEL_MEV_TYPE_FILTER_TYPE, "model": MATCH}, "value"),
        Input({"type": layout.MODEL_MEV_SINGLE_VALUE_TYPE, "model": MATCH}, "value"),
        Input({"type": layout.MODEL_MEV_FILTER_TYPE, "model": MATCH}, "value"),
        State({"type": layout.MODEL_MEV_SINGLE_VALUE_TYPE, "model": MATCH}, "options"),
        State({"type": layout.MODEL_MEV_FILTER_TYPE, "model": MATCH}, "options"),
    )
    def sync_model_mev_toggle(selected_mev_mode, selected_mev_single, selected_mevs_multi, single_mev_options, multi_mev_options):
        return _mev_toggle_label(
            selected_mev_mode,
            selected_mev_single,
            selected_mevs_multi,
            single_mev_options or [],
            multi_mev_options or [],
        )

    @app.callback(
        Output({"type": layout.MODEL_MEV_SINGLE_VALUE_TYPE, "model": MATCH}, "value", allow_duplicate=True),
        Output({"type": layout.MODEL_MEV_MENU_TYPE, "model": MATCH}, "className", allow_duplicate=True),
        Input({"type": layout.MODEL_MEV_SINGLE_OPTION_TYPE, "model": MATCH, "value": ALL}, "n_clicks"),
        prevent_initial_call=True,
    )
    def select_model_mev_single_value(_clicks):
        triggered = ctx.triggered_id
        if not triggered:
            return no_update, no_update
        return triggered["value"], "checkbox-dropdown-menu"

    @app.callback(
        Output(layout.RUN_FOR_ID, "value", allow_duplicate=True),
        Output(layout.RUN_FOR_MENU_ID, "className", allow_duplicate=True),
        Input({"type": shared_filters.SINGLE_SELECT_OPTION_ID, "filter": layout.RUN_FOR_FILTER_KEY, "value": ALL}, "n_clicks"),
        prevent_initial_call=True,
    )
    def select_run_for(_clicks):
        triggered = ctx.triggered_id
        if not triggered:
            return no_update, no_update
        return triggered["value"], "checkbox-dropdown-menu single-select-menu"

    @app.callback(
        Output(layout.SEGMENT_NAME_ID, "value", allow_duplicate=True),
        Output(layout.SEGMENT_MENU_ID, "className", allow_duplicate=True),
        Input({"type": shared_filters.SINGLE_SELECT_OPTION_ID, "filter": layout.SEGMENT_FILTER_KEY, "value": ALL}, "n_clicks"),
        prevent_initial_call=True,
    )
    def select_segment(_clicks):
        triggered = ctx.triggered_id
        if not triggered:
            return no_update, no_update
        return triggered["value"], "checkbox-dropdown-menu single-select-menu"

    @app.callback(
        Output(layout.SUBNAV_VIEW_ID, "value", allow_duplicate=True),
        Output(layout.SUBNAV_VIEW_MENU_ID, "className", allow_duplicate=True),
        Input({"type": shared_filters.SINGLE_SELECT_OPTION_ID, "filter": layout.SUBNAV_VIEW_FILTER_KEY, "value": ALL}, "n_clicks"),
        prevent_initial_call=True,
    )
    def select_subnav_view(_clicks):
        triggered = ctx.triggered_id
        if not triggered:
            return no_update, no_update
        return triggered["value"], "checkbox-dropdown-menu single-select-menu"

    @app.callback(
        Output(layout.REFERENCE_LINES_ID, "value", allow_duplicate=True),
        Output(layout.REFERENCE_LINES_MENU_ID, "className", allow_duplicate=True),
        Input({"type": shared_filters.SINGLE_SELECT_OPTION_ID, "filter": layout.REFERENCE_LINES_FILTER_KEY, "value": ALL}, "n_clicks"),
        prevent_initial_call=True,
    )
    def select_reference_lines(_clicks):
        triggered = ctx.triggered_id
        if not triggered:
            return no_update, no_update
        return triggered["value"], "checkbox-dropdown-menu single-select-menu"

    @app.callback(
        Output(layout.MEV_LABEL_MODE_ID, "value", allow_duplicate=True),
        Output(layout.MEV_LABEL_MODE_MENU_ID, "className", allow_duplicate=True),
        Input({"type": shared_filters.SINGLE_SELECT_OPTION_ID, "filter": layout.MEV_LABEL_MODE_FILTER_KEY, "value": ALL}, "n_clicks"),
        prevent_initial_call=True,
    )
    def select_mev_label_mode(_clicks):
        triggered = ctx.triggered_id
        if not triggered:
            return no_update, no_update
        return triggered["value"], "checkbox-dropdown-menu single-select-menu"

    @app.callback(
        Output(layout.HISTORICAL_STATS_ID, "value", allow_duplicate=True),
        Input({"type": shared_filters.SINGLE_SELECT_OPTION_ID, "filter": layout.HISTORICAL_STATS_FILTER_KEY, "value": ALL}, "n_clicks"),
        prevent_initial_call=True,
    )
    def select_historical_stats_toggle(_clicks):
        triggered = ctx.triggered_id
        if not triggered:
            return no_update
        return triggered["value"]

    @app.callback(
        Output(layout.HISTORICAL_STATS_FILTER_ID, "className"),
        Input(layout.SUBNAV_VIEW_ID, "value"),
    )
    def sync_historical_stats_filter_visibility(snapshot_period):
        base_class = "monitoring-filter saas-historical-stats-filter"
        if _normalize_snapshot_period(snapshot_period) != "history":
            return f"{base_class} is-hidden"
        return base_class

    @app.callback(
        Output(layout.DOWNLOAD_DATA_ID, "data"),
        Input(layout.DOWNLOAD_REPORT_ID, "n_clicks"),
        State(layout.APPLIED_FILTERS_STORE_ID, "data"),
        prevent_initial_call=True,
    )
    def download_saas_report(_n_clicks, applied):
        applied = applied or {}
        run_for = applied.get("run_for")
        compare_against = applied.get("compare_against")
        segment = applied.get("segment")
        selected_models = applied.get("selected_models")
        snapshot_period = applied.get("snapshot_period")
        reference_lines = applied.get("reference_lines")
        mev_label_mode = applied.get("mev_label_mode")
        sections = _build_report_figures(
            run_for, compare_against, segment, selected_models,
            snapshot_period, reference_lines, mev_label_mode,
        )

        effective_models = _effective_model_names(segment, selected_models)
        if _is_segment_active(segment):
            scope_label = f"Segment: {layout.format_segment_label(segment)}"
        elif effective_models:
            scope_label = f"Models: {', '.join(effective_models)}"
        else:
            scope_label = "Models: None selected"

        meta_lines = [
            f"Reporting Cycle: {_run_for_meta_label(run_for)}",
            f"Compare To: {_compare_against_toggle_label(compare_against, _primary_run_for_value(run_for))}",
            scope_label,
            f"Snapshot Period: {subnav_view_labels.get(_normalize_snapshot_period(snapshot_period), snapshot_period)}",
            f"Reference Lines: {reference_line_labels.get(reference_lines or layout.DEFAULT_REFERENCE_LINES, reference_lines)}",
            f"MEV Label: {mev_label_mode_labels.get(_normalize_mev_label_mode(mev_label_mode), mev_label_mode)}",
        ]

        html_doc = _build_saas_report_html(sections, meta_lines)
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        prefix = _run_for_filename_prefix(run_for)
        return dcc.send_string(html_doc, filename=f"{prefix}-saas-charts-{timestamp}.html")

    @app.callback(
        Output(layout.EXCEL_MODAL_ID, "className"),
        Input(layout.EXCEL_OPEN_ID, "n_clicks"),
        Input(layout.EXCEL_CANCEL_ID, "n_clicks"),
        Input(layout.EXCEL_GENERATE_ID, "n_clicks"),
        prevent_initial_call=True,
    )
    def toggle_excel_modal(_open_clicks, _cancel_clicks, _generate_clicks):
        if ctx.triggered_id == layout.EXCEL_OPEN_ID:
            return "saas-modal-overlay is-open"
        return "saas-modal-overlay"

    @app.callback(
        Output(layout.EXCEL_DOWNLOAD_DATA_ID, "data"),
        Input(layout.EXCEL_GENERATE_ID, "n_clicks"),
        State(layout.EXCEL_SCENARIO_ID, "value"),
        State(layout.RUN_FOR_ID, "value"),
        State(layout.SEGMENT_NAME_ID, "value"),
        State(layout.MODEL_NAME_ID, "value"),
        State(layout.MEV_LABEL_MODE_ID, "value"),
        prevent_initial_call=True,
    )
    def download_saas_excel(_n_clicks, scenario, run_for, segment, selected_models, mev_label_mode):
        metric_rows, chart_specs, primary_run_for = _compute_saas_metrics(
            run_for, segment, selected_models, mev_label_mode, scenario,
        )
        scenario_value = str(scenario or "").strip().lower()
        scenario_label = SAAS_SCENARIO_LABEL_MAP.get(scenario_value, scenario_value.replace("_", " ").title() or "—")
        effective_models = _effective_model_names(segment, selected_models)
        if _is_segment_active(segment):
            scope_label = f"Segment: {layout.format_segment_label(segment)}"
        elif effective_models:
            scope_label = f"Models: {', '.join(effective_models)}"
        else:
            scope_label = "Models: None selected"

        meta_lines = [
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            f"Reporting Cycle: {primary_run_for or '—'}",
            f"Scenario (Severely Adverse): {scenario_label}",
            scope_label,
            f"MEV Label: {mev_label_mode_labels.get(_normalize_mev_label_mode(mev_label_mode), mev_label_mode)}",
            "History = Quarter <= 0; Projection = Quarter > 0. Metrics use the primary Reporting Cycle and the full history + projection (independent of the on-screen Snapshot Period). Standard deviation is population (ddof=0).",
        ]

        workbook = _build_saas_excel_workbook(metric_rows, chart_specs, meta_lines, scenario_label)
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        prefix = _run_for_filename_prefix(run_for)
        return dcc.send_bytes(lambda buffer: workbook.save(buffer), filename=f"{prefix}-saas-metrics-{timestamp}.xlsx")

    @app.callback(
        Output(layout.APPLIED_FILTERS_STORE_ID, "data"),
        Input(layout.APPLY_FILTERS_ID, "n_clicks"),
        State(layout.RUN_FOR_ID, "value"),
        State(layout.COMPARE_AGAINST_ID, "value"),
        State(layout.SEGMENT_NAME_ID, "value"),
        State(layout.MODEL_NAME_ID, "value"),
        State(layout.SUBNAV_VIEW_ID, "value"),
        State(layout.REFERENCE_LINES_ID, "value"),
        State(layout.MEV_LABEL_MODE_ID, "value"),
        State(layout.HISTORICAL_STATS_ID, "value"),
        prevent_initial_call=True,
    )
    def apply_saas_filters(
        _n_clicks,
        run_for,
        compare_against,
        segment,
        selected_models,
        snapshot_period,
        reference_lines,
        mev_label_mode,
        show_historical_statistics_values,
    ):
        """Snapshot the current top filters so the content renders only on Apply."""
        return {
            "run_for": run_for,
            "compare_against": compare_against,
            "segment": segment,
            "selected_models": selected_models,
            "snapshot_period": snapshot_period,
            "reference_lines": reference_lines,
            "mev_label_mode": mev_label_mode,
            "historical_stats": show_historical_statistics_values,
        }

    @app.callback(
        Output(layout.SUBNAV_MODELS_ID, "children"),
        Input(layout.APPLIED_FILTERS_STORE_ID, "data"),
        prevent_initial_call=True,
    )
    def render_saas_subnav_models(applied):
        applied = applied or {}
        return _build_subnav_models(applied.get("segment"), applied.get("selected_models"))

    @app.callback(
        Output(layout.MEV_MODEL_PANELS_ID, "children"),
        Input(layout.APPLIED_FILTERS_STORE_ID, "data"),
        prevent_initial_call=True,
    )
    def render_saas_mev_chart(applied):
        applied = applied or {}
        run_for = applied.get("run_for")
        compare_against = applied.get("compare_against")
        segment = applied.get("segment")
        selected_models = applied.get("selected_models")
        snapshot_period = applied.get("snapshot_period")
        reference_lines = applied.get("reference_lines")
        mev_label_mode = applied.get("mev_label_mode")
        show_historical_statistics_values = applied.get("historical_stats")
        selected_run_fors = _normalize_selected_run_fors(run_for)
        scoped_run_fors = _scoped_run_for_values(run_for, compare_against)
        effective_models = _effective_model_names(segment, selected_models)
        show_historical_statistics = (
            _normalize_snapshot_period(snapshot_period) == "history"
            and _show_historical_statistics(show_historical_statistics_values)
        )

        if not selected_run_fors:
            return _build_empty_state(
                "Choose Reporting Cycle value",
                "Select one Reporting Cycle value to render the SAAS workbook charts.",
            )

        if not effective_models:
            return _build_empty_state(
                "No models match the current filters",
                "Adjust Segment or Specific Models to bring one or more SAAS models into scope.",
            )

        time_series_df = SAAS_PAGE_DATA.get("mev_time_series")
        if time_series_df is None or time_series_df.empty:
            return _build_empty_state(
                "No SAAS MEV data is available",
                "The workbook did not return any MEV time-series records for this page.",
            )

        filtered_df = time_series_df[time_series_df["Model Name"].isin(effective_models)]
        filtered_df = filtered_df[filtered_df["Run For"].isin(scoped_run_fors)]

        records = filtered_df.to_dict(orient="records")
        panels = []
        for panel_index, model_name in enumerate(effective_models, start=1):
            model_records = [row for row in records if row.get("Model Name") == model_name]
            panels.append(
                _build_model_panel(
                    panel_index,
                    model_name,
                    model_records,
                    selected_run_fors,
                    compare_against,
                    snapshot_period,
                    mev_label_mode,
                    reference_lines,
                    show_historical_statistics,
                )
            )

        return panels or _build_empty_state(
            "No MEV charts match the current filters",
            "Adjust the Reporting Cycle, Segment, or Specific Models filters to broaden the SAAS workbook selection.",
        )

    @app.callback(
        Output({"type": layout.MODEL_MEV_GRID_TYPE, "model": MATCH}, "children"),
        Output({"type": layout.MODEL_DATE_RANGE_CONTROLS_TYPE, "model": MATCH}, "children"),
        Input({"type": layout.MODEL_MEV_TYPE_FILTER_TYPE, "model": MATCH}, "value"),
        Input({"type": layout.MODEL_SCENARIO_FILTER_TYPE, "model": MATCH}, "value"),
        Input({"type": layout.MODEL_MEV_FILTER_TYPE, "model": MATCH}, "value"),
        Input({"type": layout.MODEL_MEV_SINGLE_VALUE_TYPE, "model": MATCH}, "value"),
        Input({"type": layout.MODEL_DATE_RANGE_WINDOW_TYPE, "model": MATCH}, "value"),
        Input({"type": layout.MODEL_DATE_RANGE_FROM_TYPE, "model": MATCH}, "value"),
        Input({"type": layout.MODEL_DATE_RANGE_TO_TYPE, "model": MATCH}, "value"),
        State(layout.APPLIED_FILTERS_STORE_ID, "data"),
        prevent_initial_call=True,
    )
    def update_model_mev_chart_controls(selected_mev_mode, selected_scenario, selected_mevs_multi, selected_mev_single, window_value, from_value, to_value, applied):
        model_name = ctx.triggered_id["model"]
        time_series_df = SAAS_PAGE_DATA.get("mev_time_series")
        if time_series_df is None or time_series_df.empty:
            return no_update, no_update

        applied = applied or {}
        run_for = applied.get("run_for")
        compare_against = applied.get("compare_against")
        snapshot_period = applied.get("snapshot_period")
        reference_lines = applied.get("reference_lines")
        mev_label_mode = applied.get("mev_label_mode")
        show_historical_statistics_values = applied.get("historical_stats")

        selected_run_fors = _normalize_selected_run_fors(run_for)
        scoped_run_fors = _scoped_run_for_values(run_for, compare_against)
        snapshot_period_value = _normalize_snapshot_period(snapshot_period)
        show_historical_statistics = (
            snapshot_period_value == "history"
            and _show_historical_statistics(show_historical_statistics_values)
        )
        filtered_df = time_series_df[time_series_df["Model Name"] == model_name]
        if scoped_run_fors:
            filtered_df = filtered_df[filtered_df["Run For"].isin(scoped_run_fors)]
        else:
            filtered_df = filtered_df.iloc[0:0]

        base_records = _filter_records_by_snapshot_period(
            filtered_df.to_dict(orient="records"),
            snapshot_period_value,
        )
        periods = _available_date_periods(base_records)
        range_value = _resolve_date_range_selection(periods, window_value, from_value, to_value, ctx.triggered_id)
        range_controls = [
            layout.build_model_date_range_controls(
                model_name,
                periods,
                range_value,
                disabled=snapshot_period_value == "projection",
            )
        ]
        records = _filter_records_by_date_range(base_records, range_value)
        active_selected_mevs = _active_selected_mevs(
            model_name,
            selected_mev_mode,
            selected_mev_single,
            selected_mevs_multi,
            base_records,
        )
        mev_names = sorted({str(row.get("MEV Name") or "").strip() for row in records if str(row.get("MEV Name") or "").strip()})
        scenario_names = sorted({str(row.get("Scenario") or "").strip().lower() for row in records if str(row.get("Scenario") or "").strip()})
        date_values = sorted({row.get("Date") for row in records if row.get("Date") is not None})

        meta_parts = [f"Reporting Cycle: {_run_for_meta_label(selected_run_fors)}"]
        snapshot_label = next(
            (option["label"] for option in layout.SUBNAV_VIEW_OPTIONS if option["value"] == snapshot_period_value),
            None,
        )
        if snapshot_label:
            meta_parts.append(snapshot_label)
        if mev_names:
            meta_parts.append(_pluralize(len(mev_names), "MEV"))
        if scenario_names:
            meta_parts.append(_pluralize(len(scenario_names), "scenario"))
        if date_values:
            meta_parts.append(f"{_format_month_year(date_values[0])} to {_format_month_year(date_values[-1])}")

        return (
            _build_model_chart_cards(
                model_name,
                records,
                filtered_df.to_dict(orient="records"),
                selected_mev_mode,
                selected_scenario,
                snapshot_period_value,
                mev_label_mode,
                range_value,
                reference_lines,
                active_selected_mevs,
                meta_parts,
                selected_run_fors[0] if selected_run_fors else None,
                show_historical_statistics,
            ),
            range_controls,
        )
