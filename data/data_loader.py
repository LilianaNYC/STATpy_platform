"""Source-data loading for the PD Performance Dash app.

This is a trimmed port of ``data_manager.py`` and the data-layer helpers in
``callbacks/monitoring_pd_models_callbacks.py`` from the original monitoring
dashboard, keeping only what the PD Performance tab needs:

- the portfolio extract (with ``_quarter`` / ``_snapshot_date`` derived columns)
- the PD/CRR/RAG-assignment threshold tables
- row-level PD performance observations (1y / 2y / NCO 1y)
- worst-grade rating-migration observations
- the dummy MEV catalog and rank-ordering facility data

Functions ``_build_model_rows``, ``_build_model_quarter_breakdown``,
``_build_model_segment_quarter_breakdown`` and ``_build_threshold_summary``
from the original ``build_monitoring_pd_models`` are intentionally NOT
ported -- they feed the model-overview tables on other tabs and are not
referenced by ``renderPdModels()``'s live PD Performance sections. The model
and segment filter option lists are instead derived directly from the
portfolio frame.
"""

from __future__ import annotations

import json
import logging
import math
from datetime import date, datetime
from pathlib import Path
from typing import Any

import polars as pl
from openpyxl import load_workbook

from .. import monitoring_config as config

log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Small shared helpers
# ---------------------------------------------------------------------------
def _is_missing(value: Any) -> bool:
    return value is None or (isinstance(value, float) and math.isnan(value))


def _to_float(value: Any) -> float | None:
    if _is_missing(value):
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return None if math.isnan(result) else result


def _label(value: date | datetime) -> str:
    """Convert a quarter-end timestamp to a label like '2025Q1'."""
    q = (value.month - 1) // 3 + 1
    return f"{value.year}Q{q}"


def _normalize_model_name(value: Any) -> str:
    if _is_missing(value):
        return ""
    return str(value).strip()


def _format_rating_value(value: Any) -> str:
    if _is_missing(value):
        return ""
    if isinstance(value, (int, float)) and float(value).is_integer():
        return str(int(value))
    return str(value).strip()


def _clean_null_sentinels(df: pl.DataFrame) -> pl.DataFrame:
    return df.with_columns(
        [
            pl.when(pl.col(column).cast(pl.String).is_in(config.NULL_SENTINELS))
            .then(None)
            .otherwise(pl.col(column))
            .alias(column)
            for column in df.columns
        ]
    )


def _read_excel_sheet(path: Path, sheet_name: str) -> pl.DataFrame:
    """Read an Excel sheet into Polars while preserving openpyxl cell values."""
    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        return pl.DataFrame()

    headers = [str(value).strip() if value is not None else f"column_{index}" for index, value in enumerate(rows[0])]
    body = [row for row in rows[1:] if any(value is not None for value in row)]
    if not body:
        return pl.DataFrame(schema=headers)

    return pl.DataFrame(body, schema=headers, orient="row", infer_schema_length=None)


# ---------------------------------------------------------------------------
# Portfolio loading
# ---------------------------------------------------------------------------
def load_portfolio() -> pl.DataFrame:
    """Load the portfolio Excel file, clean null sentinels, derive quarter labels."""
    log.info("Loading portfolio from %s [%s]", config.PORTFOLIO_FILE, config.PORTFOLIO_SHEET_NAME)

    df = _read_excel_sheet(config.PORTFOLIO_FILE, config.PORTFOLIO_SHEET_NAME)
    df = _clean_null_sentinels(df)
    df = df.with_columns(
        pl.col(config.DATE_COLUMN).cast(pl.Date, strict=False).alias(config.DATE_COLUMN)
    )
    df = df.with_columns(
        [
            pl.col(config.DATE_COLUMN)
            .map_elements(_label, return_dtype=pl.String)
            .alias("_quarter"),
            pl.col(config.DATE_COLUMN).alias("_snapshot_date"),
        ]
    )

    log.info("Loaded %d records across %d quarters", df.height, df["_quarter"].n_unique())
    return df


def get_quarters(df: pl.DataFrame) -> list[str]:
    """Return sorted list of all available quarter labels."""
    dates = df.select(pl.col(config.DATE_COLUMN).sort().unique()).to_series().to_list()
    labels = [_label(value) for value in dates if value is not None]
    return sorted(set(labels))


def get_quarter_df(df: pl.DataFrame, quarter: str) -> pl.DataFrame:
    """Return rows for a specific quarter label."""
    return df.filter(pl.col("_quarter") == quarter)


def get_snapshot_date(df: pl.DataFrame, quarter: str) -> str:
    """Return the snapshot date string for a quarter."""
    rows = df.filter(pl.col("_quarter") == quarter)
    if rows.is_empty():
        return ""
    return str(rows["_snapshot_date"][0])


def get_model_names(df: pl.DataFrame) -> list[str]:
    values = df.select(config.PD_MODEL_COLUMN).to_series().to_list()
    return sorted({_normalize_model_name(value) for value in values if _normalize_model_name(value)})


def get_segment_values(df: pl.DataFrame) -> list[str]:
    values = df.select(config.SEGMENT_COLUMN).to_series().to_list()
    clean = {_normalize_model_name(value) for value in values}
    return sorted({value for value in clean if value and value.lower() != "nan"})


# ---------------------------------------------------------------------------
# Monitoring thresholds
# ---------------------------------------------------------------------------
def _records(df: pl.DataFrame) -> list[dict[str, Any]]:
    """Convert a DataFrame to a list of JSON-safe records."""
    if df.is_empty():
        return []
    return df.to_dicts()


def load_monitoring_thresholds() -> dict[str, list[dict[str, Any]]]:
    """Load the PD / CRR-master-scale / RAG-assignment sheets.

    The original ``load_monitoring_thresholds`` only loads sheets whose
    config key ends with ``_thresholds_sheet_name`` or equals
    ``crr_master_scale_sheet_name`` -- which means ``RAG_Assignment_PD`` is
    never loaded by the existing pipeline. We load all three sheets the PD
    Performance tab needs explicitly here.
    """
    thresholds: dict[str, list[dict[str, Any]]] = {}

    for key, sheet_name in (
        ("pd_thresholds", config.PD_THRESHOLDS_SHEET_NAME),
        ("crr_master_scale", config.CRR_MASTER_SCALE_SHEET_NAME),
        ("rag_assignment_pd", config.RAG_ASSIGNMENT_PD_SHEET_NAME),
        ("lgd_thresholds", config.LGD_THRESHOLDS_SHEET_NAME),
        ("loss_thresholds", config.LOSS_THRESHOLDS_SHEET_NAME),
    ):
        try:
            df = _read_excel_sheet(config.MONITORING_THRESHOLDS_FILE, sheet_name)
            thresholds[key] = _records(df)
            log.info("Loaded thresholds sheet '%s' as %s (%d rows)", sheet_name, key, df.height)
        except Exception as exc:  # noqa: BLE001 - mirror original best-effort loading
            log.warning("Unable to load sheet '%s' for %s: %s", sheet_name, key, exc)
            thresholds[key] = []

    return thresholds


# ---------------------------------------------------------------------------
# PD performance observations (1y / 2y / NCO 1y)
# ---------------------------------------------------------------------------
def build_pd_performance_observations(df: pl.DataFrame) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    """Return the compact row-level payload needed for interactive PD metrics.

    Port of ``_build_pd_performance_observations``.
    """
    model_column = config.PD_MODEL_COLUMN
    segment_column = config.SEGMENT_COLUMN
    rating_column = config.RATING_COLUMN
    horizon_columns = config.PD_HORIZON_COLUMNS

    required = [model_column, segment_column, "_quarter"]
    missing = [column for column in required if column not in df.columns]
    if missing:
        log.warning("PD performance columns not found: %s", missing)
        return horizon_columns, []

    available_horizons = {}
    for horizon, columns in horizon_columns.items():
        observed_column = columns["observed_column"]
        if observed_column not in df.columns:
            log.warning("PD %s observed-default column not found: %s", horizon, observed_column)
            continue
        predicted_column = columns["predicted_column"]
        if predicted_column not in df.columns:
            log.warning("PD %s performance column not found: %s", horizon, predicted_column)
            continue
        available_horizons[horizon] = columns

    payload_columns = list(dict.fromkeys(required + (
        [rating_column] if rating_column in df.columns else []
    ) + [
        column
        for columns in available_horizons.values()
        for column in [columns["observed_column"], columns["predicted_column"], columns.get("ead_column")]
        if column
    ]))
    observations = df.select(payload_columns).with_columns(
        [
            pl.col(model_column).map_elements(_normalize_model_name, return_dtype=pl.String),
            pl.col(segment_column).map_elements(_normalize_model_name, return_dtype=pl.String),
        ]
    )
    observations = observations.filter(
        pl.col("_quarter").is_not_null()
        & (pl.col(model_column) != "")
        & (pl.col(segment_column) != "")
    )

    numeric_columns = list(dict.fromkeys([
        column
        for columns in available_horizons.values()
        for column in [columns["observed_column"], columns["predicted_column"], columns.get("ead_column")]
        if column and column in observations.columns
    ]))
    if numeric_columns:
        observations = observations.with_columns(
            [pl.col(column).cast(pl.Float64, strict=False).alias(column) for column in numeric_columns]
        )

    result = []
    for row in observations.iter_rows(named=True):
        horizons = {}
        for horizon, columns in available_horizons.items():
            observed = row[columns["observed_column"]]
            if _is_missing(observed) or observed not in (0, 1):
                continue
            predicted = row[columns["predicted_column"]]
            if not _is_missing(predicted) and 0 <= predicted <= 1:
                ead_column = columns.get("ead_column")
                ead = row[ead_column] if ead_column and ead_column in observations.columns else None
                horizons[horizon] = {
                    "observed": int(observed),
                    "predicted": round(float(predicted), 8),
                    "ead": round(float(ead), 2) if not _is_missing(ead) and ead >= 0 else None,
                }
        if horizons:
            result.append(
                {
                    "quarter": row["_quarter"],
                    "model": row[model_column],
                    "segment": row[segment_column],
                    "rating": _format_rating_value(row[rating_column]) if rating_column in observations.columns else "",
                    "horizons": horizons,
                }
            )

    return horizon_columns, result


# ---------------------------------------------------------------------------
# Rating migration (worst-grade per facility per quarter)
# ---------------------------------------------------------------------------
def build_rating_migration_observations(df: pl.DataFrame) -> tuple[list[str], list[dict[str, Any]]]:
    """Return one worst-grade rating row per facility and quarter.

    Port of ``_build_rating_migration_observations``.
    """
    model_column = config.PD_MODEL_COLUMN
    segment_column = config.SEGMENT_COLUMN
    id_column = config.FACILITY_ID_COLUMN
    rating_column = config.RATING_COLUMN

    required = [id_column, model_column, segment_column, "_quarter", rating_column]
    missing = [column for column in required if column not in df.columns]
    if missing:
        log.warning("Rating migration columns not found: %s", missing)
        return [], []

    observations = df.select(required).with_columns(
        [
            pl.col(id_column).map_elements(_normalize_model_name, return_dtype=pl.String),
            pl.col(model_column).map_elements(_normalize_model_name, return_dtype=pl.String),
            pl.col(segment_column).map_elements(_normalize_model_name, return_dtype=pl.String),
            pl.col(rating_column).cast(pl.Float64, strict=False).alias("_rating_numeric"),
        ]
    )
    observations = observations.filter(
        pl.col("_quarter").is_not_null()
        & pl.col("_rating_numeric").is_not_null()
        & (pl.col(id_column) != "")
        & (pl.col(model_column) != "")
        & (pl.col(segment_column) != "")
    )

    observations = observations.sort("_rating_numeric").unique(
        subset=["_quarter", id_column],
        keep="last",
        maintain_order=True,
    )

    ratings = sorted(observations["_rating_numeric"].unique().to_list())

    def rating_label(value: float) -> str:
        return str(int(value)) if float(value).is_integer() else str(value)

    return (
        [rating_label(value) for value in ratings],
        [
            {
                "quarter": row["_quarter"],
                "account": row[id_column],
                "model": row[model_column],
                "segment": row[segment_column],
                "rating": rating_label(row["_rating_numeric"]),
            }
            for row in observations.iter_rows(named=True)
        ],
    )


# ---------------------------------------------------------------------------
# MEV catalog
# ---------------------------------------------------------------------------
def load_pd_mev_catalog() -> dict[str, Any]:
    """Load the dummy MEV time-series catalog used by the PD MEV range section.

    Port of ``_load_pd_mev_catalog``.
    """
    path = config.MEV_DUMMY_DATA_FILE
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError:
        log.warning("PD MEV dummy data file not found: %s", path)
        return {}
    except json.JSONDecodeError as exc:
        log.warning("PD MEV dummy data file could not be parsed: %s", exc)
        return {}

    if not isinstance(payload, dict):
        log.warning("PD MEV dummy data must be a model-keyed JSON object.")
        return {}

    catalog: dict[str, Any] = {}
    for raw_model_name, raw_model_payload in payload.items():
        if not isinstance(raw_model_payload, dict):
            continue

        model_name = _normalize_model_name(raw_model_name)
        if not model_name:
            continue

        raw_mevs = raw_model_payload.get("mevs", {})
        if not isinstance(raw_mevs, dict):
            raw_mevs = {}

        mevs: dict[str, Any] = {}
        for raw_mev_name, raw_mev_payload in raw_mevs.items():
            if not isinstance(raw_mev_payload, dict):
                continue

            dev_range = raw_mev_payload.get("dev_range", {})
            time_series = raw_mev_payload.get("time_series", {})
            if not isinstance(dev_range, dict):
                dev_range = {}
            if not isinstance(time_series, dict):
                time_series = {}

            clean_range = {
                "min": _to_float(dev_range.get("min")),
                "max": _to_float(dev_range.get("max")),
                "mean": _to_float(dev_range.get("mean")),
                "2std_lower": _to_float(dev_range.get("2std_lower")),
                "2std_upper": _to_float(dev_range.get("2std_upper")),
                "development_date": str(dev_range.get("development_date") or ""),
            }
            clean_range = {
                key: (round(float(value), 6) if value is not None else None)
                for key, value in clean_range.items()
                if key != "development_date"
            } | {"development_date": clean_range["development_date"]}

            clean_series = {
                str(period): round(float(value), 6)
                for period, value in time_series.items()
                if _to_float(value) is not None
            }
            mevs[str(raw_mev_name).strip()] = {
                "dev_range": clean_range,
                "time_series": clean_series,
            }

        catalog[model_name] = {
            "segments": [
                str(segment).strip()
                for segment in raw_model_payload.get("segments", [])
                if str(segment).strip()
            ],
            "severe_scenario_date": str(raw_model_payload.get("severe_scenario_date") or ""),
            "mevs": mevs,
        }

    return catalog


# ---------------------------------------------------------------------------
# Rank-ordering facilities
# ---------------------------------------------------------------------------
def _clean_rank_ordering_series(raw_series: Any) -> dict[str, float]:
    if not isinstance(raw_series, dict):
        return {}
    clean: dict[str, float] = {}
    for raw_period, raw_value in raw_series.items():
        period = str(raw_period).strip()
        value = _to_float(raw_value)
        if period and value is not None:
            clean[period] = round(float(value), 6)
    return clean


def _clean_rank_ordering_forecast(raw_forecast: Any) -> dict[str, dict[str, float]]:
    if not isinstance(raw_forecast, dict):
        raw_forecast = {}
    return {
        "base": _clean_rank_ordering_series(raw_forecast.get("base", {})),
        "severe": _clean_rank_ordering_series(raw_forecast.get("severe", {})),
    }


def load_pd_rank_ordering_facilities() -> dict[str, Any]:
    """Load the dummy facility-level PD paths used by the scenario rank-ordering section.

    Port of ``_load_pd_rank_ordering_facilities``.
    """
    path = config.FACILITIES_DUMMY_DATA_FILE
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError:
        log.warning("PD rank-ordering dummy data file not found: %s", path)
        return {}
    except json.JSONDecodeError as exc:
        log.warning("PD rank-ordering dummy data file could not be parsed: %s", exc)
        return {}

    if not isinstance(payload, dict):
        log.warning("PD rank-ordering dummy data must be a facility-keyed JSON object.")
        return {}

    facilities: dict[str, Any] = {}
    for raw_facility_id, raw_payload in payload.items():
        if not isinstance(raw_payload, dict):
            continue

        facility_id = str(raw_facility_id).strip()
        if not facility_id:
            continue

        facilities[facility_id] = {
            "segment": str(raw_payload.get("segment") or "").strip(),
            "pd_model": _normalize_model_name(raw_payload.get("pd_model")),
            "severe_scenario_date": str(raw_payload.get("severe_scenario_date") or ""),
            "acl_pd_historical": _clean_rank_ordering_series(raw_payload.get("acl_pd_historical", {})),
            "nco_pd_historical": _clean_rank_ordering_series(raw_payload.get("nco_pd_historical", {})),
            "acl_pd_forecast": _clean_rank_ordering_forecast(raw_payload.get("acl_pd_forecast", {})),
            "nco_pd_forecast": _clean_rank_ordering_forecast(raw_payload.get("nco_pd_forecast", {})),
        }

    return facilities


# ---------------------------------------------------------------------------
# Top-level loader
# ---------------------------------------------------------------------------
def load_pd_performance_data() -> dict[str, Any]:
    """Load and assemble everything the PD Performance tab needs."""
    portfolio = load_portfolio()
    quarters = get_quarters(portfolio)
    latest_quarter = quarters[-1] if quarters else ""
    previous_quarter = quarters[-2] if len(quarters) > 1 else ""
    monitoring_thresholds = load_monitoring_thresholds()
    performance_horizons, performance_observations = build_pd_performance_observations(portfolio)
    rating_values, rating_migration_observations = build_rating_migration_observations(portfolio)

    return {
        "portfolio": portfolio,
        "quarters": quarters,
        "latest_quarter": latest_quarter,
        "previous_quarter": previous_quarter,
        "latest_snapshot_date": get_snapshot_date(portfolio, latest_quarter),
        "previous_snapshot_date": get_snapshot_date(portfolio, previous_quarter) if previous_quarter else "",
        "source_file": config.PORTFOLIO_FILE.name,
        "model_names": get_model_names(portfolio),
        "segment_values": get_segment_values(portfolio),
        "monitoring_thresholds": monitoring_thresholds,
        "performance_horizons": performance_horizons,
        "performance_observations": performance_observations,
        "rating_values": rating_values,
        "rating_migration_observations": rating_migration_observations,
        "mev_catalog": load_pd_mev_catalog(),
        "rank_ordering_facilities": load_pd_rank_ordering_facilities(),
    }
