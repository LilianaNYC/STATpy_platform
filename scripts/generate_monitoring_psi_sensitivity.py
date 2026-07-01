"""Generate the LGD/EAD PSI + Sensitivity data the dashboards need.

The PD Performance tab has a "Post Subjective Review Analysis" chapter with PSI,
Scenario Ranking and Sensitivity Analysis sub-sections. The LGD and EAD tabs are
gaining the same sub-sections, but the underlying data only existed for PD:

  * ``population_stability_index`` lived only on ``PD_Performance_Metrics``.
  * Only ``PD_Sensitivity_Projections`` existed (no LGD/EAD equivalent).
  * ``LGD_Thresholds`` / ``EAD_Thresholds`` had no PSI RAG row.
    (``Scenario_Test_Thresholds`` already carried LGD/EAD rows.)

This script writes that missing information into the committed source workbooks,
mirroring the PD structure so the new sub-sections behave exactly like PD's:

  portfolio.xlsx
    * adds ``population_stability_index`` to LGD_/EAD_Performance_Metrics
      (reusing PD's horizon-agnostic PSI for matching cycle/level/entity/quarter,
      with a deterministic synthetic fallback).
    * creates LGD_Sensitivity_Projections / EAD_Sensitivity_Projections with the
      same schema/x-axis as PD_Sensitivity_Projections (``projected_lgd`` /
      ``projected_ead`` instead of ``projected_pd``), with scenario paths ordered
      baseline < intsevere < baseline_2std_shock so the ranking test passes and
      the sensitivity impact straddles the 0.15 scenario-test threshold.

  statpy_monitoring_thresholds.xlsm
    * adds a "Population Stability Index" RAG row to LGD_Thresholds / EAD_Thresholds
      (same rule as PD: green <= 0.10, amber <= 0.25, red > 0.25, lower-is-better).

Run from the directory that contains the ``STATpy_platform`` package:

    python -m STATpy_platform.scripts.generate_monitoring_psi_sensitivity

It is idempotent: re-running overwrites the generated column/sheets/rows rather
than appending duplicates.
"""

from __future__ import annotations

import hashlib
from collections import defaultdict
from pathlib import Path

import openpyxl

SOURCE_DATA_DIR = Path(__file__).resolve().parent.parent / "source_data"
PORTFOLIO_FILE = SOURCE_DATA_DIR / "portfolio.xlsx"
THRESHOLDS_FILE = SOURCE_DATA_DIR / "statpy_monitoring_thresholds.xlsm"

PSI_COLUMN = "population_stability_index"
SCENARIOS = ("baseline", "intsevere", "baseline_2std_shock")

# Scenario-test threshold the Sensitivity Analysis RAGs against (Scenario_Test_Thresholds).
SENSITIVITY_THRESHOLD = 0.15

TAB_CONFIG = {
    "LGD": {
        "metrics_sheet": "LGD_Performance_Metrics",
        "predicted_col": "predicted_lgd",
        "sensitivity_sheet": "LGD_Sensitivity_Projections",
        "projected_col": "projected_lgd",
        "fallback_base": 0.38,
        "clamp": (0.01, 0.99),
        "thresholds_sheet": "LGD_Thresholds",
    },
    "EAD": {
        "metrics_sheet": "EAD_Performance_Metrics",
        "predicted_col": "predicted_ead",
        "sensitivity_sheet": "EAD_Sensitivity_Projections",
        "projected_col": "projected_ead",
        "fallback_base": 0.78,
        "clamp": (0.05, 1.50),
        "thresholds_sheet": "EAD_Thresholds",
    },
}


# ---------------------------------------------------------------------------
# Small worksheet helpers
# ---------------------------------------------------------------------------
def _header(ws) -> list[str]:
    return [str(c) if c is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]


def _records(ws) -> list[dict]:
    header = _header(ws)
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(c is None for c in row):
            continue
        out.append({header[i]: row[i] for i in range(len(header))})
    return out


def _clamp(value: float, lo: float, hi: float) -> float:
    return max(lo, min(hi, value))


def _synthetic_psi(*key_parts) -> float:
    """Deterministic small PSI in [0.02, 0.22] from a key (stable across runs)."""
    digest = hashlib.md5("|".join(str(p) for p in key_parts).encode()).hexdigest()
    return round(0.02 + (int(digest[:8], 16) % 2000) / 10000.0, 6)


# ---------------------------------------------------------------------------
# PSI: add population_stability_index to LGD/EAD metrics
# ---------------------------------------------------------------------------
def _build_pd_psi_lookup(wb) -> dict:
    """Horizon-agnostic PSI keyed by (cycle, level, entity, quarter)."""
    ws = wb["PD_Performance_Metrics"]
    lookup: dict[tuple, float] = {}
    for rec in _records(ws):
        psi = rec.get(PSI_COLUMN)
        if psi is None:
            continue
        key = (rec.get("reporting_cycle"), rec.get("level"), rec.get("model_or_segment"), rec.get("quarter"))
        lookup.setdefault(key, float(psi))
    return lookup


def add_psi_column(wb, cfg: dict, pd_psi: dict) -> int:
    ws = wb[cfg["metrics_sheet"]]
    header = _header(ws)
    if PSI_COLUMN in header:
        col_idx = header.index(PSI_COLUMN) + 1
    else:
        col_idx = len(header) + 1
        ws.cell(row=1, column=col_idx, value=PSI_COLUMN)
    name_idx = {name: i for i, name in enumerate(header)}
    written = 0
    for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(c is None for c in row):
            continue
        cycle = row[name_idx["reporting_cycle"]]
        level = row[name_idx["level"]]
        entity = row[name_idx["model_or_segment"]]
        quarter = row[name_idx["quarter"]]
        value = pd_psi.get((cycle, level, entity, quarter))
        if value is None:
            value = _synthetic_psi(cycle, level, entity, quarter)
        ws.cell(row=r, column=col_idx, value=round(float(value), 6))
        written += 1
    return written


# ---------------------------------------------------------------------------
# Sensitivity: create LGD_/EAD_Sensitivity_Projections from PD's x-axis
# ---------------------------------------------------------------------------
def _entity_base_values(wb, cfg: dict) -> tuple[dict, float]:
    """Mean predicted value per (cycle, level, entity); plus a global fallback."""
    ws = wb[cfg["metrics_sheet"]]
    sums: dict[tuple, list[float]] = defaultdict(list)
    all_values: list[float] = []
    for rec in _records(ws):
        val = rec.get(cfg["predicted_col"])
        if val is None:
            continue
        key = (rec.get("reporting_cycle"), rec.get("level"), rec.get("model_or_segment"))
        sums[key].append(float(val))
        all_values.append(float(val))
    base = {key: sum(vals) / len(vals) for key, vals in sums.items()}
    global_base = (sum(all_values) / len(all_values)) if all_values else cfg["fallback_base"]
    return base, global_base


def build_sensitivity_sheet(wb, cfg: dict) -> int:
    """Mirror PD_Sensitivity_Projections x-axis with tab-appropriate projected values."""
    pd_ws = wb["PD_Sensitivity_Projections"]
    pd_header = _header(pd_ws)
    out_header = [(cfg["projected_col"] if h == "projected_pd" else h) for h in pd_header]

    base_by_entity, global_base = _entity_base_values(wb, cfg)
    lo, hi = cfg["clamp"]

    # Unique (cycle, level, entity) -> ordered list of (quarter, projection_quarter, MM_P0, MM_Pm)
    axis: dict[tuple, dict] = {}
    for rec in _records(pd_ws):
        key = (rec.get("reporting_cycle"), rec.get("level"), rec.get("model_or_segment"))
        q = rec.get("quarter")
        axis.setdefault(key, {})[q] = (
            rec.get("projection_quarter"),
            rec.get("MM_P0"),
            rec.get("MM_Pm"),
        )

    rows: list[list] = []
    for (cycle, level, entity), quarters in axis.items():
        base = base_by_entity.get((cycle, level, entity))
        if base is None:
            base = base_by_entity.get((cycle, level, "All Models"), global_base)
        for q in sorted(quarters, key=lambda v: (v is None, v)):
            projection_quarter, mm_p0, mm_pm = quarters[q]
            q_int = int(q) if q is not None else 0
            # Mild baseline drift; intsevere and 2std-shock ordered above it. The
            # 2SD relative impact ramps through the 0.15 scenario-test threshold.
            baseline_path = base * (1.0 + 0.004 * q_int)
            shock_impact = 0.10 + 0.012 * q_int
            scenario_values = {
                "baseline": baseline_path,
                "intsevere": baseline_path * (1.0 + 0.5 * shock_impact),
                "baseline_2std_shock": baseline_path * (1.0 + shock_impact),
            }
            for scenario in SCENARIOS:
                value = round(_clamp(scenario_values[scenario], lo, hi), 8)
                base_scenario = "baseline" if scenario == "baseline_2std_shock" else scenario
                shock_std = 2 if scenario == "baseline_2std_shock" else 0
                shock_direction = "adverse" if scenario == "baseline_2std_shock" else "none"
                record = {
                    "reporting_cycle": cycle,
                    "level": level,
                    "model_or_segment": entity,
                    "quarter": q,
                    "projection_quarter": projection_quarter,
                    "scenario_variant": scenario,
                    "base_scenario": base_scenario,
                    "shock_std": shock_std,
                    "shock_direction": shock_direction,
                    cfg["projected_col"]: value,
                    "MM_P0": mm_p0,
                    "MM_Pm": mm_pm,
                }
                rows.append([record[h] for h in out_header])

    # (Re)create the sheet
    if cfg["sensitivity_sheet"] in wb.sheetnames:
        del wb[cfg["sensitivity_sheet"]]
    ws = wb.create_sheet(cfg["sensitivity_sheet"])
    ws.append(out_header)
    for row in rows:
        ws.append(row)
    return len(rows)


# ---------------------------------------------------------------------------
# Thresholds: add a PSI RAG row to LGD/EAD threshold tables
# ---------------------------------------------------------------------------
def add_psi_threshold(wb, sheet_name: str) -> bool:
    ws = wb[sheet_name]
    header = _header(ws)
    # Skip if a PSI row already exists (idempotent)
    for rec in _records(ws):
        if str(rec.get("metric", "")).strip().lower() == "population stability index":
            return False
    psi_values = {
        "metric": "Population Stability Index",
        "dimension": "Performance",
        "green_rule": "value <= 0.10",
        "amber_rule": "0.10 < value <= 0.25",
        "red_rule": "value > 0.25",
        "green_min": None,
        "green_max": 0.10,
        "amber_min": 0.10,
        "amber_max": 0.25,
        "red_condition": "above amber_max",
        "higher_is_better": False,
        "lower_is_better": True,
        "target_value": None,
        "notes": "Lower is better; population stability on the key-driver distribution.",
    }
    # Append at the first fully-empty row so we don't leave gaps.
    target_row = ws.max_row + 1
    for r in range(2, ws.max_row + 1):
        if all(c.value is None for c in ws[r]):
            target_row = r
            break
    for i, col in enumerate(header, start=1):
        ws.cell(row=target_row, column=i, value=psi_values.get(col))
    return True


# ---------------------------------------------------------------------------
def main() -> None:
    print(f"portfolio:  {PORTFOLIO_FILE}")
    wb = openpyxl.load_workbook(PORTFOLIO_FILE)
    pd_psi = _build_pd_psi_lookup(wb)
    for tab, cfg in TAB_CONFIG.items():
        n_psi = add_psi_column(wb, cfg, pd_psi)
        n_sens = build_sensitivity_sheet(wb, cfg)
        print(f"  {tab}: PSI rows={n_psi}, {cfg['sensitivity_sheet']} rows={n_sens}")
    wb.save(PORTFOLIO_FILE)

    print(f"thresholds: {THRESHOLDS_FILE}")
    twb = openpyxl.load_workbook(THRESHOLDS_FILE, keep_vba=True)
    for tab, cfg in TAB_CONFIG.items():
        added = add_psi_threshold(twb, cfg["thresholds_sheet"])
        print(f"  {cfg['thresholds_sheet']}: PSI row {'added' if added else 'already present'}")
    twb.save(THRESHOLDS_FILE)
    print("done.")


if __name__ == "__main__":
    main()
